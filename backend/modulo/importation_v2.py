"""
Importação v2 — Página dinâmica e moldável.

Endpoints:
    POST   /api/importation-v2/calculate     — calcula consumo, SS, P.Rep, ruptura, desvios
    GET    /api/importation-v2/modelos       — lista modelos do usuário + padrão
    POST   /api/importation-v2/modelos       — cria modelo
    DELETE /api/importation-v2/modelos/{id}  — exclui modelo do usuário
    GET    /api/importation-v2/defaults      — devolve parâmetros default + lista de SKUs padrão

Diferenças da v1 (importation.py):
    - Sem `target_calc_months` hard-coded (BUG-01 corrigido)
    - Safety stock clássico: Z × σ_diario × √LT (BUG-02 corrigido)
    - Cobertura inclui pipeline (BUG-03 corrigido)
    - Lead time / nível serviço / pipeline editáveis por SKU (overrides)
    - Janela de análise default = 15 meses (fallback: usa o que tiver)
    - Cutoff ADS < 0.05 removido
    - Detecta meses outliers (±1.5σ default) → campo `desvio`

Permissões: module_id = 'importation_v2'
"""
from __future__ import annotations

import json
import math
import os
from calendar import monthrange
from datetime import date
from typing import Optional, List, Dict, Any

import numpy as np
import pandas as pd
from dateutil.relativedelta import relativedelta
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from pydantic import BaseModel, Field

from auth_utils import get_user_id_from_session
from db_utils import get_db_connection
from permission_utils import check_module_permission
from core.config import IMPORTED_ITEM_CODES, PROJECT_ID, FILE_PARAMETROS
from modulo.importation import get_bq_client, _load_sheet_parametros_importacao  # reaproveita conexão e parser de planilha
from core import dummy  # fundação dummy determinística (sem fontes externas)


def _load_moq_map_db() -> Dict[str, float]:
    """Lê MOQ da tabela importacao_v2_moq (fonte primária)."""
    out: Dict[str, float] = {}
    try:
        conn = get_db_connection(); cur = conn.cursor()
        cur.execute("SELECT codigo, moq FROM importacao_v2_moq WHERE moq > 0")
        for r in cur.fetchall():
            out[str(r[0]).strip()] = float(r[1])
        cur.close(); conn.close()
    except Exception as e:
        print(f"[importation_v2] _load_moq_map_db error: {e}")
    return out


def _load_moq_map() -> Dict[str, float]:
    """MOQ por SKU. Prioridade: tabela importacao_v2_moq → fallback UNIT/CTN do ParametrosImportacao.xlsx.

    Aceita variações de nome (MOQ, Moq, Lote_Min, UNIT/CTN, UNITCTN, etc.).
    """
    # 1) DB (gerenciado pelo usuário, fonte primária)
    db_map = _load_moq_map_db()
    # 2) Fallback do Excel (para SKUs ainda não cadastrados no DB)
    try:
        if not os.path.exists(FILE_PARAMETROS):
            print(f"[importation_v2] MOQ: arquivo {FILE_PARAMETROS} não encontrado")
            return {}
        df = pd.read_excel(FILE_PARAMETROS)
        cols_norm = {c: str(c).strip().lower().replace(' ', '').replace('_', '').replace('/', '') for c in df.columns}

        # 1) coluna de código
        col_cod = None
        for c, cn in cols_norm.items():
            if cn in ('codigo', 'codigoempresa', 'codprod', 'codproduto', 'codigoproduto'):
                col_cod = c; break
        if col_cod is None:
            print(f"[importation_v2] MOQ: coluna de código não encontrada. Disponíveis: {list(df.columns)}")
            return {}

        # 2) MOQ explícito (preferência)
        col_moq = None
        for c, cn in cols_norm.items():
            if cn in ('moq', 'lotemin', 'loteminimo', 'qtdmin', 'qtdminima', 'minorderquantity'):
                col_moq = c; break
        # 3) fallback: UNIT/CTN (caixa fechada)
        if col_moq is None:
            for c, cn in cols_norm.items():
                if cn in ('unitctn', 'unitsctn', 'unidadescaixa', 'undcaixa', 'pcsctn'):
                    col_moq = c; break

        if col_moq is None:
            print(f"[importation_v2] MOQ: nenhuma coluna MOQ ou UNIT/CTN encontrada. Disponíveis: {list(df.columns)}")
            return {}

        out: Dict[str, float] = {}
        for _, row in df.iterrows():
            raw = row[col_cod]
            if pd.isna(raw):
                continue
            cod = str(raw).strip()
            if cod.endswith('.0'):
                cod = cod[:-2]
            try:
                v = float(row[col_moq])
                if v > 0:
                    out[cod] = v
            except Exception:
                continue
        print(f"[importation_v2] MOQ Excel fallback de '{col_moq}': {len(out)} itens")
        # DB tem prioridade — sobrescreve fallback do Excel
        for cod, v in db_map.items():
            out[cod] = v
        return out
    except Exception as e:
        print(f"[importation_v2] _load_moq_map error: {e}")
        return db_map


def _load_container_params_db() -> Dict[str, Dict[str, Any]]:
    """Lê parâmetros (UNIT/CTN, CBM, G.W, etc.) da tabela importacao_v2_moq.
    Esta é a fonte primária — gerenciada pelo usuário na página /importacao-v2/moq."""
    out: Dict[str, Dict[str, Any]] = {}
    try:
        conn = get_db_connection(); cur = conn.cursor()
        cur.execute("""
            SELECT codigo, COALESCE(unit_ctn, 0), COALESCE(cbm, 0),
                   COALESCE(gw, 0), COALESCE(nw, 0),
                   COALESCE(comprimento, 0), COALESCE(largura, 0), COALESCE(altura, 0),
                   COALESCE(price, 0), COALESCE(ncm, ''), COALESCE(unit, ''),
                   COALESCE(barcode, ''), COALESCE(name_cn, ''),
                   COALESCE(remark, ''), COALESCE(obs, ''), COALESCE(observacoes, ''),
                   COALESCE(english_description, '')
            FROM importacao_v2_moq
        """)
        for r in cur.fetchall():
            cod = str(r[0]).strip()
            out[cod] = {
                'unit_ctn': float(r[1]) if r[1] else 0.0,
                'cbm': float(r[2]) if r[2] else 0.0,
                'gw': float(r[3]) if r[3] else 0.0,
                'peso_liquido': float(r[4]) if r[4] else 0.0,
                'l': float(r[5]) if r[5] else 0.0,
                'w': float(r[6]) if r[6] else 0.0,
                'h': float(r[7]) if r[7] else 0.0,
                'price': float(r[8]) if r[8] else 0.0,
                'ncm': str(r[9] or ''),
                'unit': str(r[10] or ''),
                'barcode': str(r[11] or ''),
                'name_cn': str(r[12] or ''),
                'remark': str(r[13] or ''),
                'obs': str(r[14] or ''),
                'observacoes': str(r[15] or ''),
                'english_description': str(r[16] or ''),
            }
        cur.close(); conn.close()
    except Exception as e:
        print(f"[importation_v2] _load_container_params_db: {e}")
    return out


def _load_container_params() -> Dict[str, Dict[str, float]]:
    """Carrega params dimensionais/preço por SKU.
    Prioridade: tabela importacao_v2_moq (primária) → fallback ParametrosImportacao.xlsx.
    """
    # 1) Banco (fonte primária — gerenciada pelo RH/Logística na página /moq)
    out: Dict[str, Dict[str, float]] = dict(_load_container_params_db())
    # 2) Fallback Excel para SKUs ainda não cadastrados no banco
    try:
        if not os.path.exists(FILE_PARAMETROS):
            return out
        df = pd.read_excel(FILE_PARAMETROS)
        # Normaliza nomes
        cols_norm = {c: str(c).strip().lower().replace(' ', '').replace('_', '').replace('/', '').replace('.', '') for c in df.columns}
        def find(opts):
            for c, cn in cols_norm.items():
                if cn in opts:
                    return c
            return None
        col_cod = find({'codigo', 'codigoempresa', 'codprod', 'codproduto', 'codigoproduto', 'itemno'})
        col_unit_ctn = find({'unitctn', 'unitsctn', 'unidadescaixa', 'pcsctn'})
        col_cbm = find({'cbm', 'volumecbm', 'cbmun'})
        col_price = find({'price', 'preco', 'unitprice', 'precounitario', 'upreco'})
        col_gw = find({'gw', 'gwkg', 'pesobruto', 'pesobrutoun', 'pesobrutokg'})
        col_pliq = find({'pesoliquido', 'pesoliquidoun', 'pesoliquidokg', 'plq', 'plqun'})
        col_l = find({'l', 'comprimento'})
        col_w = find({'w', 'largura'})
        col_h = find({'h', 'altura'})
        col_unit = find({'unit'})
        if not col_cod:
            return out
        for _, row in df.iterrows():
            raw = row[col_cod]
            if pd.isna(raw):
                continue
            cod = str(raw).strip()
            if cod.endswith('.0'): cod = cod[:-2]
            def num(col):
                if col is None: return 0.0
                try:
                    v = float(row[col])
                    return v if pd.notna(v) else 0.0
                except Exception:
                    return 0.0
            def text(col):
                if col is None: return ''
                v = row[col]
                if pd.isna(v): return ''
                return str(v).strip()
            # Não sobrescreve banco — só adiciona se não existir
            if cod not in out:
                out[cod] = {
                    'unit_ctn': num(col_unit_ctn),
                    'cbm': num(col_cbm),
                    'price': num(col_price),
                    'gw': num(col_gw),
                    'peso_liquido': num(col_pliq),
                    'l': num(col_l),
                    'w': num(col_w),
                    'h': num(col_h),
                    'unit': text(col_unit),
                    'ncm': '',
                }
    except Exception as e:
        print(f"[importation_v2] _load_container_params Excel fallback: {e}")
    return out


router = APIRouter(prefix="/importation-v2", tags=["Importacao V2"])

MODULE_ID = "importation_v2"


def ensure_importacao_v2_modelos_table():
    """Cria tabelas (modelos, versoes, moq) — cada uma em sua própria transação,
    para que falha em uma não impeça criação das demais (idempotente)."""

    def _run(label: str, sqls: list):
        conn = get_db_connection()
        cur = conn.cursor()
        try:
            for sql in sqls:
                cur.execute(sql)
            conn.commit()
        except Exception as e:
            conn.rollback()
            print(f"[importation_v2] ensure_table[{label}] error: {e}")
        finally:
            cur.close(); conn.close()

    _run("modelos", [
        """CREATE TABLE IF NOT EXISTS importacao_v2_modelos (
            id              SERIAL       PRIMARY KEY,
            user_id         INT          NOT NULL,
            nome            VARCHAR(120) NOT NULL,
            codigos         JSONB        NOT NULL,
            qtd_meses       SMALLINT     NOT NULL CHECK (qtd_meses BETWEEN 1 AND 36),
            modo            VARCHAR(10)  NOT NULL CHECK (modo IN ('corrido','vendas')),
            overrides       JSONB        NOT NULL DEFAULT '{}'::jsonb,
            threshold_sigma NUMERIC(4,2) NOT NULL DEFAULT 1.5,
            created_at      TIMESTAMP    NOT NULL DEFAULT NOW(),
            updated_at      TIMESTAMP    NOT NULL DEFAULT NOW(),
            UNIQUE(user_id, nome)
        )""",
        "CREATE INDEX IF NOT EXISTS idx_importacao_v2_modelos_user ON importacao_v2_modelos(user_id)",
    ])

    _run("versoes", [
        """CREATE TABLE IF NOT EXISTS importacao_v2_versoes (
            id          SERIAL       PRIMARY KEY,
            user_id     INT,
            user_nome   VARCHAR(160),
            nome        VARCHAR(180) NOT NULL,
            labels      JSONB        NOT NULL DEFAULT '[]'::jsonb,
            observacao  TEXT,
            parametros  JSONB        NOT NULL,
            resultado   JSONB        NOT NULL,
            created_at  TIMESTAMP    NOT NULL DEFAULT NOW()
        )""",
        "CREATE INDEX IF NOT EXISTS idx_importacao_v2_versoes_created ON importacao_v2_versoes(created_at DESC)",
        "CREATE INDEX IF NOT EXISTS idx_importacao_v2_versoes_user ON importacao_v2_versoes(user_id)",
    ])

    # MOQ + Medidas — sem FK pra users
    _run("moq", [
        """CREATE TABLE IF NOT EXISTS importacao_v2_moq (
            codigo       VARCHAR(40)   PRIMARY KEY,
            descricao    VARCHAR(180),
            moq          NUMERIC(12,2) NOT NULL,
            origem       VARCHAR(40)   DEFAULT 'manual',
            updated_by   INT,
            updated_at   TIMESTAMP     NOT NULL DEFAULT NOW()
        )""",
        # Colunas novas (medidas) — ALTER IF NOT EXISTS pra ser idempotente
        "ALTER TABLE importacao_v2_moq ADD COLUMN IF NOT EXISTS unit_ctn NUMERIC(12,2)",
        "ALTER TABLE importacao_v2_moq ADD COLUMN IF NOT EXISTS cbm NUMERIC(12,6)",
        "ALTER TABLE importacao_v2_moq ADD COLUMN IF NOT EXISTS gw NUMERIC(12,3)",
        "ALTER TABLE importacao_v2_moq ADD COLUMN IF NOT EXISTS nw NUMERIC(12,3)",
        "ALTER TABLE importacao_v2_moq ADD COLUMN IF NOT EXISTS comprimento NUMERIC(10,2)",
        "ALTER TABLE importacao_v2_moq ADD COLUMN IF NOT EXISTS largura NUMERIC(10,2)",
        "ALTER TABLE importacao_v2_moq ADD COLUMN IF NOT EXISTS altura NUMERIC(10,2)",
        "ALTER TABLE importacao_v2_moq ADD COLUMN IF NOT EXISTS price NUMERIC(12,4)",
        "ALTER TABLE importacao_v2_moq ADD COLUMN IF NOT EXISTS ncm VARCHAR(40)",
        "ALTER TABLE importacao_v2_moq ADD COLUMN IF NOT EXISTS unit VARCHAR(20)",
        # Colunas extras da planilha completa (moq.xlsx)
        "ALTER TABLE importacao_v2_moq ADD COLUMN IF NOT EXISTS barcode VARCHAR(60)",
        "ALTER TABLE importacao_v2_moq ADD COLUMN IF NOT EXISTS name_cn VARCHAR(255)",
        "ALTER TABLE importacao_v2_moq ADD COLUMN IF NOT EXISTS remark TEXT",
        "ALTER TABLE importacao_v2_moq ADD COLUMN IF NOT EXISTS obs TEXT",
        "ALTER TABLE importacao_v2_moq ADD COLUMN IF NOT EXISTS observacoes TEXT",
        "ALTER TABLE importacao_v2_moq ADD COLUMN IF NOT EXISTS english_description VARCHAR(255)",
        "ALTER TABLE importacao_v2_moq ADD COLUMN IF NOT EXISTS ctns NUMERIC(12,2)",
        "ALTER TABLE importacao_v2_moq ADD COLUMN IF NOT EXISTS qty NUMERIC(12,2)",
        "ALTER TABLE importacao_v2_moq ADD COLUMN IF NOT EXISTS amount NUMERIC(14,2)",
        "ALTER TABLE importacao_v2_moq ADD COLUMN IF NOT EXISTS cbm_total NUMERIC(12,4)",
        "ALTER TABLE importacao_v2_moq ADD COLUMN IF NOT EXISTS tgw NUMERIC(12,3)",
        "ALTER TABLE importacao_v2_moq ADD COLUMN IF NOT EXISTS tnw NUMERIC(12,3)",
    ])

    _run("order_lists", [
        """CREATE TABLE IF NOT EXISTS importacao_v2_order_lists (
            id            SERIAL       PRIMARY KEY,
            user_id       INT,
            user_nome     VARCHAR(160),
            nome          VARCHAR(180) NOT NULL,
            labels        JSONB        NOT NULL DEFAULT '[]'::jsonb,
            observacao    TEXT,
            items         JSONB        NOT NULL,
            datas_chegada JSONB        NOT NULL DEFAULT '{}'::jsonb,
            created_at    TIMESTAMP    NOT NULL DEFAULT NOW()
        )""",
        "CREATE INDEX IF NOT EXISTS idx_importacao_v2_order_lists_created ON importacao_v2_order_lists(created_at DESC)",
    ])

    _run("container_modelos", [
        """CREATE TABLE IF NOT EXISTS importacao_v2_container_modelos (
            id            SERIAL       PRIMARY KEY,
            user_id       INT,
            user_nome     VARCHAR(160),
            nome          VARCHAR(180) NOT NULL,
            tipo_container VARCHAR(20) NOT NULL,
            capacidade_cbm FLOAT NOT NULL,
            containers    JSONB        NOT NULL,
            created_at    TIMESTAMP    NOT NULL DEFAULT NOW()
        )""",
        "CREATE INDEX IF NOT EXISTS idx_imp_v2_cont_modelos_created ON importacao_v2_container_modelos(created_at DESC)",
    ])


def seed_dummy_moq(admin_id: str) -> dict:
    """Popula parâmetros do produto (MOQ + medidas) em importacao_v2_moq — a fonte
    primária da página /importacao-v2/moq ("MOQ por SKU") e a origem das dimensões
    usadas em /containers (via _load_container_params_db). IDEMPOTENTE: usa
    ON CONFLICT (codigo) DO NOTHING e só insere os códigos ainda ausentes.

    Cobre a UNIÃO de:
      - dummy.PRODUTOS         → 12 códigos 104000xx (descrições "Produto Demo …")
      - IMPORTED_ITEM_CODES    → 30 códigos 10490001..10490030 (descrições
                                 genéricas "Produto Importado N")

    Isso corrige o descasamento que fazia o container "não encontrar as dimensões
    dos produtos": /calculate usa IMPORTED_ITEM_CODES por padrão, mas a tabela só
    tinha os 104000xx — agora ambos os conjuntos têm MOQ + dimensões.

    Conexão/commit/rollback próprios (try/finally). Garante a tabela via
    ensure_importacao_v2_modelos_table() e insere 1 linha por código, com MOQ,
    UNIT/CTN, CBM, peso (G.W/N.W), dimensões (LxWxH), preço e origem — todos
    determinísticos (dummy.rng). Retorna dict com a contagem inserida."""
    ensure_importacao_v2_modelos_table()

    # União determinística: (codigo, descricao, unidade) — dummy.PRODUTOS primeiro,
    # depois os importados (sem duplicar caso algum código se sobreponha).
    seen: set = set()
    catalogo: List[tuple] = []
    for codigo, descricao, unidade, _categoria in dummy.PRODUTOS:
        cod = str(codigo).strip()
        if cod in seen:
            continue
        seen.add(cod)
        catalogo.append((cod, descricao, unidade))
    for i, codigo in enumerate(IMPORTED_ITEM_CODES, start=1):
        cod = str(codigo).strip()
        if cod in seen:
            continue
        seen.add(cod)
        catalogo.append((cod, f"Produto Importado {i}", "UN"))

    conn = get_db_connection()
    cur = conn.cursor()
    try:
        origens = ["China", "Brasil"]
        n = 0
        for codigo, descricao, unidade in catalogo:
            r = dummy.rng("importacao_v2_moq", codigo)
            moq        = float(r.choice([500, 1000, 1500, 2000, 2500, 3000]))
            unit_ctn   = float(r.choice([12, 24, 36, 48, 60, 100]))
            comprimento = round(r.uniform(20.0, 60.0), 2)   # cm
            largura     = round(r.uniform(15.0, 45.0), 2)   # cm
            altura      = round(r.uniform(10.0, 40.0), 2)   # cm
            cbm = round((comprimento * largura * altura) / 1_000_000.0, 6)  # m³ por carton
            gw  = round(r.uniform(8.0, 25.0), 3)            # peso bruto (kg) por carton
            nw  = round(gw * r.uniform(0.80, 0.95), 3)      # peso líquido < bruto
            price  = round(r.uniform(2.5, 45.0), 4)         # preço unitário
            origem = r.choice(origens)
            cur.execute(
                """INSERT INTO importacao_v2_moq
                       (codigo, descricao, moq, origem, updated_by,
                        unit_ctn, cbm, gw, nw, comprimento, largura, altura,
                        price, unit)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                   ON CONFLICT (codigo) DO NOTHING""",
                (codigo, descricao, moq, origem,
                 int(admin_id) if admin_id and str(admin_id).isdigit() else None,
                 unit_ctn, cbm, gw, nw, comprimento, largura, altura,
                 price, unidade),
            )
            n += cur.rowcount  # conta só os realmente inseridos (ON CONFLICT = 0)

        conn.commit()
        return {"ok": True, "skipped": n == 0, "parametros": n,
                "total_catalogo": len(catalogo)}
    except Exception as e:
        conn.rollback()
        print(f"[importation_v2] seed_dummy_moq error: {e}")
        return {"ok": False, "erro": str(e)}
    finally:
        cur.close(); conn.close()


def seed_dummy_importacao_v2(admin_id: str) -> dict:
    """Popula o RESTO do fluxo de Importação V2 para a página ter dados ao abrir:
    modelos de cálculo, modelos de container e uma order list de exemplo.
    IDEMPOTENTE (ON CONFLICT / checagem de existência) e com conexão própria.

    Popula:
      - importacao_v2_modelos          (modelos de cálculo; codigos = IMPORTED_ITEM_CODES)
      - importacao_v2_container_modelos (20GP / 40GP / 40HC, com capacidade CBM)
      - importacao_v2_order_lists       (1 lista de exemplo p/ histórico)

    Também garante que as dimensões existam chamando seed_dummy_moq(admin_id),
    para o container achar CBM dos produtos importados.
    Retorna dict com a contagem por tabela."""
    ensure_importacao_v2_modelos_table()
    # Garante dimensões/MOQ para a UNIÃO (idempotente).
    moq_res = seed_dummy_moq(admin_id)

    uid = int(admin_id) if admin_id and str(admin_id).isdigit() else None
    nome_usuario = _fetch_user_nome(uid) if uid else None
    out: Dict[str, Any] = {"ok": True, "moq": moq_res}

    # ---- 1) Modelos de cálculo (importacao_v2_modelos) ----
    # UNIQUE(user_id, nome) → ON CONFLICT DO NOTHING garante idempotência.
    conn = get_db_connection(); cur = conn.cursor()
    try:
        modelos = [
            ("Importados — janela 15m (corrido)", IMPORTED_ITEM_CODES,
             DEFAULT_QTD_MESES, "corrido"),
            ("Importados — só meses com venda", IMPORTED_ITEM_CODES,
             12, "vendas"),
        ]
        n_mod = 0
        for nome, codigos, qtd_meses, modo in modelos:
            cur.execute(
                """INSERT INTO importacao_v2_modelos
                       (user_id, nome, codigos, qtd_meses, modo, overrides, threshold_sigma)
                   VALUES (%s, %s, %s, %s, %s, %s, %s)
                   ON CONFLICT (user_id, nome) DO NOTHING""",
                (uid, nome, json.dumps(list(codigos)), qtd_meses, modo,
                 json.dumps({}), DEFAULT_THRESHOLD_SIGMA),
            )
            n_mod += cur.rowcount
        conn.commit()
        out["modelos"] = n_mod
    except Exception as e:
        conn.rollback()
        out["modelos_erro"] = str(e)
        print(f"[importation_v2] seed_dummy_importacao_v2 modelos: {e}")
    finally:
        cur.close(); conn.close()

    # ---- 2) Modelos de container (importacao_v2_container_modelos) ----
    # Sem UNIQUE — usa checagem por nome p/ não duplicar.
    conn = get_db_connection(); cur = conn.cursor()
    try:
        container_modelos = [
            ("Container 20GP (padrão)",  "20'",  28.0),
            ("Container 40GP (padrão)",  "40'",  58.0),
            ("Container 40HC (padrão)",  "40HC", 68.0),
        ]
        n_cm = 0
        for nome, tipo, cap in container_modelos:
            cur.execute(
                "SELECT 1 FROM importacao_v2_container_modelos WHERE nome = %s LIMIT 1",
                (nome,),
            )
            if cur.fetchone():
                continue
            cur.execute(
                """INSERT INTO importacao_v2_container_modelos
                       (user_id, user_nome, nome, tipo_container, capacidade_cbm, containers)
                   VALUES (%s, %s, %s, %s, %s, %s)""",
                (uid, nome_usuario, nome, tipo, cap, json.dumps([])),
            )
            n_cm += 1
        conn.commit()
        out["container_modelos"] = n_cm
    except Exception as e:
        conn.rollback()
        out["container_modelos_erro"] = str(e)
        print(f"[importation_v2] seed_dummy_importacao_v2 container_modelos: {e}")
    finally:
        cur.close(); conn.close()

    # ---- 3) Order list de exemplo (importacao_v2_order_lists) ----
    conn = get_db_connection(); cur = conn.cursor()
    try:
        nome_ol = "Order List exemplo (importados)"
        cur.execute(
            "SELECT 1 FROM importacao_v2_order_lists WHERE nome = %s LIMIT 1",
            (nome_ol,),
        )
        n_ol = 0
        if not cur.fetchone():
            data_chegada = f"{dummy.ANO_BASE}-06-15"
            items = []
            datas_chegada: Dict[str, str] = {}
            for codigo in IMPORTED_ITEM_CODES[:10]:
                cod = str(codigo).strip()
                rr = dummy.rng("importacao_v2_orderlist", cod)
                items.append({"codigo": cod,
                              "qty": float(rr.choice([500, 1000, 1500, 2000])),
                              "data": data_chegada})
                datas_chegada[cod] = data_chegada
            cur.execute(
                """INSERT INTO importacao_v2_order_lists
                       (user_id, user_nome, nome, labels, observacao, items, datas_chegada)
                   VALUES (%s, %s, %s, %s, %s, %s, %s)""",
                (uid, nome_usuario, nome_ol, json.dumps(["Teste/Simulação"]),
                 "Gerado por seed_dummy_importacao_v2",
                 json.dumps(items), json.dumps(datas_chegada)),
            )
            n_ol = 1
        conn.commit()
        out["order_lists"] = n_ol
    except Exception as e:
        conn.rollback()
        out["order_lists_erro"] = str(e)
        print(f"[importation_v2] seed_dummy_importacao_v2 order_lists: {e}")
    finally:
        cur.close(); conn.close()

    return out


LABELS_PADRAO = [
    "Planejamento Q1", "Planejamento Q2", "Planejamento Q3", "Planejamento Q4",
    "Revisão mensal", "Revisão semanal", "Teste/Simulação",
    "Aprovado", "Aguarda revisão", "Compra confirmada", "Rejeitado",
]

# --- Defaults (do md consolidado) ---
DEFAULT_QTD_MESES        = 15      # janela de análise
DEFAULT_LEAD_TIME_MESES  = 3.0     # 90 dias ≈ 3 meses (decisão 2026-05-19)
DEFAULT_NIVEL_SERVICO    = 0.90    # 90% → Z = 1.2816
DEFAULT_THRESHOLD_SIGMA  = 1.5     # ±1.5σ marca pico/vale
LIMITE_GLOBAL_MESES      = 36      # teto rígido na varredura retroativa
DATA_PISO                = date(2023, 1, 1)  # nunca buscar vendas antes desta data


# =====================================================================
# Helpers estatísticos
# =====================================================================

def z_from_service_level(nivel: float) -> float:
    """Z da normal padrão para nível de serviço (0–1). Fallback p/ 1.2816 se inválido."""
    try:
        from scipy.stats import norm
        return float(norm.ppf(max(min(nivel, 0.9999), 0.5)))
    except Exception:
        # Aproximação para 90% se scipy ausente
        return 1.2816


def janela_corrido(hoje: date, n_meses: int) -> List[tuple]:
    """Últimos N meses calendário (não inclui o mês corrente). Não retorna meses antes de DATA_PISO.
    Retorna [(YYYY-MM, dias_no_mes)]."""
    meses = []
    for k in range(1, n_meses + 1):
        d = hoje.replace(day=1) - relativedelta(months=k)
        if d < DATA_PISO:
            break
        meses.append((d.strftime('%Y-%m'), monthrange(d.year, d.month)[1]))
    return list(reversed(meses))


def janela_vendas(hoje: date, n_meses: int, vendas_mes: Dict[str, float], limite: int = LIMITE_GLOBAL_MESES) -> List[tuple]:
    """Varredura retroativa: só meses com QTD > 0. Para no piso 2023-01-01 ou no limite N meses.
    Retorna [(YYYY-MM, dias_no_mes, qtd)]."""
    cursor = hoje.replace(day=1) - relativedelta(months=1)
    limite_dt = max(hoje.replace(day=1) - relativedelta(months=limite), DATA_PISO)
    coletados = []
    while len(coletados) < n_meses and cursor >= limite_dt:
        mes = cursor.strftime('%Y-%m')
        v = float(vendas_mes.get(mes, 0))
        if v > 0:
            coletados.append((mes, monthrange(cursor.year, cursor.month)[1], v))
        cursor -= relativedelta(months=1)
    return list(reversed(coletados))


def calc_sku(
    codigo: str,
    vendas_mes: Dict[str, float],
    estoque: float,
    hoje: date,
    qtd_meses: int,
    modo: str,
    lead_time: float,        # agora em MESES
    nivel_servico: float,
    pipeline: float,
    threshold_sigma: float,
    moq: float = 0,
) -> Dict[str, Any]:
    """Calcula todas as métricas para um SKU. Retorna dict com colunas e meses outliers."""

    # 1. Coleta da janela
    if modo == "vendas":
        coletados = janela_vendas(hoje, qtd_meses, vendas_mes)
        vendas_lista = [v for (_, _, v) in coletados]
        meses_lista = [(m, d) for (m, d, _) in coletados]
    else:  # corrido
        meses_lista = janela_corrido(hoje, qtd_meses)
        vendas_lista = [float(vendas_mes.get(m, 0)) for (m, _) in meses_lista]
        coletados = [(m, d, v) for (m, d), v in zip(meses_lista, vendas_lista)]

    # Usa apenas meses COM venda (>0) para ambos cálculos: média e desvio
    vendas_positivas = [v for v in vendas_lista if v > 0]
    meses_efetivos = len(vendas_positivas)
    total_vendas = sum(vendas_positivas)

    # 2. Consumo MENSAL (média sobre meses com venda)
    consumo_mensal = total_vendas / meses_efetivos if meses_efetivos > 0 else 0.0

    # 3. Desvio padrão amostral (STDEV.S) sobre os meses com venda
    if meses_efetivos >= 2:
        media = consumo_mensal
        sigma_mensal = math.sqrt(sum((x - media) ** 2 for x in vendas_positivas) / (meses_efetivos - 1))
    else:
        sigma_mensal = 0.0

    # 4. Safety stock em UNIDADES — escala mensal: Z × σ_mensal × √(LT_meses)
    z = z_from_service_level(nivel_servico)
    safety_stock = math.ceil(z * sigma_mensal * math.sqrt(lead_time)) if sigma_mensal > 0 else 0

    # 5. Ponto de reposição (unidades) — LT_meses × consumo_mensal + SS
    ponto_reposicao = math.ceil(lead_time * consumo_mensal + safety_stock)

    # 6. Estoque efetivo (com pipeline)
    estoque_total = float(estoque) + float(pipeline)

    # 7. Cobertura em MESES (com pipeline)
    cobertura_meses = (estoque_total / consumo_mensal) if consumo_mensal > 0 else 999.0
    # Mantém também em dias para compatibilidade no display rápido (1 mês ≈ 30d)
    cobertura_dias = cobertura_meses * 30

    # 8. Status de ruptura — 3 zonas (S-03 da validação)
    if estoque_total <= safety_stock:
        status = "RUPTURA"
    elif estoque_total < ponto_reposicao:
        status = "ATENCAO"
    else:
        status = "OK"

    # 9. Sugestão de compra
    # qtd_sugerida_pura = só o déficit (sem considerar MOQ)
    # qtd_sugerida      = aplicando MOQ (lote mínimo do fornecedor)
    deficit = max(0, math.ceil(ponto_reposicao - estoque_total))
    qtd_sugerida_pura = deficit
    if deficit > 0 and moq > 0:
        qtd_sugerida = max(deficit, int(math.ceil(moq)))
    else:
        qtd_sugerida = deficit

    # 10. Meses outliers (pico / vale) — referência: média dos meses com venda
    outliers = []
    if sigma_mensal > 0 and meses_efetivos >= 2:
        for (m, _d, v) in coletados:
            if v <= 0:
                continue
            if sigma_mensal > 0:
                z_score = (v - consumo_mensal) / sigma_mensal
                if z_score >= threshold_sigma:
                    outliers.append({"mes": m, "valor": v, "z": round(z_score, 2), "tipo": "pico"})
                elif z_score <= -threshold_sigma:
                    outliers.append({"mes": m, "valor": v, "z": round(z_score, 2), "tipo": "vale"})

    # 11. Avisos
    aviso = ""
    if modo == "vendas" and meses_efetivos < qtd_meses:
        aviso = f"Apenas {meses_efetivos} meses de venda registrados nos últimos {LIMITE_GLOBAL_MESES} meses"
    elif modo == "corrido" and len(vendas_lista) < qtd_meses:
        aviso = f"Hist. disponível: {len(vendas_lista)} meses (solicitado: {qtd_meses})"

    if len(vendas_lista) < 2:
        aviso = (aviso + " | " if aviso else "") + "σ não calculável (< 2 pontos)"

    return {
        "codigo": codigo,
        "estoque_disponivel": float(estoque),
        "pipeline": float(pipeline),
        "estoque_total": estoque_total,
        "lead_time": float(lead_time),         # em MESES
        "nivel_servico": nivel_servico,
        "moq": float(moq),
        "z": round(z, 4),
        "qtd_meses_solicitado": qtd_meses,
        "meses_efetivos": meses_efetivos,
        "total_vendas": round(total_vendas, 2),
        "consumo_mensal": round(consumo_mensal, 2),
        "sigma_mensal": round(sigma_mensal, 2),
        "estoque_seguranca": safety_stock,
        "ponto_reposicao": ponto_reposicao,
        "cobertura_meses": round(cobertura_meses, 1),
        "cobertura_dias": round(cobertura_dias, 1),  # mantido para compat
        "status": status,
        "qtd_sugerida_pura": qtd_sugerida_pura,
        "qtd_sugerida": qtd_sugerida,
        "outliers": outliers,
        "vendas_mensais": [{"mes": m, "valor": v} for (m, _d, v) in coletados],
        "aviso": aviso,
    }


# =====================================================================
# BigQuery — fetchers
# =====================================================================

def fetch_vendas_mensais(client, codigos: List[str], lookback_meses: int = LIMITE_GLOBAL_MESES) -> pd.DataFrame:
    cods = ",".join(f"'{c}'" for c in codigos)
    piso = DATA_PISO.strftime("%Y-%m-%d")
    q = f"""
        SELECT TRIM(CODIGO_PRODUTO) AS COD,
               FORMAT_DATE('%Y-%m', SAFE.PARSE_DATE('%Y-%m-%d', SUBSTR(EMISSAO,1,10))) AS MES,
               SUM(SAFE_CAST(QUANTIDADE AS FLOAT64)) AS QTD
        FROM `{PROJECT_ID}.VENDAS.VendasHistoricasDois`
        WHERE SAFE_CAST(SUBSTR(EMISSAO,1,10) AS DATE) >= GREATEST(
                DATE_SUB(CURRENT_DATE(), INTERVAL {lookback_meses} MONTH),
                DATE('{piso}')
              )
          AND DESC_TIPODOCUMENTO NOT IN ('BONIFICACAO','SAC','MOSTRUARIO','DISPLAY','CAMPANHAS','TROCA')
          AND EMPRESA = 'STAR_'
          AND TRIM(CODIGO_PRODUTO) IN ({cods})
        GROUP BY 1, 2
    """
    # DUMMY determinístico — ignora `client`/BigQuery. Gera vendas mensais para os
    # 12 meses de ANO_BASE (2026) por código, respeitando lookback_meses (último N
    # meses do calendário dummy). Shape preservado: colunas COD, MES, QTD.
    meses_str = dummy.meses_str("%Y-%m")          # ['2026-01', ..., '2026-12']
    meses_recorte = meses_str[-int(lookback_meses):] if lookback_meses > 0 else meses_str
    rows = []
    for cod in codigos:
        cod = str(cod).strip()
        r = dummy.rng("importation_v2_vendas", cod)
        # base de consumo mensal por SKU (estável por código)
        base = r.randint(40, 600)
        for mes in meses_recorte:
            rm = dummy.rng("importation_v2_vendas", cod, mes)
            qtd = float(max(0, round(base * (1 + rm.uniform(-0.35, 0.35)))))
            rows.append({"COD": cod, "MES": mes, "QTD": qtd})
    return pd.DataFrame(rows, columns=["COD", "MES", "QTD"])


def fetch_estoque(client, codigos: List[str]) -> pd.DataFrame:
    cods = ",".join(f"'{c}'" for c in codigos)
    q = f"""
        SELECT TRIM(codigo_do_item) AS COD,
               SUM(SAFE_CAST(REPLACE(CAST(quantidade AS STRING), ',', '.') AS FLOAT64)) AS DISPONIVEL
        FROM `{PROJECT_ID}.VENDAS.View_SaldoFisicoPorItem`
        WHERE codigo_do_local_estoque_ LIKE '13%'
          AND TRIM(codigo_do_item) IN ({cods})
        GROUP BY 1
    """
    # DUMMY determinístico — ignora `client`/BigQuery. Saldo físico por SKU.
    # Shape preservado: colunas COD, DISPONIVEL.
    rows = []
    for cod in codigos:
        cod = str(cod).strip()
        r = dummy.rng("importation_v2_estoque", cod)
        rows.append({"COD": cod, "DISPONIVEL": float(r.randint(0, 2500))})
    return pd.DataFrame(rows, columns=["COD", "DISPONIVEL"])


# =====================================================================
# Schemas
# =====================================================================

class SkuOverride(BaseModel):
    lead_time: Optional[float] = None        # em MESES
    nivel_servico: Optional[float] = None
    pipeline: Optional[float] = None
    moq: Optional[float] = None


class CalculateRequest(BaseModel):
    codigos: Optional[List[str]] = None                                # default = IMPORTED_ITEM_CODES
    qtd_meses: int = Field(default=DEFAULT_QTD_MESES, ge=1, le=LIMITE_GLOBAL_MESES)
    modo: str = Field(default="corrido", pattern="^(corrido|vendas)$")
    lead_time_default: float = Field(default=DEFAULT_LEAD_TIME_MESES, gt=0, le=120)  # MESES (até 10 anos)
    nivel_servico_default: float = Field(default=DEFAULT_NIVEL_SERVICO, gt=0.5, lt=1.0)
    threshold_sigma: float = Field(default=DEFAULT_THRESHOLD_SIGMA, ge=0.5, le=4.0)
    overrides: Dict[str, SkuOverride] = Field(default_factory=dict)


class ModeloIn(BaseModel):
    nome: str = Field(min_length=1, max_length=120)
    codigos: List[str]
    qtd_meses: int = Field(ge=1, le=LIMITE_GLOBAL_MESES)
    modo: str = Field(pattern="^(corrido|vendas)$")
    overrides: Dict[str, Dict[str, Any]] = Field(default_factory=dict)
    threshold_sigma: float = DEFAULT_THRESHOLD_SIGMA


# =====================================================================
# Endpoints
# =====================================================================

@router.get("/defaults")
def get_defaults(user_id: Optional[str] = Depends(get_user_id_from_session)):
    if not check_module_permission(user_id or "", MODULE_ID):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    # Quantos meses cabem entre DATA_PISO e hoje (cap dinâmico)
    hoje = date.today()
    delta_meses = (hoje.year - DATA_PISO.year) * 12 + (hoje.month - DATA_PISO.month)
    qtd_meses_max = min(LIMITE_GLOBAL_MESES, max(1, delta_meses))
    return {
        "qtd_meses": DEFAULT_QTD_MESES,
        "modo": "corrido",
        "lead_time_default": DEFAULT_LEAD_TIME_MESES,
        "lead_time_unidade": "meses",
        "nivel_servico_default": DEFAULT_NIVEL_SERVICO,
        "threshold_sigma": DEFAULT_THRESHOLD_SIGMA,
        "limite_global_meses": LIMITE_GLOBAL_MESES,
        "qtd_meses_max": qtd_meses_max,
        "data_piso": DATA_PISO.isoformat(),
        "codigos_padrao": IMPORTED_ITEM_CODES,
    }


@router.post("/calculate")
def calculate(body: CalculateRequest, user_id: Optional[str] = Depends(get_user_id_from_session)):
    if not check_module_permission(user_id or "", MODULE_ID):
        raise HTTPException(status_code=403, detail="Acesso negado.")

    codigos = [str(c).strip() for c in (body.codigos or IMPORTED_ITEM_CODES)]
    if not codigos:
        raise HTTPException(status_code=400, detail="Lista de códigos vazia.")

    # Cliente BigQuery não é mais necessário (dados dummy). Mantém a chamada por
    # compat, mas vira no-op: se as credenciais faltarem, segue com client=None.
    try:
        client = get_bq_client()
    except Exception as e:
        print(f"[importation_v2] get_bq_client no-op (dummy): {e}")
        client = None

    df_vendas = fetch_vendas_mensais(client, codigos, lookback_meses=LIMITE_GLOBAL_MESES)
    df_estoque = fetch_estoque(client, codigos)

    vendas_por_cod: Dict[str, Dict[str, float]] = {}
    for _, r in df_vendas.iterrows():
        vendas_por_cod.setdefault(r["COD"], {})[r["MES"]] = float(r["QTD"])
    estoque_dict = dict(zip(df_estoque["COD"], df_estoque["DISPONIVEL"]))

    # Descrição (planilha + fallback BQ)
    try:
        params_sheet = _load_sheet_parametros_importacao()
    except Exception:
        params_sheet = {}

    # MOQ do Excel local (ParametrosImportacao.xlsx)
    moq_map = _load_moq_map()

    hoje = date.today()
    resultados = []
    for cod in codigos:
        ov = body.overrides.get(cod, SkuOverride())
        lead = ov.lead_time if ov.lead_time is not None else body.lead_time_default
        nivel = ov.nivel_servico if ov.nivel_servico is not None else body.nivel_servico_default
        pipe = ov.pipeline if ov.pipeline is not None else 0.0
        moq_val = ov.moq if ov.moq is not None else moq_map.get(cod, 0.0)

        row = calc_sku(
            codigo=cod,
            vendas_mes=vendas_por_cod.get(cod, {}),
            estoque=estoque_dict.get(cod, 0),
            hoje=hoje,
            qtd_meses=body.qtd_meses,
            modo=body.modo,
            lead_time=lead,
            nivel_servico=nivel,
            pipeline=pipe,
            threshold_sigma=body.threshold_sigma,
            moq=moq_val,
        )
        info = params_sheet.get(cod, {}) if isinstance(params_sheet, dict) else {}
        row["descricao"] = info.get("descricao", "")
        row["peso_liquido"] = info.get("peso_liquido", 0)
        row["peso_bruto"] = info.get("peso_bruto", 0)
        resultados.append(row)

    return {
        "data_calculo": hoje.isoformat(),
        "parametros": {
            "qtd_meses": body.qtd_meses,
            "modo": body.modo,
            "lead_time_default": body.lead_time_default,
            "nivel_servico_default": body.nivel_servico_default,
            "threshold_sigma": body.threshold_sigma,
        },
        "itens": resultados,
    }


def _uid_int(user_id):
    """Cast user_id (str|int|None) para int, ou retorna None se inválido."""
    try:
        return int(user_id) if user_id not in (None, "", 0) else None
    except (ValueError, TypeError):
        return None


@router.get("/modelos")
def list_modelos(user_id: Optional[str] = Depends(get_user_id_from_session)):
    if not check_module_permission(user_id or "", MODULE_ID):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    uid = _uid_int(user_id)
    modelos = []
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            "SELECT id, nome, codigos, qtd_meses, modo, overrides, threshold_sigma, created_at "
            "FROM importacao_v2_modelos WHERE user_id = %s ORDER BY nome",
            (uid,),
        )
        rows = cur.fetchall()
        cur.close(); conn.close()
        modelos = [
            {
                "id": r[0], "nome": r[1], "codigos": r[2], "qtd_meses": r[3],
                "modo": r[4], "overrides": r[5], "threshold_sigma": float(r[6]),
                "created_at": r[7].isoformat() if r[7] else None,
            }
            for r in rows
        ]
    except Exception as e:
        # Tabela ainda não criada ou erro de DB — retorna só o padrão, não trava o frontend
        print(f"[importation_v2] list_modelos warning: {e}")
    padrao = {
        "id": 0, "nome": "Padrão (sistema)",
        "codigos": IMPORTED_ITEM_CODES, "qtd_meses": DEFAULT_QTD_MESES,
        "modo": "corrido", "overrides": {}, "threshold_sigma": DEFAULT_THRESHOLD_SIGMA,
        "created_at": None, "is_default": True,
    }
    return {"modelos": [padrao] + modelos}


@router.post("/modelos")
def create_modelo(body: ModeloIn, user_id: Optional[str] = Depends(get_user_id_from_session)):
    if not check_module_permission(user_id or "", MODULE_ID, min_permission="can_edit"):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    uid = _uid_int(user_id)  # pode ser None — persiste NULL no DB sem bloquear
    import json
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            "INSERT INTO importacao_v2_modelos (user_id, nome, codigos, qtd_meses, modo, overrides, threshold_sigma) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id",
            (uid, body.nome, json.dumps(body.codigos), body.qtd_meses, body.modo,
             json.dumps(body.overrides), body.threshold_sigma),
        )
        new_id = cur.fetchone()[0]
        conn.commit()
        return {"id": new_id, "nome": body.nome}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=400, detail=f"Erro ao salvar modelo: {e}")
    finally:
        cur.close(); conn.close()


@router.delete("/modelos/{modelo_id}")
def delete_modelo(modelo_id: int, user_id: Optional[str] = Depends(get_user_id_from_session)):
    if not check_module_permission(user_id or "", MODULE_ID, min_permission="can_edit"):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    uid = _uid_int(user_id)  # pode ser None — persiste NULL no DB sem bloquear
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM importacao_v2_modelos WHERE id = %s AND user_id = %s", (modelo_id, uid))
    affected = cur.rowcount
    conn.commit(); cur.close(); conn.close()
    if affected == 0:
        raise HTTPException(status_code=404, detail="Modelo não encontrado.")
    return {"ok": True}


# =====================================================================
# Versões (histórico)
# =====================================================================

class VersaoIn(BaseModel):
    nome: str = Field(min_length=1, max_length=180)
    labels: List[str] = Field(default_factory=list)
    observacao: Optional[str] = None
    parametros: Dict[str, Any]
    resultado: Dict[str, Any]


def _fetch_user_nome(uid: Optional[int]) -> Optional[str]:
    if uid is None:
        return None
    try:
        conn = get_db_connection(); cur = conn.cursor()
        cur.execute("SELECT name FROM users WHERE id = %s", (uid,))
        row = cur.fetchone()
        cur.close(); conn.close()
        return row[0] if row else None
    except Exception as e:
        print(f"[importation_v2] fetch user name error: {e}")
        return None


@router.get("/labels-padrao")
def get_labels_padrao(user_id: Optional[str] = Depends(get_user_id_from_session)):
    if not check_module_permission(user_id or "", MODULE_ID):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    return {"labels": LABELS_PADRAO}


VERSOES_RETENCAO_DIAS = 30  # versoes mais antigas que isso sao deletadas automaticamente


@router.get("/versoes")
def list_versoes(user_id: Optional[str] = Depends(get_user_id_from_session)):
    if not check_module_permission(user_id or "", MODULE_ID):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    versoes = []
    try:
        conn = get_db_connection(); cur = conn.cursor()
        # Limpeza automatica: descarta versoes mais antigas que VERSOES_RETENCAO_DIAS
        try:
            cur.execute(f"DELETE FROM importacao_v2_versoes WHERE created_at < NOW() - INTERVAL '{VERSOES_RETENCAO_DIAS} days'")
            conn.commit()
        except Exception as e:
            conn.rollback()
            print(f"[importation_v2] retencao versoes: {e}")
        cur.execute("""
            SELECT id, user_id, user_nome, nome, labels, observacao, parametros, created_at
            FROM importacao_v2_versoes
            ORDER BY created_at DESC
            LIMIT 200
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
        versoes = [
            {
                "id": r[0], "user_id": r[1], "user_nome": r[2] or f"user#{r[1]}" if r[1] else "—",
                "nome": r[3], "labels": r[4] or [], "observacao": r[5],
                "parametros": r[6], "created_at": r[7].isoformat() if r[7] else None,
            }
            for r in rows
        ]
    except Exception as e:
        print(f"[importation_v2] list_versoes warning: {e}")
    return {"versoes": versoes}


@router.get("/versoes/{versao_id}")
def get_versao(versao_id: int, user_id: Optional[str] = Depends(get_user_id_from_session)):
    if not check_module_permission(user_id or "", MODULE_ID):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    try:
        conn = get_db_connection(); cur = conn.cursor()
        cur.execute("""
            SELECT id, user_id, user_nome, nome, labels, observacao, parametros, resultado, created_at
            FROM importacao_v2_versoes WHERE id = %s
        """, (versao_id,))
        r = cur.fetchone()
        cur.close(); conn.close()
        if not r:
            raise HTTPException(status_code=404, detail="Versão não encontrada.")
        return {
            "id": r[0], "user_id": r[1], "user_nome": r[2] or f"user#{r[1]}" if r[1] else "—",
            "nome": r[3], "labels": r[4] or [], "observacao": r[5],
            "parametros": r[6], "resultado": r[7],
            "created_at": r[8].isoformat() if r[8] else None,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao carregar versão: {e}")


@router.post("/versoes")
def create_versao(body: VersaoIn, user_id: Optional[str] = Depends(get_user_id_from_session)):
    if not check_module_permission(user_id or "", MODULE_ID, min_permission="can_edit"):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    # Self-heal: garante tabela criada
    try:
        ensure_importacao_v2_modelos_table()
    except Exception as e:
        print(f"[importation_v2] create_versao ensure: {e}")
    uid = _uid_int(user_id)
    nome_usuario = _fetch_user_nome(uid) if uid else (str(user_id) if user_id else None)
    conn = get_db_connection(); cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO importacao_v2_versoes (user_id, user_nome, nome, labels, observacao, parametros, resultado)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            RETURNING id, created_at
        """, (
            uid, nome_usuario, body.nome,
            json.dumps(body.labels), body.observacao,
            json.dumps(body.parametros), json.dumps(body.resultado),
        ))
        new_id, created_at = cur.fetchone()
        conn.commit()
        return {"id": new_id, "nome": body.nome, "user_nome": nome_usuario,
                "created_at": created_at.isoformat() if created_at else None}
    except Exception as e:
        conn.rollback()
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=400, detail=f"Erro ao salvar versão ({type(e).__name__}): {e}")
    finally:
        cur.close(); conn.close()


# =====================================================================
# Sugestão de Container (bin-packing por CBM)
# =====================================================================

# Capacidades padrão (CBM) — pode ser editado pelo usuário no payload
CAPACIDADES_CONTAINER = {
    "20'":   28.0,   # 20 pés standard
    "40'":   58.0,   # 40 pés standard
    "40HC":  68.0,   # 40 pés high cube (padrão EMPRESA)
    "45HC":  86.0,   # 45 pés high cube
}


class ContainerItemIn(BaseModel):
    codigo: str
    descricao: Optional[str] = ""
    qtd: float = Field(ge=0)


class ContainersRequest(BaseModel):
    items: List[ContainerItemIn]  # itens a comprar (codigo + qtd)
    tipo: str = "40HC"            # "20'" | "40'" | "40HC" | "45HC" | "custom"
    capacidade_custom: Optional[float] = None  # se tipo=custom, em CBM


@router.post("/containers")
def sugestao_containers(body: ContainersRequest,
                        user_id: Optional[str] = Depends(get_user_id_from_session)):
    if not check_module_permission(user_id or "", MODULE_ID):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    if not body.items:
        raise HTTPException(status_code=400, detail="Lista de itens vazia.")

    capacidade = body.capacidade_custom if body.tipo == "custom" else CAPACIDADES_CONTAINER.get(body.tipo, 68.0)
    if not capacidade or capacidade <= 0:
        raise HTTPException(status_code=400, detail="Capacidade do container inválida.")

    params_map = _load_container_params()

    # Enriquece cada item com dados do Excel + calcula CBM/peso totais
    items_enriched: List[Dict[str, Any]] = []
    for it in body.items:
        cod = it.codigo.strip()
        if cod.endswith('.0'): cod = cod[:-2]
        if it.qtd <= 0:
            continue
        p = params_map.get(cod, {})
        unit_ctn = p.get('unit_ctn', 0) or 0
        cbm_unit = p.get('cbm', 0) or 0
        # CBM_total = CTNS (arredondado pra cima) × CBM_unit
        ctns = (math.ceil(it.qtd / unit_ctn) if unit_ctn > 0 else 0)
        cbm_total = round(ctns * cbm_unit, 4)
        peso_total = round(it.qtd * (p.get('gw', 0) or 0), 2)
        peso_liq_total = round(it.qtd * (p.get('peso_liquido', 0) or 0), 2)
        amount = round(it.qtd * (p.get('price', 0) or 0), 2)
        items_enriched.append({
            'codigo': cod,
            'descricao': it.descricao or '',
            'qtd': float(it.qtd),
            'unit_ctn': int(unit_ctn),
            'ctns': ctns,
            'cbm_unit': cbm_unit,
            'cbm_total': cbm_total,
            'price': p.get('price', 0) or 0,
            'amount': amount,
            'peso_unit': p.get('gw', 0) or 0,
            'peso_total': peso_total,
            'peso_liquido_unit': p.get('peso_liquido', 0) or 0,
            'peso_liquido_total': peso_liq_total,
            'medidas': f"{p.get('l', 0)}x{p.get('w', 0)}x{p.get('h', 0)}",
            'unit': p.get('unit', ''),
            'ncm': p.get('ncm', ''),
            'l': p.get('l', 0) or 0,
            'w': p.get('w', 0) or 0,
            'h': p.get('h', 0) or 0,
            'barcode': p.get('barcode', ''),
            'name_cn': p.get('name_cn', ''),
            'remark': p.get('remark', ''),
            'obs': p.get('obs', ''),
            'observacoes': p.get('observacoes', ''),
            'english_description': p.get('english_description', ''),
            'cbm_faltante': cbm_unit == 0,  # flag se faltam dimensoes do produto
        })

    # Itens sem CBM ficam separados (não dá pra empacotar)
    items_sem_cbm = [it for it in items_enriched if it['cbm_total'] <= 0]
    items_pack = [it for it in items_enriched if it['cbm_total'] > 0]

    # Sort por CBM_total desc (greedy - First Fit Decreasing)
    items_pack.sort(key=lambda x: x['cbm_total'], reverse=True)

    containers: List[Dict[str, Any]] = []
    current_id = 1
    current_items: List[Dict[str, Any]] = []
    current_volume = 0.0

    label_tipo = body.tipo if body.tipo != "custom" else f"custom {capacidade:.0f}m³"

    def close_container():
        nonlocal current_id, current_items, current_volume
        if not current_items:
            return
        containers.append({
            'id': current_id,
            'label': f"Container {current_id} ({label_tipo})",
            'capacidade_cbm': capacidade,
            'volume_usado_cbm': round(current_volume, 2),
            'volume_livre_cbm': round(capacidade - current_volume, 2),
            'ocupacao_pct': round((current_volume / capacidade) * 100, 1),
            'itens': current_items,
            'total_qtd': sum(it['qtd'] for it in current_items),
            'total_amount': round(sum(it['amount'] for it in current_items), 2),
            'total_peso': round(sum(it['peso_total'] for it in current_items), 2),
            'total_ctns': sum(it['ctns'] for it in current_items),
        })
        current_id += 1
        current_items = []
        current_volume = 0.0

    for it in items_pack:
        # Se item sozinho passa da capacidade, divide em quantos containers precisar
        if it['cbm_total'] > capacidade:
            close_container()  # fecha o que tava aberto
            # Quantos containers cheios pra esse item?
            cbm_per_ctn = it['cbm_unit']
            ctns_per_container = math.floor(capacidade / cbm_per_ctn) if cbm_per_ctn > 0 else 0
            remaining_ctns = it['ctns']
            qtd_per_ctn = it['unit_ctn'] or 1
            while remaining_ctns > 0:
                ctns_neste = min(remaining_ctns, ctns_per_container)
                qtd_neste = ctns_neste * qtd_per_ctn
                cbm_neste = round(ctns_neste * cbm_per_ctn, 4)
                containers.append({
                    'id': current_id,
                    'label': f"Container {current_id} ({label_tipo})",
                    'capacidade_cbm': capacidade,
                    'volume_usado_cbm': cbm_neste,
                    'volume_livre_cbm': round(capacidade - cbm_neste, 2),
                    'ocupacao_pct': round((cbm_neste / capacidade) * 100, 1),
                    'itens': [{**it, 'qtd': qtd_neste, 'ctns': ctns_neste, 'cbm_total': cbm_neste,
                               'amount': round(qtd_neste * it['price'], 2),
                               'peso_total': round(qtd_neste * it['peso_unit'], 2)}],
                    'total_qtd': qtd_neste,
                    'total_amount': round(qtd_neste * it['price'], 2),
                    'total_peso': round(qtd_neste * it['peso_unit'], 2),
                    'total_ctns': ctns_neste,
                })
                current_id += 1
                remaining_ctns -= ctns_neste
            continue

        if current_volume + it['cbm_total'] > capacidade and current_items:
            close_container()
        current_items.append(it)
        current_volume += it['cbm_total']

    close_container()

    return {
        "tipo": body.tipo,
        "capacidade_cbm": capacidade,
        "total_containers": len(containers),
        "total_amount": round(sum(c['total_amount'] for c in containers), 2),
        "total_qtd": sum(c['total_qtd'] for c in containers),
        "containers": containers,
        "itens_sem_dimensoes": items_sem_cbm,  # produtos sem CBM cadastrado
        "capacidades_disponiveis": CAPACIDADES_CONTAINER,
    }


# =====================================================================
# Order Lists — histórico de uploads (mantém 30 mais recentes)
# =====================================================================

ORDER_LISTS_MAX = 30


class OrderListItem(BaseModel):
    codigo: str
    qty: float
    data: Optional[str] = None  # YYYY-MM-DD


class OrderListIn(BaseModel):
    nome: str = Field(min_length=1, max_length=180)
    labels: List[str] = Field(default_factory=list)
    observacao: Optional[str] = None
    items: List[OrderListItem]
    datas_chegada: Dict[str, str] = Field(default_factory=dict)


@router.get("/order-lists")
def list_order_lists(user_id: Optional[str] = Depends(get_user_id_from_session)):
    if not check_module_permission(user_id or "", MODULE_ID):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    rows = []
    try:
        conn = get_db_connection(); cur = conn.cursor()
        # Mantém só os 30 mais recentes — apaga os antigos
        try:
            cur.execute(f"""
                DELETE FROM importacao_v2_order_lists
                WHERE id NOT IN (
                    SELECT id FROM importacao_v2_order_lists
                    ORDER BY created_at DESC
                    LIMIT {ORDER_LISTS_MAX}
                )
            """)
            conn.commit()
        except Exception as e:
            conn.rollback()
            print(f"[importation_v2] retencao order_lists: {e}")
        cur.execute("""
            SELECT id, user_id, user_nome, nome, labels, observacao,
                   jsonb_array_length(items) AS qtd_itens, created_at
            FROM importacao_v2_order_lists
            ORDER BY created_at DESC
            LIMIT 200
        """)
        for r in cur.fetchall():
            rows.append({
                "id": r[0], "user_id": r[1],
                "user_nome": r[2] or (f"user#{r[1]}" if r[1] else "—"),
                "nome": r[3], "labels": r[4] or [], "observacao": r[5],
                "qtd_itens": r[6], "created_at": r[7].isoformat() if r[7] else None,
            })
        cur.close(); conn.close()
    except Exception as e:
        print(f"[importation_v2] list_order_lists: {e}")
    return {"order_lists": rows, "limite": ORDER_LISTS_MAX}


@router.get("/order-lists/{ol_id}")
def get_order_list(ol_id: int, user_id: Optional[str] = Depends(get_user_id_from_session)):
    if not check_module_permission(user_id or "", MODULE_ID):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    conn = get_db_connection(); cur = conn.cursor()
    try:
        cur.execute("""
            SELECT id, user_id, user_nome, nome, labels, observacao, items, datas_chegada, created_at
            FROM importacao_v2_order_lists WHERE id = %s
        """, (ol_id,))
        r = cur.fetchone()
        if not r:
            raise HTTPException(status_code=404, detail="Order List não encontrada.")
        return {
            "id": r[0], "user_id": r[1],
            "user_nome": r[2] or (f"user#{r[1]}" if r[1] else "—"),
            "nome": r[3], "labels": r[4] or [], "observacao": r[5],
            "items": r[6] or [], "datas_chegada": r[7] or {},
            "created_at": r[8].isoformat() if r[8] else None,
        }
    finally:
        cur.close(); conn.close()


@router.post("/order-lists")
def create_order_list(body: OrderListIn, user_id: Optional[str] = Depends(get_user_id_from_session)):
    if not check_module_permission(user_id or "", MODULE_ID, min_permission="can_edit"):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    try:
        ensure_importacao_v2_modelos_table()
    except Exception as e:
        print(f"[importation_v2] create_order_list ensure: {e}")
    uid = _uid_int(user_id)
    nome_usuario = _fetch_user_nome(uid) if uid else (str(user_id) if user_id else None)
    conn = get_db_connection(); cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO importacao_v2_order_lists
              (user_id, user_nome, nome, labels, observacao, items, datas_chegada)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            RETURNING id, created_at
        """, (
            uid, nome_usuario, body.nome,
            json.dumps(body.labels), body.observacao,
            json.dumps([it.dict() for it in body.items]),
            json.dumps(body.datas_chegada),
        ))
        new_id, created_at = cur.fetchone()
        conn.commit()
        # Garante o limite de 30
        try:
            cur.execute(f"""
                DELETE FROM importacao_v2_order_lists
                WHERE id NOT IN (
                    SELECT id FROM importacao_v2_order_lists
                    ORDER BY created_at DESC LIMIT {ORDER_LISTS_MAX}
                )
            """)
            conn.commit()
        except Exception: conn.rollback()
        return {"id": new_id, "nome": body.nome, "user_nome": nome_usuario,
                "created_at": created_at.isoformat() if created_at else None}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=400, detail=f"Erro ao salvar order list: {e}")
    finally:
        cur.close(); conn.close()


@router.delete("/order-lists/{ol_id}")
def delete_order_list(ol_id: int, user_id: Optional[str] = Depends(get_user_id_from_session)):
    if not check_module_permission(user_id or "", MODULE_ID, min_permission="can_edit"):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    conn = get_db_connection(); cur = conn.cursor()
    cur.execute("DELETE FROM importacao_v2_order_lists WHERE id = %s", (ol_id,))
    affected = cur.rowcount
    conn.commit(); cur.close(); conn.close()
    if affected == 0:
        raise HTTPException(status_code=404, detail="Order List não encontrada.")
    return {"ok": True}


# =====================================================================
# Container Modelos — Salvar / Listar / Carregar / Excluir
# =====================================================================

@router.get("/container-modelos")
def list_container_modelos(user_id: Optional[str] = Depends(get_user_id_from_session)):
    if not check_module_permission(user_id or "", MODULE_ID):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    try: ensure_importacao_v2_modelos_table()
    except: pass
    conn = get_db_connection(); cur = conn.cursor()
    try:
        cur.execute("""
            SELECT id, user_nome, nome, tipo_container, capacidade_cbm, created_at
            FROM importacao_v2_container_modelos
            ORDER BY created_at DESC LIMIT 30
        """)
        cols = [c[0] for c in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
        for r in rows:
            if r.get('created_at'): r['created_at'] = r['created_at'].isoformat()
        return {"modelos": rows}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close(); conn.close()


@router.post("/container-modelos")
def create_container_modelo(body: ContainerModeloIn, user_id: Optional[str] = Depends(get_user_id_from_session)):
    if not check_module_permission(user_id or "", MODULE_ID, min_permission="can_edit"):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    try: ensure_importacao_v2_modelos_table()
    except: pass
    uid = _uid_int(user_id)
    nome_usuario = _fetch_user_nome(uid) if uid else None
    conn = get_db_connection(); cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO importacao_v2_container_modelos
              (user_id, user_nome, nome, tipo_container, capacidade_cbm, containers)
            VALUES (%s, %s, %s, %s, %s, %s)
            RETURNING id, created_at
        """, (uid, nome_usuario, body.nome, body.tipo_container, body.capacidade_cbm, json.dumps(body.containers)))
        new_id, created_at = cur.fetchone()
        conn.commit()
        try:
            cur.execute("""DELETE FROM importacao_v2_container_modelos WHERE id NOT IN (
                SELECT id FROM importacao_v2_container_modelos ORDER BY created_at DESC LIMIT 30)""")
            conn.commit()
        except: conn.rollback()
        return {"id": new_id, "nome": body.nome, "user_nome": nome_usuario, "created_at": created_at.isoformat() if created_at else None}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=400, detail=f"Erro ao salvar modelo: {e}")
    finally:
        cur.close(); conn.close()


@router.delete("/container-modelos/{cm_id}")
def delete_container_modelo(cm_id: int, user_id: Optional[str] = Depends(get_user_id_from_session)):
    if not check_module_permission(user_id or "", MODULE_ID, min_permission="can_edit"):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    conn = get_db_connection(); cur = conn.cursor()
    cur.execute("DELETE FROM importacao_v2_container_modelos WHERE id = %s", (cm_id,))
    affected = cur.rowcount
    conn.commit(); cur.close(); conn.close()
    if affected == 0:
        raise HTTPException(status_code=404, detail="Modelo não encontrado.")
    return {"ok": True}


@router.get("/container-modelos/{cm_id}")
def get_container_modelo(cm_id: int, user_id: Optional[str] = Depends(get_user_id_from_session)):
    if not check_module_permission(user_id or "", MODULE_ID):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    conn = get_db_connection(); cur = conn.cursor()
    try:
        cur.execute("SELECT id, user_nome, nome, tipo_container, capacidade_cbm, containers, created_at FROM importacao_v2_container_modelos WHERE id = %s", (cm_id,))
        row = cur.fetchone()
        if not row: raise HTTPException(status_code=404, detail="Modelo não encontrado.")
        cols = [c[0] for c in cur.description]
        r = dict(zip(cols, row))
        if isinstance(r['containers'], str): r['containers'] = json.loads(r['containers'])
        if r.get('created_at'): r['created_at'] = r['created_at'].isoformat()
        return r
    finally:
        cur.close(); conn.close()


# =====================================================================
# MOQ — CRUD + Upload de planilha
# =====================================================================

class ContainerModeloIn(BaseModel):
    nome: str
    tipo_container: str
    capacidade_cbm: float
    containers: list


class MoqIn(BaseModel):
    codigo: str = Field(min_length=1, max_length=40)
    descricao: Optional[str] = None
    moq: float = Field(ge=0)
    unit_ctn: Optional[float] = None
    cbm: Optional[float] = None
    gw: Optional[float] = None
    nw: Optional[float] = None
    comprimento: Optional[float] = None
    largura: Optional[float] = None
    altura: Optional[float] = None
    price: Optional[float] = None
    ncm: Optional[str] = None
    unit: Optional[str] = None
    barcode: Optional[str] = None
    name_cn: Optional[str] = None
    remark: Optional[str] = None
    obs: Optional[str] = None
    observacoes: Optional[str] = None
    english_description: Optional[str] = None
    ctns: Optional[float] = None
    qty: Optional[float] = None
    amount: Optional[float] = None
    cbm_total: Optional[float] = None
    tgw: Optional[float] = None
    tnw: Optional[float] = None


@router.get("/moq")
def list_moq(user_id: Optional[str] = Depends(get_user_id_from_session)):
    if not check_module_permission(user_id or "", MODULE_ID):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    rows = []
    user_names: Dict[int, str] = {}
    try:
        conn = get_db_connection(); cur = conn.cursor()
        # 1) Lê MOQ sem join — mais robusto
        cur.execute("""
            SELECT codigo, descricao, moq, origem, updated_at, updated_by,
                   COALESCE(unit_ctn, 0), COALESCE(cbm, 0),
                   COALESCE(gw, 0), COALESCE(nw, 0),
                   COALESCE(comprimento, 0), COALESCE(largura, 0), COALESCE(altura, 0),
                   COALESCE(price, 0), COALESCE(ncm, ''), COALESCE(unit, ''),
                   COALESCE(barcode, ''), COALESCE(name_cn, ''),
                   COALESCE(remark, ''), COALESCE(obs, ''), COALESCE(observacoes, ''),
                   COALESCE(english_description, ''),
                   COALESCE(ctns, 0), COALESCE(qty, 0), COALESCE(amount, 0),
                   COALESCE(cbm_total, 0), COALESCE(tgw, 0), COALESCE(tnw, 0)
            FROM importacao_v2_moq
            ORDER BY codigo
        """)
        raw_rows = cur.fetchall()
        # 2) Mapa de nomes de usuário (tenta 'name' depois 'nome')
        uids = list({r[5] for r in raw_rows if r[5] is not None})
        if uids:
            for col in ('name', 'nome', 'full_name', 'fullname', 'email'):
                try:
                    cur.execute(f"SELECT id, {col} FROM users WHERE id = ANY(%s)", (uids,))
                    for u_id, u_name in cur.fetchall():
                        if u_name:
                            user_names[u_id] = str(u_name)
                    if user_names:
                        break
                except Exception:
                    continue
        for r in raw_rows:
            rows.append({
                "codigo": r[0], "descricao": r[1], "moq": float(r[2]),
                "origem": r[3], "updated_at": r[4].isoformat() if r[4] else None,
                "updated_by": user_names.get(r[5], f"user#{r[5]}" if r[5] is not None else "—"),
                "unit_ctn": float(r[6]) if r[6] else 0,
                "cbm": float(r[7]) if r[7] else 0,
                "gw": float(r[8]) if r[8] else 0,
                "nw": float(r[9]) if r[9] else 0,
                "comprimento": float(r[10]) if r[10] else 0,
                "largura": float(r[11]) if r[11] else 0,
                "altura": float(r[12]) if r[12] else 0,
                "price": float(r[13]) if r[13] else 0,
                "ncm": r[14] or '',
                "unit": r[15] or '',
                "barcode": r[16] or '',
                "name_cn": r[17] or '',
                "remark": r[18] or '',
                "obs": r[19] or '',
                "observacoes": r[20] or '',
                "english_description": r[21] or '',
                "ctns": float(r[22]) if r[22] else 0,
                "qty": float(r[23]) if r[23] else 0,
                "amount": float(r[24]) if r[24] else 0,
                "cbm_total": float(r[25]) if r[25] else 0,
                "tgw": float(r[26]) if r[26] else 0,
                "tnw": float(r[27]) if r[27] else 0,
            })
        cur.close(); conn.close()
    except Exception as e:
        import traceback; traceback.print_exc()
        print(f"[importation_v2] list_moq error: {e}")
    return {"items": rows}


@router.put("/moq/{codigo}")
def upsert_moq(codigo: str, body: MoqIn, user_id: Optional[str] = Depends(get_user_id_from_session)):
    if not check_module_permission(user_id or "", MODULE_ID, min_permission="can_edit"):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    uid = _uid_int(user_id)
    conn = get_db_connection(); cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO importacao_v2_moq
              (codigo, descricao, moq, origem, updated_by, updated_at,
               unit_ctn, cbm, gw, nw, comprimento, largura, altura, price, ncm, unit,
               barcode, name_cn, remark, obs, observacoes, english_description,
               ctns, qty, amount, cbm_total, tgw, tnw)
            VALUES (%s, %s, %s, 'manual', %s, NOW(),
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s)
            ON CONFLICT (codigo) DO UPDATE SET
              descricao   = EXCLUDED.descricao,
              moq         = EXCLUDED.moq,
              origem      = 'manual',
              updated_by  = EXCLUDED.updated_by,
              updated_at  = NOW(),
              unit_ctn    = EXCLUDED.unit_ctn,
              cbm         = EXCLUDED.cbm,
              gw          = EXCLUDED.gw,
              nw          = EXCLUDED.nw,
              comprimento = EXCLUDED.comprimento,
              largura     = EXCLUDED.largura,
              altura      = EXCLUDED.altura,
              price       = EXCLUDED.price,
              ncm         = EXCLUDED.ncm,
              unit        = EXCLUDED.unit,
              barcode             = EXCLUDED.barcode,
              name_cn             = EXCLUDED.name_cn,
              remark              = EXCLUDED.remark,
              obs                 = EXCLUDED.obs,
              observacoes         = EXCLUDED.observacoes,
              english_description = EXCLUDED.english_description,
              ctns                = EXCLUDED.ctns,
              qty                 = EXCLUDED.qty,
              amount              = EXCLUDED.amount,
              cbm_total           = EXCLUDED.cbm_total,
              tgw                 = EXCLUDED.tgw,
              tnw                 = EXCLUDED.tnw
        """, (
            codigo.strip(), (body.descricao or '').strip() or None, body.moq, uid,
            body.unit_ctn, body.cbm, body.gw, body.nw,
            body.comprimento, body.largura, body.altura,
            body.price, body.ncm, body.unit,
            body.barcode, body.name_cn, body.remark, body.obs, body.observacoes,
            body.english_description,
            body.ctns, body.qty, body.amount, body.cbm_total, body.tgw, body.tnw,
        ))
        conn.commit()
        return {"ok": True}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=400, detail=f"Erro ao salvar parâmetros: {e}")
    finally:
        cur.close(); conn.close()


@router.delete("/moq/{codigo}")
def delete_moq(codigo: str, user_id: Optional[str] = Depends(get_user_id_from_session)):
    if not check_module_permission(user_id or "", MODULE_ID, min_permission="can_edit"):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    conn = get_db_connection(); cur = conn.cursor()
    try:
        cur.execute("DELETE FROM importacao_v2_moq WHERE codigo = %s", (codigo.strip(),))
        affected = cur.rowcount
        conn.commit()
        if affected == 0:
            raise HTTPException(status_code=404, detail="SKU não encontrado.")
        return {"ok": True}
    finally:
        cur.close(); conn.close()


class MoqBulkItem(BaseModel):
    codigo: str
    descricao: Optional[str] = None
    moq: float
    unit_ctn: Optional[float] = None
    cbm: Optional[float] = None
    gw: Optional[float] = None
    nw: Optional[float] = None
    comprimento: Optional[float] = None
    largura: Optional[float] = None
    altura: Optional[float] = None
    price: Optional[float] = None
    ncm: Optional[str] = None
    unit: Optional[str] = None
    barcode: Optional[str] = None
    name_cn: Optional[str] = None
    remark: Optional[str] = None
    obs: Optional[str] = None
    observacoes: Optional[str] = None
    english_description: Optional[str] = None
    ctns: Optional[float] = None
    qty: Optional[float] = None
    amount: Optional[float] = None
    cbm_total: Optional[float] = None
    tgw: Optional[float] = None
    tnw: Optional[float] = None


class MoqBulkBody(BaseModel):
    items: List[MoqBulkItem]


@router.post("/moq/bulk")
def bulk_upsert_moq(body: MoqBulkBody, user_id: Optional[str] = Depends(get_user_id_from_session)):
    """Recebe JSON com lista de {codigo, descricao, moq} (já parseado no frontend) e faz upsert.
    Bem mais leve que /moq/upload — evita problemas de upload de planilhas grandes."""
    if not check_module_permission(user_id or "", MODULE_ID, min_permission="can_edit"):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    uid = _uid_int(user_id)
    # Cria tabela inline (sem FK) — propaga erro real se falhar
    try:
        conn_e = get_db_connection(); cur_e = conn_e.cursor()
        try:
            cur_e.execute("""
                CREATE TABLE IF NOT EXISTS importacao_v2_moq (
                    codigo       VARCHAR(40)   PRIMARY KEY,
                    descricao    VARCHAR(180),
                    moq          NUMERIC(12,2) NOT NULL,
                    origem       VARCHAR(40)   DEFAULT 'manual',
                    updated_by   INT,
                    updated_at   TIMESTAMP     NOT NULL DEFAULT NOW()
                )
            """)
            # Self-heal: garante colunas de medidas
            for col_def in [
                "unit_ctn NUMERIC(12,2)", "cbm NUMERIC(12,6)",
                "gw NUMERIC(12,3)", "nw NUMERIC(12,3)",
                "comprimento NUMERIC(10,2)", "largura NUMERIC(10,2)", "altura NUMERIC(10,2)",
                "price NUMERIC(12,4)", "ncm VARCHAR(40)", "unit VARCHAR(20)",
                "barcode VARCHAR(60)", "name_cn VARCHAR(255)",
                "remark TEXT", "obs TEXT", "observacoes TEXT",
                "english_description VARCHAR(255)",
                "ctns NUMERIC(12,2)", "qty NUMERIC(12,2)", "amount NUMERIC(14,2)",
                "cbm_total NUMERIC(12,4)", "tgw NUMERIC(12,3)", "tnw NUMERIC(12,3)",
            ]:
                try:
                    cur_e.execute(f"ALTER TABLE importacao_v2_moq ADD COLUMN IF NOT EXISTS {col_def}")
                except Exception as e:
                    print(f"[importation_v2] ALTER {col_def}: {e}")
            conn_e.commit()
        finally:
            cur_e.close(); conn_e.close()
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Falha ao criar tabela importacao_v2_moq: {type(e).__name__}: {e}")

    valid_uid = uid  # sem FK — qualquer int vale; None tb vale

    # SKUs existentes
    existentes = set()
    try:
        conn0 = get_db_connection(); cur0 = conn0.cursor()
        cur0.execute("SELECT codigo FROM importacao_v2_moq")
        existentes = {r[0] for r in cur0.fetchall()}
        cur0.close(); conn0.close()
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Erro lendo tabela: {type(e).__name__}: {e}")

    inseridos = 0; atualizados = 0; ignorados = 0
    erros: List[str] = []
    for it in body.items:
        cod = (it.codigo or "").strip()
        if cod.endswith('.0'): cod = cod[:-2]
        if not cod:
            ignorados += 1; continue
        # MOQ vazio/negativo vira 0 — todas as linhas com ITEM NO entram
        moq_val = float(it.moq) if (it.moq is not None and it.moq >= 0) else 0.0
        desc = (it.descricao or "").strip()[:180] or None
        conn = get_db_connection(); cur = conn.cursor()
        try:
            cur.execute("""
                INSERT INTO importacao_v2_moq
                  (codigo, descricao, moq, origem, updated_by, updated_at,
                   unit_ctn, cbm, gw, nw, comprimento, largura, altura, price, ncm, unit,
                   barcode, name_cn, remark, obs, observacoes, english_description,
                   ctns, qty, amount, cbm_total, tgw, tnw)
                VALUES (%s, %s, %s, 'upload', %s, NOW(),
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s, %s)
                ON CONFLICT (codigo) DO UPDATE SET
                  descricao   = COALESCE(EXCLUDED.descricao, importacao_v2_moq.descricao),
                  moq         = EXCLUDED.moq,
                  origem      = 'upload',
                  updated_by  = EXCLUDED.updated_by,
                  updated_at  = NOW(),
                  unit_ctn    = COALESCE(EXCLUDED.unit_ctn, importacao_v2_moq.unit_ctn),
                  cbm         = COALESCE(EXCLUDED.cbm, importacao_v2_moq.cbm),
                  gw          = COALESCE(EXCLUDED.gw, importacao_v2_moq.gw),
                  nw          = COALESCE(EXCLUDED.nw, importacao_v2_moq.nw),
                  comprimento = COALESCE(EXCLUDED.comprimento, importacao_v2_moq.comprimento),
                  largura     = COALESCE(EXCLUDED.largura, importacao_v2_moq.largura),
                  altura      = COALESCE(EXCLUDED.altura, importacao_v2_moq.altura),
                  price       = COALESCE(EXCLUDED.price, importacao_v2_moq.price),
                  ncm         = COALESCE(EXCLUDED.ncm, importacao_v2_moq.ncm),
                  unit        = COALESCE(EXCLUDED.unit, importacao_v2_moq.unit),
                  barcode             = COALESCE(EXCLUDED.barcode, importacao_v2_moq.barcode),
                  name_cn             = COALESCE(EXCLUDED.name_cn, importacao_v2_moq.name_cn),
                  remark              = COALESCE(EXCLUDED.remark, importacao_v2_moq.remark),
                  obs                 = COALESCE(EXCLUDED.obs, importacao_v2_moq.obs),
                  observacoes         = COALESCE(EXCLUDED.observacoes, importacao_v2_moq.observacoes),
                  english_description = COALESCE(EXCLUDED.english_description, importacao_v2_moq.english_description),
                  ctns                = COALESCE(EXCLUDED.ctns, importacao_v2_moq.ctns),
                  qty                 = COALESCE(EXCLUDED.qty, importacao_v2_moq.qty),
                  amount              = COALESCE(EXCLUDED.amount, importacao_v2_moq.amount),
                  cbm_total           = COALESCE(EXCLUDED.cbm_total, importacao_v2_moq.cbm_total),
                  tgw                 = COALESCE(EXCLUDED.tgw, importacao_v2_moq.tgw),
                  tnw                 = COALESCE(EXCLUDED.tnw, importacao_v2_moq.tnw)
            """, (
                cod, desc, moq_val, valid_uid,
                it.unit_ctn, it.cbm, it.gw, it.nw,
                it.comprimento, it.largura, it.altura,
                it.price, it.ncm, it.unit,
                it.barcode, it.name_cn, it.remark, it.obs, it.observacoes,
                it.english_description,
                it.ctns, it.qty, it.amount, it.cbm_total, it.tgw, it.tnw,
            ))
            conn.commit()
            if cod in existentes:
                atualizados += 1
            else:
                inseridos += 1
                existentes.add(cod)
        except Exception as row_err:
            conn.rollback()
            msg = f"{cod}: {type(row_err).__name__}: {row_err}"
            print(f"[importation_v2] bulk skip {msg}")
            erros.append(msg[:200])
            ignorados += 1
        finally:
            cur.close(); conn.close()

    return {"ok": True, "inseridos": inseridos, "atualizados": atualizados, "ignorados": ignorados, "erros": erros[:10]}


@router.post("/moq/upload")
async def upload_moq_xlsx(
    file: UploadFile = File(...),
    user_id: Optional[str] = Depends(get_user_id_from_session),
):
    """Lê uma planilha de 1 aba e faz upsert em massa.
    Procura colunas: ITEM NO / Código / Codigo EMPRESA  +  MOQ.
    Descrição opcional via DESCRIPTION / Descrição.
    """
    if not check_module_permission(user_id or "", MODULE_ID, min_permission="can_edit"):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    uid = _uid_int(user_id)
    if not file.filename or not file.filename.lower().endswith(('.xlsx', '.xls', '.xlsm')):
        raise HTTPException(status_code=400, detail="Envie um arquivo Excel (.xlsx).")
    # Garante que a tabela existe (self-healing caso ensure_table no startup não tenha rodado)
    try:
        ensure_importacao_v2_modelos_table()
    except Exception as e:
        print(f"[importation_v2] ensure_table self-heal error: {e}")

    import io as _io
    content = await file.read()
    try:
        xls = pd.ExcelFile(_io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel inválido: {e}")

    # Obriga 1 aba — evita ambiguidade
    if len(xls.sheet_names) > 1:
        raise HTTPException(
            status_code=400,
            detail=f"A planilha tem {len(xls.sheet_names)} abas ({', '.join(xls.sheet_names)}). Deixe apenas 1 aba com os dados de MOQ."
        )
    sheet = xls.sheet_names[0]
    df = xls.parse(sheet)

    # Detecta colunas — normaliza removendo acentos e tudo que não é alfanumérico
    import unicodedata as _ud
    def _norm(s):
        s = _ud.normalize('NFKD', str(s)).encode('ascii', 'ignore').decode('ascii')
        return ''.join(ch for ch in s.lower() if ch.isalnum())
    def find_col(df, options):
        opts = {_norm(o) for o in options}
        for c in df.columns:
            if _norm(c) in opts:
                return c
        return None

    col_cod = find_col(df, {'itemno', 'codigo', 'codigoempresa', 'codprod', 'codproduto', 'codigoproduto'})
    col_moq = find_col(df, {'moq', 'lotemin', 'loteminimo', 'qtdmin', 'qtdminima', 'minorderquantity'})
    col_desc = find_col(df, {'description', 'descricao', 'descricaoproduto', 'descricaodoproduto'})
    # Colunas extras (planilha completa)
    col_barcode = find_col(df, {'barcodenumber', 'barcode', 'codbarras', 'codigodebarras', 'ean'})
    col_name_cn = find_col(df, {'name', 'nomecn', 'nomechines'})
    col_remark  = find_col(df, {'remark', 'remarks'})
    col_obs     = find_col(df, {'obs'})
    col_observ  = find_col(df, {'observacoes', 'observacao', 'observacoes1303', 'observacao1303'})
    col_engdesc = find_col(df, {'englishdescription', 'engdescription', 'descriptionen'})
    col_ctns    = find_col(df, {'ctns', 'cartons', 'cxs'})
    col_qty     = find_col(df, {'qty', 'quantity', 'quantidade'})
    col_amount  = find_col(df, {'amount', 'valortotal', 'total'})
    col_cbmtot  = find_col(df, {'cbmtotal'})
    col_tgw     = find_col(df, {'tgw', 'totalgw', 'totalgrossweight'})
    col_tnw     = find_col(df, {'tnw', 'totalnw', 'totalnetweight'})
    col_unit_ctn = find_col(df, {'unitctn', 'unitsctn', 'pcsctn'})
    col_cbm     = find_col(df, {'cbm'})
    col_gw      = find_col(df, {'gw', 'grossweight', 'pesobruto'})
    col_nw      = find_col(df, {'nw', 'netweight', 'pesoliquido'})
    col_l       = find_col(df, {'l', 'comprimento', 'length'})
    col_w       = find_col(df, {'w', 'largura', 'width'})
    col_h       = find_col(df, {'h', 'altura', 'height'})
    col_price   = find_col(df, {'uprice', 'price', 'preco', 'precounitario'})
    col_unit    = find_col(df, {'unit', 'unidade'})
    col_ncm     = find_col(df, {'ncm'})

    if not col_cod:
        raise HTTPException(status_code=400, detail=f"Coluna de código não encontrada na aba '{sheet}'. Colunas: {list(df.columns)}")
    if not col_moq:
        raise HTTPException(status_code=400, detail=f"Coluna MOQ não encontrada na aba '{sheet}'. Colunas: {list(df.columns)}")

    inseridos = 0; atualizados = 0; ignorados = 0
    erros: List[str] = []
    try:
        # Conexão única só para o SELECT inicial
        conn0 = get_db_connection(); cur0 = conn0.cursor()
        try:
            cur0.execute("SELECT codigo FROM importacao_v2_moq")
            existentes = {r[0] for r in cur0.fetchall()}
        finally:
            cur0.close(); conn0.close()

        # Resolve user_id: se uid não existir em users, usa NULL (evita FK violation)
        valid_uid = None
        if uid is not None:
            conn_u = get_db_connection(); cur_u = conn_u.cursor()
            try:
                cur_u.execute("SELECT id FROM users WHERE id = %s", (uid,))
                if cur_u.fetchone():
                    valid_uid = uid
            finally:
                cur_u.close(); conn_u.close()

        # Cada INSERT em sua própria conexão/transação (evita poisoning do Postgres)
        for _, row in df.iterrows():
            raw_cod = row[col_cod]
            if pd.isna(raw_cod):
                ignorados += 1; continue
            cod = str(raw_cod).strip()
            if cod.endswith('.0'): cod = cod[:-2]
            if not cod:
                ignorados += 1; continue
            try:
                moq_val = float(row[col_moq])
            except Exception:
                moq_val = 0.0
            if moq_val < 0 or (moq_val != moq_val):  # NaN guard
                moq_val = 0.0
            desc = None
            if col_desc and not pd.isna(row[col_desc]):
                desc = str(row[col_desc]).strip()[:180]

            # Helpers de extração robustos
            def _num(c):
                if c is None: return None
                v = row[c]
                if pd.isna(v): return None
                try: return float(v)
                except Exception: return None
            def _txt(c, maxlen=None):
                if c is None: return None
                v = row[c]
                if pd.isna(v): return None
                s = str(v).strip()
                if not s: return None
                if maxlen: s = s[:maxlen]
                return s

            vals = {
                'unit_ctn': _num(col_unit_ctn), 'cbm': _num(col_cbm),
                'gw': _num(col_gw), 'nw': _num(col_nw),
                'comprimento': _num(col_l), 'largura': _num(col_w), 'altura': _num(col_h),
                'price': _num(col_price), 'ncm': _txt(col_ncm, 40), 'unit': _txt(col_unit, 20),
                'barcode': _txt(col_barcode, 60), 'name_cn': _txt(col_name_cn, 255),
                'remark': _txt(col_remark), 'obs': _txt(col_obs), 'observacoes': _txt(col_observ),
                'english_description': _txt(col_engdesc, 255),
                'ctns': _num(col_ctns), 'qty': _num(col_qty), 'amount': _num(col_amount),
                'cbm_total': _num(col_cbmtot), 'tgw': _num(col_tgw), 'tnw': _num(col_tnw),
            }

            conn = get_db_connection(); cur = conn.cursor()
            try:
                cur.execute("""
                    INSERT INTO importacao_v2_moq
                      (codigo, descricao, moq, origem, updated_by, updated_at,
                       unit_ctn, cbm, gw, nw, comprimento, largura, altura, price, ncm, unit,
                       barcode, name_cn, remark, obs, observacoes, english_description,
                       ctns, qty, amount, cbm_total, tgw, tnw)
                    VALUES (%s, %s, %s, 'upload', %s, NOW(),
                            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                            %s, %s, %s, %s, %s, %s,
                            %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (codigo) DO UPDATE
                      SET descricao = COALESCE(EXCLUDED.descricao, importacao_v2_moq.descricao),
                          moq       = EXCLUDED.moq,
                          origem    = 'upload',
                          updated_by= EXCLUDED.updated_by,
                          updated_at= NOW(),
                          unit_ctn  = COALESCE(EXCLUDED.unit_ctn, importacao_v2_moq.unit_ctn),
                          cbm       = COALESCE(EXCLUDED.cbm, importacao_v2_moq.cbm),
                          gw        = COALESCE(EXCLUDED.gw, importacao_v2_moq.gw),
                          nw        = COALESCE(EXCLUDED.nw, importacao_v2_moq.nw),
                          comprimento = COALESCE(EXCLUDED.comprimento, importacao_v2_moq.comprimento),
                          largura   = COALESCE(EXCLUDED.largura, importacao_v2_moq.largura),
                          altura    = COALESCE(EXCLUDED.altura, importacao_v2_moq.altura),
                          price     = COALESCE(EXCLUDED.price, importacao_v2_moq.price),
                          ncm       = COALESCE(EXCLUDED.ncm, importacao_v2_moq.ncm),
                          unit      = COALESCE(EXCLUDED.unit, importacao_v2_moq.unit),
                          barcode             = COALESCE(EXCLUDED.barcode, importacao_v2_moq.barcode),
                          name_cn             = COALESCE(EXCLUDED.name_cn, importacao_v2_moq.name_cn),
                          remark              = COALESCE(EXCLUDED.remark, importacao_v2_moq.remark),
                          obs                 = COALESCE(EXCLUDED.obs, importacao_v2_moq.obs),
                          observacoes         = COALESCE(EXCLUDED.observacoes, importacao_v2_moq.observacoes),
                          english_description = COALESCE(EXCLUDED.english_description, importacao_v2_moq.english_description),
                          ctns                = COALESCE(EXCLUDED.ctns, importacao_v2_moq.ctns),
                          qty                 = COALESCE(EXCLUDED.qty, importacao_v2_moq.qty),
                          amount              = COALESCE(EXCLUDED.amount, importacao_v2_moq.amount),
                          cbm_total           = COALESCE(EXCLUDED.cbm_total, importacao_v2_moq.cbm_total),
                          tgw                 = COALESCE(EXCLUDED.tgw, importacao_v2_moq.tgw),
                          tnw                 = COALESCE(EXCLUDED.tnw, importacao_v2_moq.tnw)
                """, (cod, desc, moq_val, valid_uid,
                      vals['unit_ctn'], vals['cbm'], vals['gw'], vals['nw'],
                      vals['comprimento'], vals['largura'], vals['altura'],
                      vals['price'], vals['ncm'], vals['unit'],
                      vals['barcode'], vals['name_cn'], vals['remark'], vals['obs'],
                      vals['observacoes'], vals['english_description'],
                      vals['ctns'], vals['qty'], vals['amount'],
                      vals['cbm_total'], vals['tgw'], vals['tnw']))
                conn.commit()
                if cod in existentes:
                    atualizados += 1
                else:
                    inseridos += 1
                    existentes.add(cod)
            except Exception as row_err:
                conn.rollback()
                msg = f"{cod}: {type(row_err).__name__}: {row_err}"
                print(f"[importation_v2] upload skip {msg}")
                erros.append(msg[:200])
                ignorados += 1
            finally:
                cur.close(); conn.close()

        return {"ok": True, "inseridos": inseridos, "atualizados": atualizados, "ignorados": ignorados, "aba": sheet, "erros": erros[:10]}
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Erro no upload: {type(e).__name__}: {e}")


# =====================================================================
# Export / WhatsApp
# =====================================================================

def _gerar_xlsx_importacao_v2(parametros: Dict[str, Any], itens: List[Dict[str, Any]]) -> bytes:
    """Gera XLSX da análise de Importação v2 com estilo (header colorido, totais, status colorido)."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Análise de Ruptura"

    # Cores
    HEADER_FILL = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    THIN = Side(style="thin", color="CBD5E1")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    RUPT_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    ATEN_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    OK_FILL   = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")

    # Linha 1: título
    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value = "Importação · Análise de Ruptura"
    c.font = Font(name="Calibri", size=14, bold=True, color="6366F1")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    # Linha 2: metadados
    from datetime import datetime
    ws["A2"] = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(name="Calibri", size=9, italic=True, color="64748B")
    ws.merge_cells("A2:N2")

    # Resumo dos parâmetros (linha 3)
    p = parametros or {}
    resumo = (
        f"Janela: {p.get('qtd_meses', '?')}m ({p.get('modo', '?')})  ·  "
        f"Prazo padrão: {p.get('lead_time_default', '?')}d  ·  "
        f"Confiança: {(p.get('nivel_servico_default', 0) * 100):.0f}%  ·  "
        f"Sensibilidade: {p.get('threshold_sigma', '?')}σ"
    )
    ws["A3"] = resumo
    ws["A3"].font = Font(name="Calibri", size=9, color="475569")
    ws.merge_cells("A3:N3")

    # Linha 5: headers
    headers = ["Código", "Descrição", "Estoque", "Em Trânsito", "Prazo (meses)", "Confiança",
               "MOQ", "Venda/mês", "Variação", "Col. Segurança", "Quando Comprar", "Meses p/ Zerar",
               "Sugerido", "Comprar c/ MOQ", "Status", "Picos / Vales"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[5].height = 24

    # Larguras
    widths = [12, 36, 10, 12, 10, 11, 9, 11, 11, 13, 14, 12, 11, 14, 12, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Status → label + fill
    STATUS_MAP = {
        "RUPTURA": ("Ruptura", RUPT_FILL),
        "ATENCAO": ("Atenção", ATEN_FILL),
        "OK":      ("OK",      OK_FILL),
    }

    for idx, it in enumerate(itens, start=6):
        outliers_txt = ", ".join(
            f"{o.get('mes', '?')} {o.get('tipo', '?')}({o.get('valor', '?')})"
            for o in (it.get("outliers") or [])
        )
        status_key = it.get("status", "OK")
        status_label, status_fill = STATUS_MAP.get(status_key, ("—", None))
        row_vals = [
            it.get("codigo", ""),
            it.get("descricao", "") or "",
            it.get("estoque_disponivel", 0),
            it.get("pipeline", 0),
            it.get("lead_time", 0),
            it.get("nivel_servico", 0),
            it.get("moq", 0),
            round(float(it.get("consumo_mensal", 0)), 0),
            round(float(it.get("sigma_mensal", 0)), 2),
            it.get("estoque_seguranca", 0),
            it.get("ponto_reposicao", 0),
            round(float(it.get("cobertura_meses", 0)), 1),
            it.get("qtd_sugerida_pura", 0),
            it.get("qtd_sugerida", 0),
            status_label,
            outliers_txt,
        ]
        # col indices: 1 Cod | 2 Desc | 3 Est | 4 Pipe | 5 Prazo | 6 Conf | 7 MOQ |
        # 8 Venda/d | 9 Variação | 10 EstMin | 11 QdComprar | 12 DiasZerar | 13 Sugerido | 14 Comprar | 15 Status | 16 Picos
        for col, val in enumerate(row_vals, 1):
            cell = ws.cell(row=idx, column=col, value=val)
            cell.border = BORDER
            cell.font = Font(name="Calibri", size=9)
            if col == 1:
                cell.font = Font(name="Calibri", size=9, bold=True)
            if col in (3, 4, 5, 7, 8, 9, 10, 11, 12, 13, 14):
                cell.alignment = Alignment(horizontal="right")
                if col in (3, 4, 7, 8, 10, 11, 13, 14):
                    cell.number_format = "#,##0"
                elif col == 5:
                    cell.number_format = '0.0" m"'
                elif col == 6:
                    cell.number_format = "0%"
                elif col == 12:
                    cell.number_format = "0.0"
                elif col == 9:
                    cell.number_format = "#,##0.00"
            if col == 14:
                # Comprar c/ MOQ — bold indigo
                cell.font = Font(name="Calibri", size=9, bold=True, color="4F46E5")
            if col == 15 and status_fill is not None:
                cell.fill = status_fill
                cell.font = Font(name="Calibri", size=9, bold=True, color="334155")
                cell.alignment = Alignment(horizontal="center")

    # Freeze pane abaixo do header
    ws.freeze_panes = "A6"

    # Aba 2: parâmetros completos
    ws2 = wb.create_sheet("Parâmetros")
    ws2["A1"] = "Parâmetros da Análise"
    ws2["A1"].font = Font(name="Calibri", size=14, bold=True, color="6366F1")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 40
    r = 3
    for k, v in (parametros or {}).items():
        ws2.cell(row=r, column=1, value=str(k)).font = Font(bold=True)
        ws2.cell(row=r, column=2, value=str(v) if not isinstance(v, (dict, list)) else json.dumps(v, ensure_ascii=False))
        r += 1

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


class GerarXlsxBody(BaseModel):
    parametros: Dict[str, Any]
    itens: List[Dict[str, Any]]


@router.post("/gerar-xlsx")
def gerar_xlsx_endpoint(body: GerarXlsxBody, user_id: Optional[str] = Depends(get_user_id_from_session)):
    from fastapi.responses import StreamingResponse
    if not check_module_permission(user_id or "", MODULE_ID):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    import io
    xlsx_bytes = _gerar_xlsx_importacao_v2(body.parametros, body.itens)
    filename = f"Importacao_v2_{date.today().strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


class EnviarWhatsAppBody(BaseModel):
    numero: str
    parametros: Dict[str, Any]
    itens: List[Dict[str, Any]]
    mensagem: Optional[str] = None


@router.post("/enviar-whatsapp")
def enviar_whatsapp_endpoint(body: EnviarWhatsAppBody, user_id: Optional[str] = Depends(get_user_id_from_session)):
    if not check_module_permission(user_id or "", MODULE_ID, min_permission="can_edit"):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    import base64
    from modulo.whatsapp_config import enviar_arquivo_whatsapp

    xlsx_bytes = _gerar_xlsx_importacao_v2(body.parametros, body.itens)
    filename = f"Importacao_v2_{date.today().strftime('%Y-%m-%d')}.xlsx"
    data_b64 = base64.b64encode(xlsx_bytes).decode("ascii")

    rupturas = sum(1 for it in body.itens if it.get("status") == "RUPTURA")
    atencoes = sum(1 for it in body.itens if it.get("status") == "ATENCAO")
    oks = sum(1 for it in body.itens if it.get("status") == "OK")
    total = len(body.itens)
    caption_default = (
        f"*Importação · Análise de Ruptura*\n"
        f"_{date.today().strftime('%d/%m/%Y')}_\n\n"
        f"📦 Total: {total}\n"
        f"🔴 Em ruptura: {rupturas}\n"
        f"🟠 Em atenção: {atencoes}\n"
        f"🟢 OK: {oks}\n\n"
        f"Detalhes no arquivo anexo."
    )
    caption = body.mensagem or caption_default

    result = enviar_arquivo_whatsapp(
        user_id=user_id,
        numero=body.numero,
        origem="importacao_v2",
        referencia_id=None,
        caption=caption,
        filename=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        data_base64=data_b64,
    )
    return {"ok": True, "result": result}


@router.delete("/versoes/{versao_id}")
def delete_versao(versao_id: int, user_id: Optional[str] = Depends(get_user_id_from_session)):
    if not check_module_permission(user_id or "", MODULE_ID, min_permission="can_edit"):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    uid = _uid_int(user_id)  # pode ser None — persiste NULL no DB sem bloquear
    conn = get_db_connection(); cur = conn.cursor()
    # Apenas o dono pode excluir (super_user/ceo poderá futuramente — manter simples agora)
    cur.execute("DELETE FROM importacao_v2_versoes WHERE id = %s AND user_id = %s", (versao_id, uid))
    affected = cur.rowcount
    conn.commit(); cur.close(); conn.close()
    if affected == 0:
        raise HTTPException(status_code=404, detail="Versão não encontrada ou sem permissão para excluir.")
    return {"ok": True}


# =====================================================================
# RELATÓRIO — Conversão para dados dummy determinísticos (sem fontes externas)
# =====================================================================
# (a) Externas substituídas:
#     - fetch_vendas_mensais: BigQuery client.query(...).to_dataframe(...) →
#       gerador dummy.rng() determinístico. Ignora `client`.
#     - fetch_estoque: BigQuery client.query(...).to_dataframe(...) →
#       gerador dummy.rng() determinístico. Ignora `client`.
#     - /calculate: get_bq_client() agora é no-op (não levanta 500 sem
#       credenciais; client=None e segue com dummy). Postgres do app intacto.
#     (Não substituídas neste módulo, pois NÃO são chamadas externas próprias:
#      _load_sheet_parametros_importacao e get_bq_client vivem em
#      modulo/importation.py — fora do escopo deste arquivo; ver "não confirmados".)
#
# (b) Shape exato (preservado):
#     - fetch_vendas_mensais → DataFrame colunas: ['COD'(str), 'MES'(str 'YYYY-MM'), 'QTD'(float64)]
#       1 linha por (código × mês); cobre os 12 meses de 2026 (respeita lookback_meses).
#     - fetch_estoque → DataFrame colunas: ['COD'(str), 'DISPONIVEL'(float64)]; 1 linha por código.
#     Consumidores (calculate → calc_sku) usam r["COD"], r["MES"], r["QTD"],
#     df_estoque["COD"]/["DISPONIVEL"] — todos mantidos.
#
# (c) Teste real (cd backend && /c/Python312/python -c "..."):
#     fetch_vendas_mensais(None, ['10401085','10400044','10402210']):
#       VENDAS cols: ['COD','MES','QTD'] | QTD dtype float64
#       meses cobertos: 12 -> 2026-01 ... 2026-12 (todos 2026: True, 12 meses: True)
#       rows por cod: 12 cada | determinismo: True (2ª chamada idêntica)
#       amostra: 10401085 2026-01 387.0 / 2026-02 634.0 / 2026-03 440.0
#     fetch_estoque(None, mesmos cods):
#       ESTOQUE cols: ['COD','DISPONIVEL'] | {'10401085':1544.0,'10400044':1153.0,'10402210':2360.0}
#     lookback_meses=3 → ['2026-10','2026-11','2026-12'] (recorte correto)
#     import do módulo OK (router.prefix == '/importation-v2').
#
# (d) Não confirmados / fora de escopo:
#     - _load_sheet_parametros_importacao (gspread/Drive) e get_bq_client/_load_moq_map
#       lendo Excel (FILE_PARAMETROS) ainda apontam para fontes externas, mas estão
#       em modulo/importation.py ou dependem de arquivos locais; o alvo era APENAS
#       este arquivo. Aqui já tratados de forma resiliente (try/except → fallback
#       vazio / client=None), então /calculate funciona só com dummy.
# =====================================================================
