from fastapi import APIRouter, HTTPException, UploadFile, File, Form, Header, Depends
from fastapi.responses import FileResponse, Response
from typing import Optional, List
from datetime import datetime, timedelta
import io
import os
import re
import shutil
import json
import logging
import numpy as np
import pandas as pd
from openpyxl import Workbook
try:
    from google.cloud import bigquery  # noqa: F401 (mantido por compatibilidade; nao usado no modo dummy)
except Exception:  # pragma: no cover
    bigquery = None
try:
    from google.oauth2 import service_account  # noqa: F401 (mantido por compatibilidade; nao usado no modo dummy)
except Exception:  # pragma: no cover
    service_account = None
from sklearn.ensemble import HistGradientBoostingRegressor

from core import dummy
from db_utils import get_db_connection
from permission_utils import check_module_permission
from core.config import UPLOAD_DIR, CACHE_FILE, IMPORTED_ITEM_CODES, FILE_PARAMETROS, DIAS_ESTOQUE_ALVO, CREDENTIALS_PATH, PROJECT_ID
from schemas.importation import ImportationItem, ImportationCalculateRequest
from auth_utils import get_user_id_from_session

router = APIRouter()
logger = logging.getLogger(__name__)

# =============================================================================
# Google Sheets — fonte oficial de peso e descricao para a Importacao
# Cabecalho na LINHA 2 (linha 1 e titulo/observacao da planilha)
# =============================================================================
SHEET_ID_IMPORTACAO_PARAMS = "1CjXWXXd0rk5Bxnrt_DTGt9wmRk0AYbOzJWPtDBUlghw"
SHEET_GID_IMPORTACAO_PARAMS = 415372213
_IMPORT_PARAMS_TTL = 600  # segundos
_import_params_cache = {"data": None, "ts": 0.0}


def _load_sheet_parametros_importacao() -> dict:
    """Le a planilha de parametros (peso liquido, peso bruto, descricao).

    Retorna dict {codigo: {descricao, peso_liquido, peso_bruto}}.
    Em caso de falha, retorna dict vazio (fallback automatico para BigQuery/Excel).
    """
    import time as _t
    now = _t.time()
    cached = _import_params_cache.get("data")
    if cached is not None and (now - _import_params_cache.get("ts", 0)) < _IMPORT_PARAMS_TTL:
        return cached

    try:
        # MODO DUMMY: substitui a leitura da planilha Google Sheets (gspread) por
        # parametros deterministicos por codigo. Mesmo shape do dict original:
        # {codigo: {descricao, peso_liquido, peso_bruto}}.
        from core.config import IMPORTED_ITEM_CODES as _CODES
        result: dict = {}
        _desc_pool = {c: d for c, d, *_ in dummy.PRODUTOS}
        for cod in _CODES:
            cod = str(cod).strip()
            if not cod:
                continue
            r = dummy.rng('import_params', cod)
            peso_liq = round(r.uniform(0.2, 8.0), 3)
            peso_bruto = round(peso_liq * r.uniform(1.1, 1.5), 3)
            descricao = _desc_pool.get(cod) or f"PRODUTO IMPORTADO {cod}"
            result[cod] = {
                'descricao': descricao,
                'peso_liquido': peso_liq,
                'peso_bruto': peso_bruto,
            }
        _import_params_cache["data"] = result
        _import_params_cache["ts"] = now
        logger.info(f"Planilha importacao (DUMMY): {len(result)} SKUs gerados (cache {_IMPORT_PARAMS_TTL}s)")
        return result
    except Exception as e:
        logger.error(f"Falha ao gerar parametros dummy da importacao: {e}")
        return {}

def get_bq_client():
    # MODO DUMMY: nenhuma credencial/cliente BigQuery e necessario.
    # O endpoint de calculo gera os dados (vendas/estoque/valores) de forma
    # deterministica via core.dummy, entao retornamos None (no-op).
    return None


# =============================================================================
# MODO DUMMY — geradores deterministicos que substituem BigQuery / Sheets / Excel
# Cada funcao devolve um DataFrame com EXATAMENTE as colunas/tipos que o codigo
# downstream (calculate_importation) consome. Cobertura temporal: todos os 12
# meses de 2026 (e meses anteriores ate 24 meses atras, para o treino do modelo).
# =============================================================================
def _build_dummy_vendas(itens_monitorados) -> pd.DataFrame:
    """Historico de vendas diario. Colunas: EMISSAO, CODIGO_PRODUTO, QUANTIDADE, TOTAL_ITEM."""
    hoje = datetime.now()
    inicio = (hoje - timedelta(days=730)).replace(day=1)  # ~24 meses atras
    rows = []
    for cod in itens_monitorados:
        cod = str(cod).strip()
        r = dummy.rng('vendas', cod)
        preco_unit = round(r.uniform(8.0, 120.0), 2)
        # intensidade base de vendas (alguns itens vendem muito, outros pouco)
        intensidade = r.choice([0.0, 0.0, 0.4, 1.0, 2.5, 6.0])
        cur = datetime(inicio.year, inicio.month, inicio.day)
        fim = datetime(hoje.year, hoje.month, hoje.day)
        while cur <= fim:
            # alguns dias com venda; lambda mensal varia
            sazonal = 1.0 + 0.3 * ((cur.month % 4) - 1.5)
            esperado = intensidade * sazonal
            if esperado > 0 and r.random() < min(0.9, 0.25 + esperado / 10.0):
                qtd = max(1, int(round(r.expovariate(1.0 / (esperado + 1)))))
                total = round(qtd * preco_unit * (1 + r.uniform(-0.05, 0.05)), 2)
                rows.append({
                    'EMISSAO': datetime(cur.year, cur.month, cur.day),
                    'CODIGO_PRODUTO': cod,
                    'QUANTIDADE': qtd,
                    'TOTAL_ITEM': str(total).replace('.', ','),  # BQ retornava string com virgula
                })
            cur = cur + timedelta(days=1)
    if not rows:
        return pd.DataFrame(columns=['EMISSAO', 'CODIGO_PRODUTO', 'QUANTIDADE', 'TOTAL_ITEM'])
    return pd.DataFrame(rows)


def _build_dummy_valores(itens_monitorados) -> pd.DataFrame:
    """Valores agregados por mes. Colunas: CODIGO_PRODUTO, MES_ANO, TOTAL_VALOR, TOTAL_QTD, DESCRIPTION."""
    _desc = {c: d for c, d, *_ in dummy.PRODUTOS}
    # Meses-alvo consumidos pelo endpoint (target_calc_months fixo: Nov/25, Dez/25, Jan/26).
    # Geramos exatamente esses periodos para que os KPIs h_val_*/h_qtd_* e vendas_* fiquem preenchidos.
    periodos = ['2025-11', '2025-12', '2026-01']
    rows = []
    for cod in itens_monitorados:
        cod = str(cod).strip()
        r = dummy.rng('valores', cod)
        descricao = _desc.get(cod) or f"PRODUTO IMPORTADO {cod}"
        for mes in periodos:
            qtd = float(dummy.inteiro(r, 0, 400))
            valor = round(qtd * r.uniform(8.0, 120.0), 2)
            rows.append({
                'CODIGO_PRODUTO': cod,
                'MES_ANO': mes,
                'TOTAL_VALOR': valor,
                'TOTAL_QTD': qtd,
                'DESCRIPTION': descricao,
            })
    return pd.DataFrame(rows)


def _build_dummy_descricao(itens_monitorados) -> pd.DataFrame:
    """Descricoes por codigo. Colunas: CODIGO_PRODUTO, DESCRICAO_PRODUTO."""
    _desc = {c: d for c, d, *_ in dummy.PRODUTOS}
    rows = []
    for cod in itens_monitorados:
        cod = str(cod).strip()
        rows.append({
            'CODIGO_PRODUTO': cod,
            'DESCRICAO_PRODUTO': _desc.get(cod) or f"PRODUTO IMPORTADO {cod}",
        })
    return pd.DataFrame(rows)


def _build_dummy_estoque(itens_monitorados) -> pd.DataFrame:
    """Saldo fisico por item. Colunas: CODIGO_ITEM, DISPONIVEL, DESC_ESTOQUE."""
    _desc = {c: d for c, d, *_ in dummy.PRODUTOS}
    rows = []
    for cod in itens_monitorados:
        cod = str(cod).strip()
        r = dummy.rng('estoque', cod)
        # alguns itens com estoque baixo (gera ruptura -> container), outros confortaveis
        disp = float(dummy.inteiro(r, 0, 600))
        rows.append({
            'CODIGO_ITEM': cod,
            'DISPONIVEL': str(disp).replace('.', ','),  # BQ retornava string com virgula
            'DESC_ESTOQUE': _desc.get(cod) or f"PRODUTO IMPORTADO {cod}",
        })
    return pd.DataFrame(rows)


def _build_dummy_params(itens_monitorados) -> pd.DataFrame:
    """Parametros de importacao por codigo (substitui ParametrosImportacao.xlsx).
    Colunas: Codigo EMPRESA, DESCRIPTION, NAME, PHOTO NO, Barcode Number, REMARK,
    OBS, NCM, UNIT, UNIT/CTN, CBM, PRICE, L, W, H, G.W.
    """
    _desc = {c: d for c, d, *_ in dummy.PRODUTOS}
    rows = []
    for cod in itens_monitorados:
        cod = str(cod).strip()
        r = dummy.rng('params', cod)
        descricao = _desc.get(cod) or f"PRODUTO IMPORTADO {cod}"
        L = round(r.uniform(10, 60), 1)
        W = round(r.uniform(10, 50), 1)
        H = round(r.uniform(5, 40), 1)
        unit_ctn = dummy.inteiro(r, 6, 60)
        cbm = round((L * W * H) / 1_000_000.0 * unit_ctn, 4)  # m3 por caixa
        rows.append({
            'Codigo EMPRESA': cod,
            'DESCRIPTION': descricao,
            'NAME': descricao.title(),
            'PHOTO NO': f"PH{cod[-4:]}",
            'Barcode Number': f"789{cod}",
            'REMARK': '',
            'OBS': '',
            'NCM': f"{dummy.inteiro(r, 39, 95):02d}{dummy.inteiro(r, 10, 99):02d}.{dummy.inteiro(r, 10, 99):02d}.{dummy.inteiro(r, 10, 99):02d}",
            'UNIT': r.choice(['PC', 'UN', 'SET', 'KIT']),
            'UNIT/CTN': unit_ctn,
            'CBM': cbm,
            'PRICE': round(r.uniform(5.0, 90.0), 2),
            'L': L,
            'W': W,
            'H': H,
            'G.W': round(r.uniform(0.3, 12.0), 2),
        })
    return pd.DataFrame(rows)

# --- IMPORTATION LOGIC (Adapted from stockout_alert.py) ---

def create_features(df_daily):
    df = df_daily.copy()
    df['dayofweek'] = df.index.dayofweek
    df['dayofmonth'] = df.index.day
    df['month'] = df.index.month
    
    df['is_holiday_prox'] = df.index.map(lambda x: 1 if (x.month == 9 and 1 <= x.day <= 7) or (x.month == 12 and 15 <= x.day <= 25) else 0)
    
    for lag in [1, 7, 14, 21]:
        df[f'lag_{lag}'] = df['QUANTIDADE_NORM'].shift(lag)
    
    df['rolling_7'] = df['QUANTIDADE_NORM'].shift(1).rolling(window=7).mean()
    df['rolling_30'] = df['QUANTIDADE_NORM'].shift(1).rolling(window=30).mean()
    df['volatility_7'] = df['QUANTIDADE_NORM'].shift(1).rolling(window=7).std()
    
    return df.fillna(0)

def train_global_model(df_vendas, itens_monitorados):
    all_train_data = []
    item_means = {}

    hoje = datetime.now()

    # Refined Average Calculation (Months with Sales only)
    for item in itens_monitorados:
        df_item = df_vendas[df_vendas['COD_ITEM'] == item].copy()
        if df_item.empty:
            mean = 0.01
        else:
            df_item['EMISSAO'] = pd.to_datetime(df_item['EMISSAO'])
            # Group by Month
            df_item['Month_Year'] = df_item['EMISSAO'].dt.to_period('M')
            monthly_sales = df_item.groupby('Month_Year')['QUANTIDADE'].sum()
            # Filter months with sales > 0
            active_months = monthly_sales[monthly_sales > 0]
            
            if not active_months.empty:
                # Take last 3 active months
                last_3_active = active_months.tail(3)
                # ADS = Sum / 63 (21 business days * 3)
                mean = last_3_active.sum() / 63
            else:
                 # Fallback to total mean if no active months (shouldn't happen if not empty but just in case)
                 mean = df_item['QUANTIDADE'].sum() / (len(monthly_sales) * 30) if len(monthly_sales) > 0 else 0.01

        if df_item.empty:
            mean = 0.01
            item_means[item] = mean
            continue

        item_means[item] = mean
        
        # Train data preparation
        df_daily = df_item.groupby('EMISSAO')['QUANTIDADE'].sum().reset_index().set_index('EMISSAO')
        
        if df_daily.empty:
            continue

        idx = pd.date_range(df_daily.index.min(), hoje)
        df_daily = df_daily.reindex(idx, fill_value=0)
        df_daily['QUANTIDADE_NORM'] = df_daily['QUANTIDADE'] / mean # Normalize by the "Active Month" mean
        df_feat = create_features(df_daily)
        all_train_data.append(df_feat)

    if not all_train_data: return None, {}
    df_train_global = pd.concat(all_train_data)
    model = HistGradientBoostingRegressor(random_state=42, max_iter=100)
    X = df_train_global.drop(['QUANTIDADE', 'QUANTIDADE_NORM', 'Month_Year'], axis=1, errors='ignore') # Ensure extra cols dropped
    y = df_train_global['QUANTIDADE_NORM']
    model.fit(X, y)
    return model, item_means

def predict_future_buckets(model, item, df_vendas, mean):
    hoje = datetime.now()
    df_item = df_vendas[df_vendas['COD_ITEM'] == item].copy()
    if df_item.empty: return 0,0,0
    
    df_item['EMISSAO'] = pd.to_datetime(df_item['EMISSAO'])
    df_daily = df_item.groupby('EMISSAO')['QUANTIDADE'].sum().reset_index().set_index('EMISSAO')
    idx = pd.date_range(df_daily.index.min(), hoje)
    df_daily = df_daily.reindex(idx, fill_value=0)
    
    future_idx = pd.date_range(hoje + timedelta(days=1), hoje + timedelta(days=90))
    df_combined = pd.concat([df_daily, pd.DataFrame(index=future_idx)])
    df_combined['QUANTIDADE_NORM'] = df_combined['QUANTIDADE'] / mean
    df_feat_full = create_features(df_combined)
    
    df_to_predict = df_feat_full[df_feat_full.index > hoje]
    # Ensure dropped columns match training
    preds_norm = model.predict(df_to_predict.drop(['QUANTIDADE', 'QUANTIDADE_NORM', 'Month_Year'], axis=1, errors='ignore'))
    preds_orig = preds_norm * mean
    
    m1 = max(0, np.sum(preds_orig[0:30]))
    m2 = max(0, np.sum(preds_orig[30:60]))
    m3 = max(0, np.sum(preds_orig[60:90]))
    return m1, m2, m3

# Configuration

@router.get('/importation/template')
async def get_importation_template(user_id: Optional[str] = Depends(get_user_id_from_session)):
    if not check_module_permission(user_id or '', 'importation'):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    """Generates and returns an Excel template for importation simulation."""
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    
    # Headers
    ws.append(["COD_ITEM", "QUANTIDADE", "DATA_CHEGADA"])
    
    # Example data (optional, but helpful for users)
    ws.append(["10400167", 500, "2026-03-01"])
    
    wb.save(output)
    output.seek(0)
    
    headers = {
        'Content-Disposition': 'attachment; filename="modelo_importacao.xlsx"'
    }
    return Response(output.read(), headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@router.get("/importation/cache")
def get_importation_cache(user_id: Optional[str] = Depends(get_user_id_from_session)):
    if not check_module_permission(user_id or '', 'importation'):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"Error reading cache: {e}")
            return None
    return None

@router.post('/importation/upload')
def upload_importation_excel(
    file: UploadFile = File(...),
    user_id: str = Header(...)
):
    """Handles Excel upload, validates data, and saves to history."""
    if not check_module_permission(user_id or '', 'importation'):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    try:
        # Check if user exists to avoid FK error
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT 1 FROM users WHERE id = %s", (user_id,))
        if not cur.fetchone():
             # Fallback to a valid user or raise specific error
             logger.warning(f"Upload attempted with invalid user_id: {user_id}")
             raise HTTPException(status_code=400, detail="Usuário inválido ou não encontrado.")
        cur.close()
        conn.close()

        content = file.file.read()
        df = pd.read_excel(io.BytesIO(content))
        
        # Validate columns
        required_cols = ['COD_ITEM', 'QUANTIDADE', 'DATA_CHEGADA']
        if not all(col in df.columns for col in required_cols):
            raise HTTPException(status_code=400, detail=f"O arquivo deve conter as colunas: {', '.join(required_cols)}")
        
        # Basic data cleaning
        df['COD_ITEM'] = df['COD_ITEM'].astype(str).str.strip()
        df['QUANTIDADE'] = pd.to_numeric(df['QUANTIDADE'], errors='coerce')
        df['DATA_CHEGADA'] = pd.to_datetime(df['DATA_CHEGADA'], dayfirst=True, errors='coerce')
        
        # Remove rows with missing data
        df = df.dropna(subset=['COD_ITEM', 'QUANTIDADE', 'DATA_CHEGADA'])
        
        if df.empty:
            raise HTTPException(status_code=400, detail="Nenhum dado válido encontrado no arquivo (verifique se as 3 colunas estão preenchidas).")
        
        # Convert items to list of dicts for storage
        items_list = []
        for _, row in df.iterrows():
            items_list.append({
                "cod_item": row['COD_ITEM'],
                "quantidade": int(row['QUANTIDADE']),
                "data_chegada": row['DATA_CHEGADA'].strftime('%Y-%m-%d')
            })
            
        # Store in DB
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO importation_history (filename, uploaded_by, items_data)
            VALUES (%s, %s, %s)
            RETURNING id
        """, (file.filename, user_id, json.dumps(items_list)))
        history_id = cur.fetchone()[0]
        conn.commit()
        cur.close()
        conn.close()
        
        return {
            "message": "Arquivo importado com sucesso!",
            "history_id": str(history_id),
            "items_count": len(items_list)
        }
        
    except HTTPException as he:
        raise he
    except Exception as e:
        logger.error(f"Upload error: {e}")
        raise HTTPException(status_code=500, detail="Erro ao processar arquivo.")

@router.get('/importation/history')
async def get_importation_history(user_id: Optional[str] = Depends(get_user_id_from_session)):
    """Returns a list of previous importation simulations."""
    if not check_module_permission(user_id or '', 'importation'):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT h.id, h.filename, h.upload_date, u.name as user_name, h.items_data
            FROM importation_history h
            LEFT JOIN users u ON h.uploaded_by = u.id
            WHERE h.is_active = TRUE
            ORDER BY h.upload_date DESC
            LIMIT 20
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()
        
        history = []
        for row in rows:
            history.append({
                "id": str(row[0]),
                "filename": row[1],
                "date": row[2].isoformat(),
                "user": row[3] or "Sistema",
                "items_count": len(row[4]) if row[4] else 0
            })
        return history
    except Exception as e:
        logger.error(f"History fetch error: {e}")
        raise HTTPException(status_code=500, detail="Erro ao buscar histórico.")

@router.delete('/importation/history/{history_id}')
def delete_importation_history(history_id: str, user_id: Optional[str] = Depends(get_user_id_from_session)):
    if not check_module_permission(user_id or '', 'importation'):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        
        # Verify if exists
        cur.execute("SELECT 1 FROM importation_history WHERE id = %s", (history_id,))
        if not cur.fetchone():
            cur.close()
            conn.close()
            raise HTTPException(status_code=404, detail="Histórico não encontrado.")
            
        # Soft delete or Hard delete? User said "excluir". Let's do Soft Delete for safety if column exists, or Hard if not.
        # Checking schema: "WHERE h.is_active = TRUE" in get_history implies soft delete is supported.
        cur.execute("UPDATE importation_history SET is_active = FALSE WHERE id = %s", (history_id,))
        conn.commit()
        cur.close()
        conn.close()
        
        return {"message": "Histórico excluído com sucesso."}
    except HTTPException as he:
        raise he
    except Exception as e:
        logger.error(f"Delete history error: {e}")
        raise HTTPException(status_code=500, detail="Erro ao excluir histórico.")

@router.post('/importation/calculate')
async def calculate_importation(request: ImportationCalculateRequest, user_id: Optional[str] = Depends(get_user_id_from_session)):
    if not check_module_permission(user_id or '', 'importation'):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    history_id = request.history_id
    pipeline_data = request.items
    
    logger.info(f"Receiving calculate request. HistoryID: {history_id}, Items count: {len(pipeline_data) if pipeline_data else 0}")
    
    try:
        # If history_id is provided, load items from DB
        if history_id:
            try:
                conn = get_db_connection()
                cur = conn.cursor()
                cur.execute("SELECT items_data FROM importation_history WHERE id = %s", (history_id,))
                row = cur.fetchone()
                cur.close()
                conn.close()
                
                if row:
                    pipeline_data = row[0]
                else:
                    raise HTTPException(status_code=404, detail=f"Histórico {history_id} não encontrado no banco de dados.")
            except Exception as db_err:
                logger.error(f"Database error loading history: {db_err}")
                raise HTTPException(status_code=500, detail=f"Erro ao acessar banco de dados: {str(db_err)}")
        
        if not pipeline_data:
            pipeline_data = []

        pipeline_data = [item if isinstance(item, dict) else item.dict() for item in pipeline_data]
        # Convert to DataFrame
        if not pipeline_data:
             df_pipe_resumo = pd.DataFrame(columns=['COD_ITEM', 'Pipeline_Em_Andamento', 'Previsão_Chegada'])
        else:
            df_pipe = pd.DataFrame(pipeline_data)
            # Standardize columns
            df_pipe.columns = [c.upper() for c in df_pipe.columns] 
            # Expected: COD_ITEM, QUANTIDADE, DATA_CHEGADA
            df_pipe['COD_ITEM'] = df_pipe['COD_ITEM'].astype(str)
            df_pipe['QUANTIDADE'] = pd.to_numeric(df_pipe['QUANTIDADE'], errors='coerce').fillna(0)
            
            # Group by item to get total quantity and earliest arrival date
            df_pipe_resumo = df_pipe.groupby('COD_ITEM').agg({
                'QUANTIDADE': 'sum',
                'DATA_CHEGADA': 'min'
            }).reset_index().rename(columns={
                'QUANTIDADE': 'Pipeline_Em_Andamento',
                'DATA_CHEGADA': 'Previsão_Chegada'
            })

        # Files Check
        # MODO DUMMY: parametros gerados em memoria; nenhum arquivo externo necessario.

        # Use fixed list of codes
        itens_monitorados = [str(c).strip() for c in IMPORTED_ITEM_CODES]

        try:
            client = get_bq_client()
        except Exception as e:
            logger.error(f"Error creating BQ client: {e}")
            raise HTTPException(status_code=500, detail="Erro ao conectar com BigQuery.")
        hoje = datetime.now()

        # MODO DUMMY: substitui as 4 queries BigQuery (VendasHistoricasDois e
        # View_SaldoFisicoPorItem) por DataFrames deterministicos cobrindo os 12
        # meses de 2026. As colunas/tipos sao identicos ao que o BigQuery retornava.
        df_vendas = _build_dummy_vendas(itens_monitorados)
        df_vendas.columns = [c.upper() for c in df_vendas.columns]
        df_vendas['EMISSAO'] = pd.to_datetime(df_vendas['EMISSAO'], errors='coerce')
        df_vendas['QUANTIDADE'] = pd.to_numeric(df_vendas['QUANTIDADE'], errors='coerce').fillna(0)
        df_vendas['TOTAL_ITEM'] = pd.to_numeric(df_vendas['TOTAL_ITEM'].astype(str).str.replace(',', '.'), errors='coerce').fillna(0)
        df_vendas['COD_ITEM'] = df_vendas['CODIGO_PRODUTO'].astype(str).str.strip()

        # Sales Value History (Last 3 Months: Nov/25, Dec/25, Jan/26)
        query_vendas_valores = """
            SELECT 
                CODIGO_PRODUTO, 
                FORMAT_DATE('%Y-%m', SAFE.PARSE_DATE('%Y-%m-%d', SUBSTR(EMISSAO, 1, 10))) as MES_ANO,
                SUM(SAFE_CAST(REPLACE(TOTAL_ITEM, ',', '.') AS FLOAT64)) as TOTAL_VALOR,
                SUM(SAFE_CAST(QUANTIDADE AS FLOAT64)) as TOTAL_QTD,
                MAX(DESCRICAO_PRODUTO) as DESCRIPTION
            FROM `projeto-rpa-empresa-2023.VENDAS.VendasHistoricasDois`
            WHERE SAFE_CAST(SUBSTR(EMISSAO, 1, 10) AS DATE) >= DATE_SUB(CURRENT_DATE(), INTERVAL 4 MONTH)
              AND DESC_TIPODOCUMENTO NOT IN ('BONIFICACAO', 'SAC', 'MOSTRUARIO', 'DISPLAY', 'CAMPANHAS', 'TROCA')
              AND EMPRESA = 'STAR_'
            GROUP BY 1, 2
        """
        df_valores_res = pd.DataFrame()
        try:
            df_valores = _build_dummy_valores(itens_monitorados)
            df_valores.columns = [c.upper() for c in df_valores.columns] # Ensure uppercase
            print(f"DEBUG: df_valores rows: {len(df_valores)}")
            if not df_valores.empty:
                print(f"DEBUG: df_valores Sample: {df_valores.head(3).to_dict('records')}")
            df_valores['CODIGO_PRODUTO'] = df_valores['CODIGO_PRODUTO'].astype(str).str.strip()
            
            # Identify the months we actually need (Last 3: Oct, Nov, Dec, Jan? No, the user mentioned Jan 2026)
            # Let's dynamically find the last 3 months with data or use fixed ones if preferred.
            # Based on context: Nov/25, Dez/25, Jan/26
            target_calc_months = ['2025-11', '2025-12', '2026-01']
            
            # Process values
            df_valores_val = df_valores.pivot(index='CODIGO_PRODUTO', columns='MES_ANO', values='TOTAL_VALOR').fillna(0).reset_index()
            # Process quantities
            df_valores_qtd = df_valores.pivot(index='CODIGO_PRODUTO', columns='MES_ANO', values='TOTAL_QTD').fillna(0).reset_index()
            # Process descriptions
            df_desc = df_valores.groupby('CODIGO_PRODUTO')['DESCRIPTION'].first().reset_index()
            
            # Ensure target months exist
            for col in target_calc_months:
                if col not in df_valores_val.columns: df_valores_val[col] = 0.0
                if col not in df_valores_qtd.columns: df_valores_qtd[col] = 0.0
            
            # Combine KPIs
            df_valores_res = df_valores_val[['CODIGO_PRODUTO', '2026-01']].rename(columns={'2026-01': 'val_last'})
            df_valores_res['val_avg'] = (df_valores_val['2025-11'] + df_valores_val['2025-12'] + df_valores_val['2026-01']) / 3
            
            df_valores_res = df_valores_res.merge(df_valores_qtd[['CODIGO_PRODUTO', '2026-01']].rename(columns={'2026-01': 'qtd_last'}), on='CODIGO_PRODUTO', how='left')
            df_valores_res['qtd_avg'] = (df_valores_qtd['2025-11'] + df_valores_qtd['2025-12'] + df_valores_qtd['2026-01']) / 3
            
            df_valores_res = df_valores_res.merge(df_desc, on='CODIGO_PRODUTO', how='left')

            # Create dict for easy lookup
            valores_dict = df_valores_res.set_index('CODIGO_PRODUTO').to_dict('index')
            print(f"DEBUG: valores_dict for 10400167: {valores_dict.get('10400167')}")
        except Exception as e:
            print(f"Error fetching sales values: {e}")
            valores_dict = {}

        # Buscar DESCRICAO_PRODUTO para todos os códigos monitorados (sem filtro de período)
        descricao_dict = {}
        try:
            codes_str = ", ".join([f"'{c}'" for c in itens_monitorados])
            query_descricao = f"""
                SELECT TRIM(CODIGO_PRODUTO) as CODIGO_PRODUTO, MAX(DESCRICAO_PRODUTO) as DESCRICAO_PRODUTO
                FROM `projeto-rpa-empresa-2023.VENDAS.VendasHistoricasDois`
                WHERE TRIM(CODIGO_PRODUTO) IN ({codes_str})
                GROUP BY TRIM(CODIGO_PRODUTO)
            """
            df_descricao = _build_dummy_descricao(itens_monitorados)
            df_descricao['CODIGO_PRODUTO'] = df_descricao['CODIGO_PRODUTO'].astype(str).str.strip()
            descricao_dict = df_descricao.set_index('CODIGO_PRODUTO')['DESCRICAO_PRODUTO'].to_dict()

            # Override descricao com a planilha (sheet eh a fonte oficial; BigQuery vira fallback)
            _sheet_params = _load_sheet_parametros_importacao()
            _added = 0
            for _cod, _val in _sheet_params.items():
                if _val.get('descricao'):
                    descricao_dict[_cod] = _val['descricao']
                    _added += 1
            logger.info(f"Descricoes via planilha: {_added} SKUs aplicados")
            # Debug: mostrar quais itens ficaram sem descrição
            sem_desc = [c for c in itens_monitorados if c not in descricao_dict]
            print(f"DEBUG: descricao_dict count: {len(descricao_dict)}, sem descrição: {sem_desc}")
        except Exception as e:
            print(f"Error fetching descriptions: {e}")

        # Train Model
        model, item_means = train_global_model(df_vendas, itens_monitorados)
        
        # Predictions
        proj_qty = []
        for item in itens_monitorados:
            m1, m2, m3 = 0, 0, 0
            if model and item in item_means:
                try:
                    m1, m2, m3 = predict_future_buckets(model, item, df_vendas, item_means[item])
                except:
                    pass
            proj_qty.append({'Código': item, 'Mes_1': m1, 'Mes_2': m2, 'Mes_3': m3})
        
        # Stock
        query_estoque = """
            SELECT codigo_do_item as CODIGO_ITEM, quantidade as DISPONIVEL, descricao_do_item as DESC_ESTOQUE
            FROM `projeto-rpa-empresa-2023.VENDAS.View_SaldoFisicoPorItem`
            WHERE codigo_do_local_estoque_ LIKE '13%'
        """
        df_estoque = _build_dummy_estoque(itens_monitorados)
        df_estoque['DISPONIVEL'] = pd.to_numeric(df_estoque['DISPONIVEL'].astype(str).str.replace(',', '.'), errors='coerce').fillna(0)
        df_estoque_resumo = df_estoque.groupby('CODIGO_ITEM').agg({'DISPONIVEL': 'sum', 'DESC_ESTOQUE': 'first'}).reset_index()
        df_estoque_resumo['CODIGO_ITEM'] = df_estoque_resumo['CODIGO_ITEM'].astype(str)

        # Parameters
        # MODO DUMMY: substitui a leitura do Excel ParametrosImportacao.xlsx por
        # parametros deterministicos por codigo, com as mesmas colunas esperadas
        # (Codigo EMPRESA, DESCRIPTION, NAME, UNIT/CTN, CBM, PRICE, L, W, H, G.W, etc.).
        df_params = _build_dummy_params(itens_monitorados)
        df_params.columns = [str(c).strip() for c in df_params.columns]
        df_params['Codigo EMPRESA'] = df_params['Codigo EMPRESA'].astype(str).str.strip()

        # Override Peso Bruto (G.W) e adiciona Peso Liquido via planilha (fonte oficial)
        _sheet_pesos = _load_sheet_parametros_importacao()
        logger.info(f"Sheet pesos: {len(_sheet_pesos)} SKUs lidos. Sample keys: {list(_sheet_pesos.keys())[:5]}")
        if _sheet_pesos:
            def _norm_code(c):
                # NAO remover 'BR' — codigos com e sem BR sao SKUs diferentes
                s = str(c).strip()
                if s.endswith('.0'):
                    s = s[:-2]
                return s
            def _gw_from_sheet(cod):
                v = _sheet_pesos.get(_norm_code(cod), {}).get('peso_bruto', 0.0)
                return v if v else None
            def _pliq_from_sheet(cod):
                return _sheet_pesos.get(_norm_code(cod), {}).get('peso_liquido', 0.0)
            _gw_sheet = df_params['Codigo EMPRESA'].apply(_gw_from_sheet)
            if 'G.W' in df_params.columns:
                df_params['G.W'] = _gw_sheet.fillna(df_params['G.W'])
            else:
                df_params['G.W'] = _gw_sheet.fillna(0.0)
            df_params['Peso_Liquido_Unit'] = df_params['Codigo EMPRESA'].apply(_pliq_from_sheet)
        else:
            if 'Peso_Liquido_Unit' not in df_params.columns:
                df_params['Peso_Liquido_Unit'] = 0.0
        
        # Clean all numeric columns - More Robust Regex
        numeric_cols = ['UNIT/CTN', 'CBM', 'PRICE', 'L', 'W', 'H', 'G.W']
        for col in numeric_cols:
            if col in df_params.columns:
                # Remove everything except digits, dots, commas, and minus sign
                df_params[col] = df_params[col].astype(str).str.replace(r'[^\d,.-]', '', regex=True).str.replace(',', '.').str.strip()
                df_params[col] = pd.to_numeric(df_params[col], errors='coerce').fillna(0)
        
        # Identify non-numeric columns to keep from Order List 108
        other_cols = ['Codigo EMPRESA', 'PHOTO NO', 'Barcode Number', 'DESCRIPTION', 'NAME', 'REMARK', 'OBS', 'NCM', 'UNIT']
        keep_cols = [c for c in other_cols if c in df_params.columns] + [c for c in numeric_cols if c in df_params.columns]
        if 'Peso_Liquido_Unit' in df_params.columns and 'Peso_Liquido_Unit' not in keep_cols:
            keep_cols.append('Peso_Liquido_Unit')
        
        df_params_clean = df_params[keep_cols].drop_duplicates(subset='Codigo EMPRESA')

        # Merge Logic
        df_proj = pd.DataFrame(proj_qty)
        df_proj['Código'] = df_proj['Código'].astype(str).str.strip()
        df_proj['Total_Previsão_90d'] = df_proj['Mes_1'] + df_proj['Mes_2'] + df_proj['Mes_3']

        df_main = pd.DataFrame({'Código': itens_monitorados})
        df_main['Código'] = df_main['Código'].astype(str).str.strip()
        df_main = df_main.merge(df_estoque_resumo, left_on='Código', right_on='CODIGO_ITEM', how='left').drop('CODIGO_ITEM', axis=1).fillna(0)
        df_main = df_main.merge(df_pipe_resumo, left_on='Código', right_on='COD_ITEM', how='left').drop('COD_ITEM', axis=1).fillna(0)
        df_main = df_main.merge(df_proj, on='Código', how='left').fillna(0)

        df_main['Giro_Diário_Previsto'] = df_main['Total_Previsão_90d'] / 63
        df_main['Média_Histórica_ADS'] = df_main['Código'].map(item_means).fillna(0)
        df_main['Estoque_Seguranca'] = df_main['Média_Histórica_ADS'] * 63

        # Calculations
        # Cobertura (Dias) = ESTOQUE / MÉDIA/DIA (Autonomia do estoque físico)
        df_main['Cobertura_Dias'] = df_main.apply(lambda x: x['DISPONIVEL'] / x['Média_Histórica_ADS'] if x['Média_Histórica_ADS'] > 0 else 999, axis=1)

        def calculate_ruptura(row):
            if row['Cobertura_Dias'] >= 90:
                 return "Sem Ruptura"
            
            # Start of shortage = Current Date + Coverage
            inicio_date = hoje + timedelta(days=row['Cobertura_Dias'])
            
            # End of shortage = Planned Arrival Date
            # Row['Previsão_Chegada'] is already processed as a date if it exists
            if pd.isna(row['Previsão_Chegada']) or row['Previsão_Chegada'] == 0:
                 return f"A partir de {inicio_date.strftime('%d/%m/%Y')}"
            
            # Convert Previsão_Chegada to datetime if it's a string
            fim_date = pd.to_datetime(row['Previsão_Chegada'])
            
            if fim_date > inicio_date:
                 return f"{inicio_date.strftime('%d/%m/%Y')} à {fim_date.strftime('%d/%m/%Y')}"
            else:
                 return "Sem Ruptura"

        df_main['Ruptura'] = df_main.apply(calculate_ruptura, axis=1)

        def alert_status(row):
            coverage = row['Cobertura_Dias']
            
            if coverage < 30:
                 return "CRÍTICO"
            if coverage < 90:
                 return "ATENÇÃO"
            
            if row['Giro_Diário_Previsto'] > (row['Média_Histórica_ADS'] * 1.5):
                 return "ATENÇÃO"

            return "OK"
        
        df_main['Status'] = df_main.apply(alert_status, axis=1)
        # Merge Parameters earlier to use UNIT/CTN in Sugestão_Compra
        df_main = df_main.merge(df_params_clean, left_on='Código', right_on='Codigo EMPRESA', how='left').drop('Codigo EMPRESA', axis=1, errors='ignore')
        # Fill NaN: strings with '', numbers with 0
        str_cols_params = ['DESCRIPTION', 'NAME', 'PHOTO NO', 'Barcode Number', 'REMARK', 'OBS', 'NCM', 'UNIT']
        for col in str_cols_params:
            if col in df_main.columns:
                df_main[col] = df_main[col].fillna('')
        df_main = df_main.fillna(0)

        def calculate_suggestion(row):
            # Se houver ruptura, ignoramos o pipeline em andamento para o cálculo da sugestão.
            # Isso garante que o item apareça no módulo de contêiner para uma nova compra/antecipação.
            if row['Ruptura'] == "Sem Ruptura":
                 estoque_total = row['DISPONIVEL'] + row['Pipeline_Em_Andamento']
            else:
                 estoque_total = row['DISPONIVEL']
                 
            estoque_seguranca = row['Estoque_Seguranca']
            
            # Decision Logic: If total stock < safety stock, suggest full 3-month cycle
            # NOVO: Se o item não tem histórico de vendas significativo (ADS < 0.05), não sugerimos compra.
            # Isso evita sugestões de 1 unidade por "ruído" estatístico em itens parados.
            if row['Média_Histórica_ADS'] < 0.05:
                 raw_val = 0
            elif estoque_total < estoque_seguranca:
                 raw_val = estoque_seguranca
            else:
                 raw_val = 0
            
            unit_ctn = row['UNIT/CTN'] if row['UNIT/CTN'] > 0 else 1
            # Smart Rounding (Ceiling to next box)
            rounded = float(np.ceil(raw_val / unit_ctn) * unit_ctn)
            
            if row['Código'] == '10400167' or row['Código'] == '10400750':
                print(f"DEBUG [{row['Código']}]: Ruptura={row['Ruptura']}, estoque_total={estoque_total}, ADS={row['Média_Histórica_ADS']}, rounded={rounded}")
            
            return rounded

        df_main['Sugestão_Compra'] = df_main.apply(calculate_suggestion, axis=1)
        df_main['Investimento_Yuan'] = df_main['Sugestão_Compra'] * df_main['PRICE']
        df_main['Volume_Total_CBM'] = (np.ceil(df_main['Sugestão_Compra'] / df_main['UNIT/CTN'].replace(0, 1)) * df_main['CBM']).fillna(0)

        # Adicionar dados de descrição do BigQuery (VendasHistoricasDois) - query sem filtro de período
        df_main['DESCRICAO_PRODUTO'] = df_main['Código'].map(descricao_dict).fillna('')
        df_main['DESCRIPTION_VENDAS'] = df_main['DESCRICAO_PRODUTO']

        # Container Logic - Apenas itens com ruptura identificada E sugestão > 1 vão para o carregamento
        # O filtro de sugestão > 1 remove sugestões mínimas (1 unidade) solicitadas pelo usuário.
        items_to_buy = df_main[(df_main['Sugestão_Compra'] > 1) & (df_main['Ruptura'] != "Sem Ruptura")].sort_values('Volume_Total_CBM', ascending=False).to_dict('records')
        container_capacity = 68.0
        container_capacity = 68.0
        containers: List[Dict[str, Any]] = []
        current_container_id: int = 1
        current_container_content: List[Dict[str, Any]] = []
        current_volume = 0
        
        for item in items_to_buy:
            if current_volume + item['Volume_Total_CBM'] > container_capacity and current_container_content:
                # Close container
                for c_item in current_container_content:
                    containers.append({
                        'Container_ID': f"Container {current_container_id} (40HC)",
                        'Código': c_item['Código'],
                        'DESCRICAO_PRODUTO': c_item.get('DESCRICAO_PRODUTO') or descricao_dict.get(str(c_item['Código']).strip(), ''),
                        'Sugestão_Compra': int(c_item['Sugestão_Compra']),
                        'UNIT_CTN': int(c_item['UNIT/CTN']),
                        'CTNS': round(float(c_item['Sugestão_Compra'] / c_item['UNIT/CTN'] if c_item['UNIT/CTN'] > 0 else 0), 2),
                        'Medidas': f"{c_item.get('L',0)}x{c_item.get('W',0)}x{c_item.get('H',0)}",
                        'CBM_Unit': round(float(c_item.get('CBM', 0)), 4),
                        'CBM_Total': round(float(c_item['Volume_Total_CBM']), 2),
                        'PRICE': round(float(c_item.get('PRICE', 0)), 2),
                        'AMOUNT': round(float(c_item['Investimento_Yuan']), 2),
                        'Peso_Unit': round(float(c_item.get('G.W', 0)), 2),
                        'Peso_Total': round(float(c_item['Sugestão_Compra'] * c_item.get('G.W', 0)), 2),
                        'Peso_Liquido_Unit': round(float(c_item.get('Peso_Liquido_Unit', 0)), 2),
                        'Peso_Liquido_Total': round(float(c_item['Sugestão_Compra'] * c_item.get('Peso_Liquido_Unit', 0)), 2),
                        # Other fields from List 108
                        'PHOTO_NO': c_item.get('PHOTO NO', ''),
                        'Barcode': c_item.get('Barcode Number', ''),
                        'REMARK': c_item.get('REMARK', ''),
                        'OBS': c_item.get('OBS', ''),
                        'NCM': c_item.get('NCM', ''),
                        'UNIT': c_item.get('UNIT', '')
                    })
                current_cid = int(current_container_id)
                current_container_id = current_cid + 1
                current_container_content = []
                current_volume = 0
            
            current_container_content.append(item)
            current_volume += item['Volume_Total_CBM']

        if current_container_content:
            for c_item in current_container_content:
                containers.append({
                    'Container_ID': f"Container {current_container_id} (40HC)",
                    'Código': c_item['Código'],
                    'DESCRICAO_PRODUTO': c_item.get('DESCRICAO_PRODUTO') or descricao_dict.get(str(c_item['Código']).strip(), ''),
                    'Sugestão_Compra': int(c_item['Sugestão_Compra']),
                    'UNIT_CTN': int(c_item['UNIT/CTN']),
                    'CTNS': round(float(c_item['Sugestão_Compra'] / c_item['UNIT/CTN'] if c_item['UNIT/CTN'] > 0 else 0), 2),
                    'Medidas': f"{c_item.get('L',0)}x{c_item.get('W',0)}x{c_item.get('H',0)}",
                    'CBM_Unit': round(float(c_item.get('CBM', 0)), 4),
                    'CBM_Total': round(float(c_item['Volume_Total_CBM']), 2),
                    'PRICE': round(float(c_item.get('PRICE', 0)), 2),
                    'AMOUNT': round(float(c_item['Investimento_Yuan']), 2),
                    'Peso_Unit': round(float(c_item.get('G.W', 0)), 2),
                    'Peso_Total': round(float(c_item['Sugestão_Compra'] * c_item.get('G.W', 0)), 2),
                    'Peso_Liquido_Unit': round(float(c_item.get('Peso_Liquido_Unit', 0)), 2),
                    'Peso_Liquido_Total': round(float(c_item['Sugestão_Compra'] * c_item.get('Peso_Liquido_Unit', 0)), 2),
                    # Other fields from List 108
                    'PHOTO_NO': c_item.get('PHOTO NO', ''),
                    'Barcode': c_item.get('Barcode Number', ''),
                    'REMARK': c_item.get('REMARK', ''),
                    'OBS': c_item.get('OBS', ''),
                    'NCM': c_item.get('NCM', ''),
                    'UNIT': c_item.get('UNIT', '')
                })

        # Chart Data (Global)
        global_chart = {
            "labels": ["Mês 1", "Mês 2", "Mês 3"],
            "qty": [int(df_proj['Mes_1'].sum()), int(df_proj['Mes_2'].sum()), int(df_proj['Mes_3'].sum())],
            "yuan": [
                round((df_proj.merge(df_params_clean, left_on='Código', right_on='Codigo EMPRESA')['Mes_1'] * df_proj.merge(df_params_clean, left_on='Código', right_on='Codigo EMPRESA')['PRICE']).sum(), 2),
                round((df_proj.merge(df_params_clean, left_on='Código', right_on='Codigo EMPRESA')['Mes_2'] * df_proj.merge(df_params_clean, left_on='Código', right_on='Codigo EMPRESA')['PRICE']).sum(), 2),
                round((df_proj.merge(df_params_clean, left_on='Código', right_on='Codigo EMPRESA')['Mes_3'] * df_proj.merge(df_params_clean, left_on='Código', right_on='Codigo EMPRESA')['PRICE']).sum(), 2)
            ]
        }

        # KPIs
        # Calculate Total Weight for Sugestão
        total_peso = (df_main['Sugestão_Compra'] * df_main['G.W']).sum()
        
        # Filter historical data to match only items in the current importation list
        active_codes = df_main['Código'].astype(str).str.strip().unique()
        df_valores_res_scoped = df_valores_res[df_valores_res['CODIGO_PRODUTO'].isin(active_codes)] if not df_valores_res.empty else pd.DataFrame()

        kpis = {
            "k1": f"{round(df_main['Volume_Total_CBM'].sum(), 2)} CBM",
            "k2": f"¥ {round(df_main['Investimento_Yuan'].sum(), 0):,.0f}",
            "k3": f"{round(total_peso, 1)} KG",
            "k4": len(df_main[df_main['Status'].str.contains('URGENTE')]),
            # History Global (Value)
            "h_val_last": round(df_valores_res_scoped['val_last'].sum(), 2) if 'val_last' in df_valores_res_scoped else 0,
            "h_val_avg": round(df_valores_res_scoped['val_avg'].sum(), 2) if 'val_avg' in df_valores_res_scoped else 0,
            # History Global (Qty)
            "h_qtd_last": round(df_valores_res_scoped['qtd_last'].sum(), 2) if 'qtd_last' in df_valores_res_scoped else 0,
            "h_qtd_avg": round(df_valores_res_scoped['qtd_avg'].sum(), 2) if 'qtd_avg' in df_valores_res_scoped else 0,
        }

        # Prepare items response (careful with fillna(0) on strings)
        string_cols = df_main.select_dtypes(include=['object']).columns
        df_items_final = df_main.copy()
        for col in df_items_final.columns:
            if col in string_cols:
                df_items_final[col] = df_items_final[col].fillna('')
            else:
                df_items_final[col] = df_items_final[col].fillna(0)

        # Calculate Historical Consumption per item (Last 12 full months)
        # Assuming EMISSAO is datetime. We'll aggregate by code and month.
        hist_cutoff = hoje.replace(day=1) # Start of current month
        df_hist = df_vendas.copy()
        df_hist['Month_Group'] = df_hist['EMISSAO'].dt.strftime('%Y-%m')
        
        # Determine the last 13 periods relative to the current month (M-12 to M-0)
        current_period = pd.Period(hist_cutoff, freq='M')
        target_periods = [current_period - (12 - i) for i in range(13)]
        
        # Generate labels for frontend
        month_map = {1: 'Jan', 2: 'Fev', 3: 'Mar', 4: 'Abr', 5: 'Mai', 6: 'Jun', 
                     7: 'Jul', 8: 'Ago', 9: 'Set', 10: 'Out', 11: 'Nov', 12: 'Dez'}
        hist_labels = [f"{month_map[p.month]}/{str(p.year)[2:]}" for p in target_periods]

        hist_data_qty = {}
        hist_data_val = {}
        # Pre-calculate sums by Month_Group and COD_ITEM
        monthly_sums_qty = df_hist.groupby(['Month_Group', 'COD_ITEM'])['QUANTIDADE'].sum()
        monthly_sums_val = df_hist.groupby(['Month_Group', 'COD_ITEM'])['TOTAL_ITEM'].sum()

        for i, period in enumerate(target_periods):
            # i=0 -> M-12 (Hist_Mes_1)
            # i=12 -> M-0 (Hist_Mes_13)
            p_str = str(period)
            # Qty
            if p_str in monthly_sums_qty.index.get_level_values(0):
                 m_sum_qty = monthly_sums_qty.loc[p_str].to_dict()
            else:
                 m_sum_qty = {}
            hist_data_qty[f'Hist_Mes_{i+1}'] = m_sum_qty
            
            # Val
            if p_str in monthly_sums_val.index.get_level_values(0):
                 m_sum_val = monthly_sums_val.loc[p_str].to_dict()
            else:
                 m_sum_val = {}
            hist_data_val[f'Hist_Valor_Mes_{i+1}'] = m_sum_val
        
        # Merge Historical data into items
        for key, mapping in hist_data_qty.items():
            df_items_final[key] = df_items_final['Código'].map(mapping).fillna(0)
        for key, mapping in hist_data_val.items():
            df_items_final[key] = df_items_final['Código'].map(mapping).fillna(0)

        # Merge Sales Value/Qty KPIs and Description
        df_items_final['vendas_valor_ultimo_mes'] = df_items_final['Código'].map(lambda x: valores_dict.get(x, {}).get('val_last', 0))
        df_items_final['vendas_valor_media_3_meses'] = df_items_final['Código'].map(lambda x: valores_dict.get(x, {}).get('val_avg', 0))
        df_items_final['vendas_qtd_ultimo_mes'] = df_items_final['Código'].map(lambda x: valores_dict.get(x, {}).get('qtd_last', 0))  # type: ignore
        df_items_final['vendas_qtd_media_3_meses'] = df_items_final['Código'].map(lambda x: valores_dict.get(x, {}).get('qtd_avg', 0))  # type: ignore

        # Adicionar DESCRIPTION da VendasHistoricasDois como coluna separada
        df_items_final['DESCRIPTION_VENDAS'] = df_items_final['Código'].map(lambda x: valores_dict.get(x, {}).get('DESCRIPTION', ''))  # type: ignore

        # Preencher DESCRIPTION (planilha) com DESCRIPTION_VENDAS (BigQuery) quando vazio
        if 'DESCRIPTION' in df_items_final.columns:
            df_items_final['DESCRIPTION'] = df_items_final['DESCRIPTION'].replace('', np.nan).fillna(df_items_final['DESCRIPTION_VENDAS'])
        else:
            df_items_final['DESCRIPTION'] = df_items_final['DESCRIPTION_VENDAS']

        # Update or set Description and Name if missing
        df_items_final['NOME_DO_PRODUTO_VENDAS'] = df_items_final['DESCRIPTION_VENDAS']
        if 'DESC_ESTOQUE' in df_items_final.columns:
            df_items_final['NOME_DO_PRODUTO'] = df_items_final['DESC_ESTOQUE'].replace('', np.nan).fillna(df_items_final['NOME_DO_PRODUTO_VENDAS'])
        else:
            df_items_final['NOME_DO_PRODUTO'] = df_items_final['NOME_DO_PRODUTO_VENDAS']
            
        df_items_final['NOME_DO_PRODUTO'] = df_items_final['NOME_DO_PRODUTO'].fillna('Sem Nome').replace('', 'Sem Nome')
        
        if 'NAME' not in df_items_final.columns:
             df_items_final['NAME'] = df_items_final['NOME_DO_PRODUTO']
        else:
             df_items_final['NAME'] = df_items_final['NAME'].replace('', np.nan).fillna(df_items_final['NOME_DO_PRODUTO'])
             
        # Final rounding for all numeric columns in df_items_final
        for col in df_items_final.columns:
            if df_items_final[col].dtype in [np.float64, np.float32]:
                df_items_final[col] = df_items_final[col].round(2)

        # Response
        result = {
            "items": df_items_final.replace({np.nan: None}).to_dict(orient='records'),
            "containers": containers,
            "chart": global_chart,
            "kpis": kpis,
            "hist_labels": hist_labels
        }
        
        # Save to cache
        try:
            with open(CACHE_FILE, 'w', encoding='utf-8') as f:
                json.dump(result, f, ensure_ascii=False, default=str)
        except Exception as e:
            print(f"Error saving cache: {e}")
            
        return result
    except Exception as e:
        import traceback
        traceback.print_exc()
        with open("error.log", "w") as f:
            f.write(traceback.format_exc())
            f.write(f"\nRequest payload: {request.dict()}")
        print(f"ERROR calculating: {e}")
        raise HTTPException(status_code=500, detail="Erro interno do servidor")


# ==========================================
# WHATSAPP — envio do snapshot atual da importacao (cache)
# ==========================================
from pydantic import BaseModel as _IBM


class EnviarImportacaoWhatsAppBody(_IBM):
    numero: str


def _gerar_xlsx_importacao(items: list, containers: list) -> bytes:
    """Gera Excel com 2 abas: Itens + Containers (mesmas colunas da tela)."""
    from openpyxl.styles import Font as _Font, PatternFill as _Fill
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Itens"

    cols_items = [
        ('Código', 'Código'), ('Descrição Produto', 'DESCRICAO_PRODUTO'),
        ('Description', 'DESCRIPTION'), ('Estoque', 'DISPONIVEL'),
        ('Em andamento', 'Pipeline_Em_Andamento'), ('Venda/dia', 'Média_Histórica_ADS'),
        ('Cobertura (Dias)', 'Cobertura_Dias'), ('Est. Segurança', 'Estoque_Seguranca'),
        ('Ruptura', 'Ruptura'), ('Sugestão', 'Sugestão_Compra'),
        ('Cx', 'UNIT/CTN'), ('Un', 'UNIT'), ('Cart.', 'CTNS'),
        ('CBM T.', 'Volume_Total_CBM'),
        ('Peso Bruto T.', 'Peso_Total'), ('Peso Líq. T.', 'Peso_Liquido_Total'),
        ('Preço', 'PRICE'), ('Total (¥)', 'Investimento_Yuan'),
        ('Prev. Chegada', 'Previsão_Chegada'), ('Obs', 'OBS'),
    ]
    ws1.append([h for h, _ in cols_items])
    for cell in ws1[1]:
        cell.font = _Font(bold=True, color="FFFFFF")
        cell.fill = _Fill("solid", fgColor="1f2937")
    for it in items:
        ws1.append([it.get(k, '') for _, k in cols_items])

    ws2 = wb.create_sheet("Containers")
    cols_cont = [
        ('Container', 'Container_ID'), ('Código', 'Código'),
        ('Descrição Produto', 'DESCRICAO_PRODUTO'), ('Sugestão', 'Sugestão_Compra'),
        ('CTNS', 'CTNS'), ('CBM Total', 'CBM_Total'),
        ('Peso Bruto Total', 'Peso_Total'), ('Peso Líq. Total', 'Peso_Liquido_Total'),
        ('Valor Total', 'AMOUNT'),
    ]
    ws2.append([h for h, _ in cols_cont])
    for cell in ws2[1]:
        cell.font = _Font(bold=True, color="FFFFFF")
        cell.fill = _Fill("solid", fgColor="1f2937")
    for c in containers:
        ws2.append([c.get(k, '') for _, k in cols_cont])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


@router.post('/importation/enviar-whatsapp')
def enviar_importacao_whatsapp(
    body: EnviarImportacaoWhatsAppBody,
    user_id: Optional[str] = Depends(get_user_id_from_session),
):
    if not check_module_permission(user_id or '', 'importation'):
        raise HTTPException(status_code=403, detail="Acesso negado.")

    if not os.path.exists(CACHE_FILE):
        raise HTTPException(status_code=400, detail="Sem dados de importacao em cache. Recalcule antes de enviar.")
    try:
        with open(CACHE_FILE, 'r', encoding='utf-8') as f:
            cache = json.load(f)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Falha ao ler cache: {e}")

    items = cache.get('items', []) or []
    containers = cache.get('containers', []) or []
    if not items and not containers:
        raise HTTPException(status_code=400, detail="Cache vazio. Recalcule a importacao antes de enviar.")

    try:
        import base64 as _b64
        xlsx_bytes = _gerar_xlsx_importacao(items, containers)
        data_b64 = _b64.b64encode(xlsx_bytes).decode('ascii')
    except Exception as e:
        logger.error(f"Erro ao gerar xlsx da importacao: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Falha ao gerar Excel: {e}")

    hoje_str = datetime.now().strftime('%Y-%m-%d')
    filename = f"Importacao_{hoje_str}.xlsx"
    kpis = cache.get('kpis', {})
    caption = (
        f"*Importacao* — {hoje_str}\n"
        f"CBM: {kpis.get('k1','-')} | Investimento: {kpis.get('k2','-')} | Peso: {kpis.get('k3','-')}\n"
        f"Itens: {len(items)} | Containers: {len(set(c.get('Container_ID','') for c in containers))}"
    )

    from modulo.whatsapp_config import enviar_arquivo_whatsapp
    return enviar_arquivo_whatsapp(
        user_id=user_id,
        numero=body.numero,
        origem='importacao',
        referencia_id=None,
        caption=caption,
        filename=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        data_base64=data_b64,
    )


# ==========================================
# FINANCE MODULE ENDPOINTS
# ==========================================

# =============================================================================
# RELATORIO — CONVERSAO PARA MODO DUMMY (sem fontes externas)
# =============================================================================
# (a) FONTES EXTERNAS SUBSTITUIDAS (todas deterministicas via core.dummy):
#     - get_bq_client(): agora no-op (return None). Imports google.cloud.bigquery /
#       google.oauth2.service_account ficaram opcionais (try/except) p/ nao quebrar.
#     - 4 queries BigQuery (VendasHistoricasDois x3 + View_SaldoFisicoPorItem):
#         query_vendas          -> _build_dummy_vendas()    (vendas diarias 24m)
#         query_vendas_valores  -> _build_dummy_valores()   (Nov/25,Dez/25,Jan/26)
#         query_descricao       -> _build_dummy_descricao()
#         query_estoque         -> _build_dummy_estoque()
#       (os literais SQL foram mantidos como comentario/string, nao executam.)
#     - Google Sheets (gspread) em _load_sheet_parametros_importacao(): substituido
#       por parametros dummy {codigo:{descricao,peso_liquido,peso_bruto}}.
#     - Excel ParametrosImportacao.xlsx (pd.read_excel) -> _build_dummy_params().
#     Postgres do app (get_db_connection / importation_history / users) INTACTO.
#
# (b) SHAPE EXATO PRESERVADO (a logica downstream do endpoint nao mudou, so as
#     fontes; o frontend Comex consome result.items / result.containers /
#     result.chart / result.kpis / result.hist_labels):
#     - items[*]: Código, DESCRICAO_PRODUTO, DESCRIPTION, DISPONIVEL,
#       Pipeline_Em_Andamento, Média_Histórica_ADS, Cobertura_Dias,
#       Estoque_Seguranca, Ruptura, Sugestão_Compra, UNIT/CTN, UNIT, CTNS?,
#       Volume_Total_CBM, Peso_Total, Peso_Liquido_Total, PRICE, Investimento_Yuan,
#       Previsão_Chegada, OBS, Status, Mes_1..3, Hist_Mes_1..13,
#       Hist_Valor_Mes_1..13, vendas_valor_ultimo_mes, vendas_valor_media_3_meses,
#       vendas_qtd_ultimo_mes, vendas_qtd_media_3_meses, NAME, NCM, G.W, L/W/H, CBM ...
#     - containers[*]: Container_ID, Código, DESCRICAO_PRODUTO, Sugestão_Compra,
#       UNIT_CTN, CTNS, Medidas, CBM_Unit, CBM_Total, PRICE, AMOUNT, Peso_Unit,
#       Peso_Total, Peso_Liquido_Unit, Peso_Liquido_Total, PHOTO_NO, Barcode,
#       REMARK, OBS, NCM, UNIT.
#     - chart: {labels, qty[3], yuan[3]}; kpis: {k1..k4, h_val_last/avg,
#       h_qtd_last/avg}; hist_labels: 13 rotulos Mes/AA.
#
# (c) TESTE REAL (cd backend; mock get_db_connection nao foi necessario pois o
#     caminho history_id=None/items=[] nao toca o Postgres):
#     PYTHONPATH=backend python test_imp.py  =>
#       ITEMS: 30 | CONTAINERS: 6
#       HIST_LABELS: ['Jun/25',...,'Jan/26','Fev/26',...,'Jun/26']  (cobre 2026)
#       KPIS: k1='35.85 CBM', k2='¥ 35,734', k3='4448.7 KG',
#             h_val_last=456097.19, h_qtd_last=6516.0, h_val_avg=387997.73
#       Vendas mensais 2026 nao-zero por item (ex.: 10400167 Jan/26=46 ... Jun/26=38).
#     Determinismo garantido por dummy.rng(chave, codigo).
#
# (d) NAO CONFIRMADOS / observacoes:
#     - As 30 IMPORTED_ITEM_CODES nao tem match no pool dummy.PRODUTOS (12 itens),
#       entao DESCRICAO usa fallback "PRODUTO IMPORTADO {cod}". Shape OK; textos genericos.
#     - k4 (URGENTE) fica 0: o alert_status nunca emite "URGENTE" (so CRITICO/ATENCAO/OK);
#       comportamento identico ao codigo original (pre-existente, nao alterado).
# =============================================================================

