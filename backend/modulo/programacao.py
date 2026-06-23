"""Módulo Fábrica — Programação.

Aba 1 (Plano): usa a ÚLTIMA versão gerada do Otimizador de Produção
(tabela plano_producao_versoes) — sequência, produto, qtd a produzir,
estoque físico e saldo reserva já vêm de lá — e cruza com as OPs em
produção (view_ORDEM_PRODUCAO, SITUACAO_DA_OP_A = 8) por CODIGO_DO_MATERIAL.
"""
from fastapi import APIRouter, HTTPException, Depends
from typing import Optional, List
from pydantic import BaseModel
from datetime import datetime, timezone, timedelta
import os
import json
import time
import hashlib
import logging

from modulo import maquinas as _maquinas


def _iso_utc(dt):
    """Datetime do banco (UTC) -> ISO com offset, para o front converter ao fuso BR."""
    return dt.replace(tzinfo=timezone.utc).isoformat() if dt else None

from db_utils import get_db_connection
from permission_utils import check_module_permission, check_sector_permission
from auth_utils import get_user_id_from_session

router = APIRouter(prefix="/programacao", tags=["programacao"])
logger = logging.getLogger(__name__)

MODULE_ID = "programacao"

_BQ_KEY_FILE = os.path.join(os.path.dirname(__file__), "..", "projeto-rpa-blackd-2023-16b15891f73c.json")
_BQ_PROJECT = "projeto-rpa-blackd-2023"
_OP_TABELA = "projeto-rpa-blackd-2023.VIEW.view_ORDEM_PRODUCAO"


def _bq_client():
    from google.cloud import bigquery
    from google.oauth2 import service_account
    info_env = os.environ.get("GOOGLE_CREDENTIALS_JSON")
    if info_env:
        credentials = service_account.Credentials.from_service_account_info(json.loads(info_env))
    else:
        credentials = service_account.Credentials.from_service_account_file(os.path.normpath(_BQ_KEY_FILE))
    return bigquery.Client(credentials=credentials, project=_BQ_PROJECT)


def _uid(user_id, edit=False):
    if not user_id:
        raise HTTPException(status_code=401, detail="Não autenticado")
    lvl = "can_edit" if edit else "can_view"
    if not check_module_permission(user_id, MODULE_ID, lvl):
        raise HTTPException(status_code=403, detail="Sem permissão para Programação")
    return user_id


# ---- OPs em produção (cache em memória) ----
_OP_CACHE = {"data": None, "ts": 0.0}
_OP_TTL = 300


def carregar_ops(refresh: bool = False):
    agora = time.time()
    if not refresh and _OP_CACHE["data"] is not None and (agora - _OP_CACHE["ts"]) < _OP_TTL:
        return _OP_CACHE["data"]
    client = _bq_client()
    # Uma linha por OP (a view tem várias linhas por operação/recurso → dedup por NUMERO_DA_OP).
    sql = f"""
        SELECT NUMERO_DA_OP AS numero_op,
               ANY_VALUE(CAST(CODIGO_DO_MATERIAL AS STRING)) AS codigo,
               ANY_VALUE(DESCRICAO_DO_MATERIAL) AS descricao,
               ANY_VALUE(UNIDADE) AS unidade,
               ANY_VALUE(QUANTIDADE_PLANEJADA) AS qtd_op,
               ANY_VALUE(QUANTIDADE_APONTADA) AS apontada,
               ANY_VALUE(QUANTIDADE_NECESSARIA_A) AS qtd_nec_a,
               MAX(QUANTIDADE_NECESSARIA_B) AS qtd_nec_material,
               MIN(INICIO_REAL) AS inicio_real,
               ANY_VALUE(K01T_001_C) AS k01t_001,
               ANY_VALUE(K01T_002_C) AS k01t_002
        FROM `{_OP_TABELA}`
        WHERE SITUACAO_DA_OP_A = 8
        GROUP BY NUMERO_DA_OP
    """
    rows = client.query(sql).result()
    ops = []
    for r in rows:
        ini = r["inicio_real"]
        ops.append({
            "numero_op": (str(r["numero_op"]).strip() if r["numero_op"] is not None else ""),
            "codigo": (str(r["codigo"]).strip() if r["codigo"] is not None else ""),
            "descricao": (r["descricao"] or "").strip(),
            "unidade": (r["unidade"] or "").strip(),
            "qtd_op": float(r["qtd_op"]) if r["qtd_op"] is not None else None,
            "apontada": float(r["apontada"]) if r["apontada"] is not None else None,
            "qtd_nec_a": float(r["qtd_nec_a"]) if r["qtd_nec_a"] is not None else None,
            "qtd_nec_material": float(r["qtd_nec_material"]) if r["qtd_nec_material"] is not None else None,
            "inicio_real": ini.isoformat() if ini is not None else None,
            "k01t_001": (r["k01t_001"] or "").strip(),
            "k01t_002": (r["k01t_002"] or "").strip(),
        })
    _OP_CACHE["data"] = ops
    _OP_CACHE["ts"] = agora
    return ops


def _carregar_versao(cur, versao_id: str):
    cur.execute(
        "SELECT id, created_at, created_by_name, hoje, plano, oficial, oficial_em "
        "FROM plano_producao_versoes WHERE id = %s", (versao_id,)
    )
    row = cur.fetchone()
    if not row:
        return None
    plano = row[4]
    if isinstance(plano, str):
        plano = json.loads(plano)
    return {
        "plano": plano or [],
        "versao": {
            "id": str(row[0]),
            "created_at": _iso_utc(row[1]),
            "created_by_name": row[2],
            "hoje": row[3].isoformat() if row[3] else None,
            "oficial": bool(row[5]),
            "oficial_em": _iso_utc(row[6]),
        },
    }


@router.get("/plano")
def obter_plano(versao_id: Optional[str] = None, user_id: Optional[str] = Depends(get_user_id_from_session)):
    """Plano de uma versão específica do otimizador (não puxa mais a última automaticamente)."""
    _uid(user_id)
    if not versao_id:
        return {"plano": [], "versao": None}
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        data = _carregar_versao(cur, versao_id)
    finally:
        cur.close()
        conn.close()
    if not data:
        raise HTTPException(status_code=404, detail="Versão não encontrada")
    return data


@router.get("/versoes-oficiais")
def versoes_oficiais(user_id: Optional[str] = Depends(get_user_id_from_session)):
    """Lista as versões marcadas como oficiais, da mais nova para a mais antiga."""
    _uid(user_id)
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            "SELECT id, created_at, created_by_name, hoje, oficial_em, oficial_por_nome "
            "FROM plano_producao_versoes WHERE oficial = TRUE "
            "ORDER BY COALESCE(oficial_em, created_at) DESC"
        )
        versoes = [{
            "id": str(r[0]),
            "created_at": _iso_utc(r[1]),
            "created_by_name": r[2],
            "hoje": r[3].isoformat() if r[3] else None,
            "oficial_em": _iso_utc(r[4]),
            "oficial_por_nome": r[5],
        } for r in cur.fetchall()]
    finally:
        cur.close()
        conn.close()
    return {"versoes": versoes, "total": len(versoes)}


@router.get("/tempos-maquina")
def tempos_maquina(maquina_id: int, user_id: Optional[str] = Depends(get_user_id_from_session)):
    """Peças/hora cadastradas (Cadastro de Máquinas) por produto, para uma máquina."""
    _uid(user_id)
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT cod_item, pecas_hora FROM maquina_produto_tempo WHERE maquina_id = %s", (maquina_id,))
    tempos = {r[0]: (float(r[1]) if r[1] is not None else None) for r in cur.fetchall()}
    cur.close()
    conn.close()
    return {"maquina_id": maquina_id, "tempos": tempos}


@router.get("/oficiais-uso")
def oficiais_uso(user_id: Optional[str] = Depends(get_user_id_from_session)):
    """Versões oficiais e quantos itens de montagem (board) cada uma tem — para o otimizador
    avisar que a produção já está usando uma versão antes de trocar a oficial."""
    _uid(user_id)
    _ensure_board()
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT v.id, v.created_at, v.oficial_em, v.created_by_name, COUNT(b.id)
        FROM plano_producao_versoes v
        LEFT JOIN programacao_board b ON b.versao_id = v.id::text
        WHERE v.oficial = TRUE
        GROUP BY v.id, v.created_at, v.oficial_em, v.created_by_name
        ORDER BY COALESCE(v.oficial_em, v.created_at) DESC
        """
    )
    out = [{
        "id": str(r[0]),
        "created_at": _iso_utc(r[1]),
        "oficial_em": _iso_utc(r[2]),
        "created_by_name": r[3],
        "board_itens": int(r[4] or 0),
    } for r in cur.fetchall()]
    cur.close()
    conn.close()
    return {"oficiais": out}


@router.get("/comparar")
def comparar(base: str, novo: str, user_id: Optional[str] = Depends(get_user_id_from_session)):
    """Diff entre dois planos (por CODIGO_PRODUTO): novos, removidos, sequência e quantidade."""
    _uid(user_id)
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        a = _carregar_versao(cur, base)
        b = _carregar_versao(cur, novo)
    finally:
        cur.close()
        conn.close()
    if not a or not b:
        raise HTTPException(status_code=404, detail="Versão (base ou nova) não encontrada")

    def idx(plano):
        m = {}
        for r in plano:
            cod = str(r.get("CODIGO_PRODUTO", "")).strip()
            if cod:
                m[cod] = r
        return m

    ma, mb = idx(a["plano"]), idx(b["plano"])
    novos, removidos, seq_mudou, qtd_mudou = [], [], [], []
    for cod, rb in mb.items():
        if cod not in ma:
            novos.append({"codigo": cod, "descricao": rb.get("DESCRICAO", ""),
                          "sequencia_novo": rb.get("SEQUENCIA"), "qtd_novo": rb.get("QTD_PRODUZIR")})
    for cod, ra in ma.items():
        if cod not in mb:
            removidos.append({"codigo": cod, "descricao": ra.get("DESCRICAO", ""),
                              "sequencia_base": ra.get("SEQUENCIA"), "qtd_base": ra.get("QTD_PRODUZIR")})
            continue
        rb = mb[cod]
        if ra.get("SEQUENCIA") != rb.get("SEQUENCIA"):
            seq_mudou.append({"codigo": cod, "descricao": ra.get("DESCRICAO", ""),
                              "sequencia_base": ra.get("SEQUENCIA"), "sequencia_novo": rb.get("SEQUENCIA")})
        qa, qb = ra.get("QTD_PRODUZIR") or 0, rb.get("QTD_PRODUZIR") or 0
        if qa != qb:
            qtd_mudou.append({"codigo": cod, "descricao": ra.get("DESCRICAO", ""),
                              "qtd_base": qa, "qtd_novo": qb, "delta": qb - qa})
    return {
        "base": a["versao"], "novo": b["versao"],
        "novos": novos, "removidos": removidos, "seq_mudou": seq_mudou, "qtd_mudou": qtd_mudou,
        "resumo": {"novos": len(novos), "removidos": len(removidos), "seq": len(seq_mudou), "qtd": len(qtd_mudou)},
    }


class SugestoesBody(BaseModel):
    codigos: List[str]


class BoardItem(BaseModel):
    cod_item: str
    sequencia: Optional[int] = 0   # sequência/grupo do otimizador (para identificar duplicatas do mesmo produto)
    maquina_id: int
    ordem: int
    lote: Optional[int] = 0        # qual programação (lote) dentro da máquina
    qtd: Optional[float] = None   # quantidade a produzir (editável pelo usuário)
    observacao: Optional[str] = None
    ativo: Optional[bool] = True
    previsao_termino: Optional[str] = None  # ISO com offset (BR); manual ou calculada via peças/hora
    concluido: Optional[bool] = False       # produção do card concluída
    qtd_produzida: Optional[float] = None   # quantidade já produzida (informada pelo operador)


class MaqInicio(BaseModel):
    maquina_id: int
    data_inicio: Optional[str] = None


class LoteIn(BaseModel):
    maquina_id: int
    lote: int
    data_inicio: Optional[str] = None
    data_fim: Optional[str] = None
    ordem: Optional[int] = 0


class BoardBody(BaseModel):
    versao_id: str
    itens: List[BoardItem]
    maquinas: Optional[List[MaqInicio]] = []   # compat (datas por máquina, legado)
    lotes: Optional[List[LoteIn]] = []         # programações (lotes) por máquina, com data


class UsoComponenteIn(BaseModel):
    versao_id: str
    cod_item: str
    cod_componente: str
    descricao: Optional[str] = None
    tipo_comp: Optional[str] = None  # 'encarte' | 'taba'
    qtd_usar: Optional[float] = None
    maquina_id: Optional[int] = None  # máquina (Sopro) onde a taba é feita


def _ensure_board():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS programacao_board (
            id SERIAL PRIMARY KEY,
            versao_id TEXT NOT NULL,
            cod_item TEXT NOT NULL,
            sequencia INTEGER NOT NULL DEFAULT 0,
            maquina_id INTEGER NOT NULL,
            ordem INTEGER NOT NULL DEFAULT 0,
            updated_at TIMESTAMP DEFAULT NOW(),
            updated_by TEXT,
            UNIQUE (versao_id, cod_item, sequencia)
        )
        """
    )
    cur.execute("ALTER TABLE programacao_board ADD COLUMN IF NOT EXISTS qtd NUMERIC")
    cur.execute("ALTER TABLE programacao_board ADD COLUMN IF NOT EXISTS observacao TEXT")
    cur.execute("ALTER TABLE programacao_board ADD COLUMN IF NOT EXISTS ativo BOOLEAN DEFAULT TRUE")
    cur.execute("ALTER TABLE programacao_board ADD COLUMN IF NOT EXISTS lote INTEGER DEFAULT 0")
    cur.execute("ALTER TABLE programacao_board ADD COLUMN IF NOT EXISTS sequencia INTEGER DEFAULT 0")
    # Previsão de término (data/hora BR) por linha (produto x máquina x lote) — preenchida
    # manualmente OU calculada a partir da peças/hora do produto na máquina (maquinas/tempos).
    cur.execute("ALTER TABLE programacao_board ADD COLUMN IF NOT EXISTS previsao_termino TIMESTAMPTZ")
    cur.execute("ALTER TABLE programacao_board ADD COLUMN IF NOT EXISTS concluido BOOLEAN DEFAULT FALSE")
    cur.execute("ALTER TABLE programacao_board ADD COLUMN IF NOT EXISTS qtd_produzida NUMERIC")
    # Auto-cura de drift de schema (ambientes antigos): o upsert usa
    # ON CONFLICT (versao_id, cod_item, sequencia). Remove constraint única legada de
    # 2 colunas (versao_id, cod_item), que travaria duplicar o produto por sequência, e
    # garante o índice único de 3 colunas que o ON CONFLICT exige.
    cur.execute(
        """
        DO $$
        DECLARE r record;
        BEGIN
          FOR r IN
            SELECT conname FROM pg_constraint
            WHERE conrelid = 'programacao_board'::regclass AND contype = 'u'
              AND array_length(conkey, 1) = 2
          LOOP
            EXECUTE format('ALTER TABLE programacao_board DROP CONSTRAINT %I', r.conname);
          END LOOP;
        END $$;
        """
    )
    cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS ux_programacao_board_vcs ON programacao_board (versao_id, cod_item, sequencia)")
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS programacao_lote (
            id SERIAL PRIMARY KEY,
            versao_id TEXT NOT NULL,
            maquina_id INTEGER NOT NULL,
            lote INTEGER NOT NULL,
            data_inicio TEXT,
            data_fim TEXT,
            ordem INTEGER NOT NULL DEFAULT 0,
            updated_at TIMESTAMP DEFAULT NOW(),
            updated_by TEXT,
            UNIQUE (versao_id, maquina_id, lote)
        )
        """
    )
    # Auto-cura de drift: garante colunas usadas pelo INSERT de lotes em tabelas antigas.
    cur.execute("ALTER TABLE programacao_lote ADD COLUMN IF NOT EXISTS data_inicio TEXT")
    cur.execute("ALTER TABLE programacao_lote ADD COLUMN IF NOT EXISTS data_fim TEXT")
    cur.execute("ALTER TABLE programacao_lote ADD COLUMN IF NOT EXISTS ordem INTEGER DEFAULT 0")
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS programacao_maquina_inicio (
            id SERIAL PRIMARY KEY,
            versao_id TEXT NOT NULL,
            maquina_id INTEGER NOT NULL,
            data_inicio TEXT,
            updated_at TIMESTAMP DEFAULT NOW(),
            updated_by TEXT,
            UNIQUE (versao_id, maquina_id)
        )
        """
    )
    cur.execute("ALTER TABLE programacao_maquina_inicio ADD COLUMN IF NOT EXISTS data_inicio TEXT")
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS programacao_uso_componente (
            id SERIAL PRIMARY KEY,
            versao_id TEXT NOT NULL,
            cod_item TEXT NOT NULL,
            cod_componente TEXT NOT NULL,
            descricao TEXT,
            tipo_comp TEXT,
            qtd_usar NUMERIC,
            updated_at TIMESTAMP DEFAULT NOW(),
            updated_by TEXT,
            UNIQUE (versao_id, cod_item, cod_componente)
        )
        """
    )
    # Auto-cura de drift: colunas usadas pelo upsert de uso de componente em tabelas antigas.
    cur.execute("ALTER TABLE programacao_uso_componente ADD COLUMN IF NOT EXISTS descricao TEXT")
    cur.execute("ALTER TABLE programacao_uso_componente ADD COLUMN IF NOT EXISTS tipo_comp TEXT")
    cur.execute("ALTER TABLE programacao_uso_componente ADD COLUMN IF NOT EXISTS qtd_usar NUMERIC")
    # Máquina (Sopro) onde a taba é produzida — opcional, só faz sentido para tipo_comp='taba'.
    cur.execute("ALTER TABLE programacao_uso_componente ADD COLUMN IF NOT EXISTS maquina_id INTEGER")
    # Cards de taba no Sopro: 1 por (produto × sequência), gerados pela programação automática.
    # Tabela própria (não mistura com uso_componente, que serve o detalhe do item).
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS programacao_taba_sopro (
            id SERIAL PRIMARY KEY,
            versao_id TEXT NOT NULL,
            cod_item TEXT NOT NULL,
            sequencia INTEGER NOT NULL DEFAULT 0,
            cod_componente TEXT NOT NULL,
            descricao TEXT,
            qtd NUMERIC,
            maquina_id INTEGER,
            ordem INTEGER NOT NULL DEFAULT 0,
            lote INTEGER NOT NULL DEFAULT 0,
            inicio TIMESTAMPTZ,
            updated_at TIMESTAMP DEFAULT NOW(),
            UNIQUE (versao_id, cod_item, sequencia, cod_componente)
        )
        """
    )
    cur.execute("CREATE INDEX IF NOT EXISTS idx_taba_sopro_versao ON programacao_taba_sopro (versao_id)")
    cur.execute("ALTER TABLE programacao_taba_sopro ADD COLUMN IF NOT EXISTS inicio TIMESTAMPTZ")
    conn.commit()
    cur.close()
    conn.close()


# ============================================================================
# Versões salvas da Programação (snapshots) — save explícito da operação.
# Cada save congela o board atual (itens/lotes/componentes/peças-hora) de uma
# versão do plano. É o que o Otimizador de Faturamento consome e a base de
# comparação ("a programação mudou").
# ============================================================================

def _user_name(cur, user_id):
    if not user_id:
        return None
    try:
        cur.execute("SELECT name FROM users WHERE id = %s", (str(user_id),))
        r = cur.fetchone()
        return r[0] if r else None
    except Exception:
        return None


def _ensure_prog_versao():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS programacao_versao (
            id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
            plano_versao_id TEXT NOT NULL,
            created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            created_by TEXT,
            created_by_name TEXT,
            snapshot JSONB NOT NULL,
            hash TEXT NOT NULL,
            oficial BOOLEAN DEFAULT FALSE,
            oficial_em TIMESTAMPTZ,
            oficial_por TEXT,
            oficial_por_nome TEXT
        )
        """
    )
    cur.execute("CREATE INDEX IF NOT EXISTS idx_prog_versao_plano ON programacao_versao(plano_versao_id, created_at DESC)")
    # Fase C (16/06): lock real — só uma `programacao_versao` por `plano_versao_id` pode ser oficial.
    # Antes era lock aplicativo (UPDATE prévio) → race condition produzia múltiplas oficiais simultâneas.
    # Migração idempotente: limpa legado uma vez (backup + desmarca mais antigas) e cria UNIQUE INDEX.
    cur.execute("""
        DO $$
        BEGIN
            IF NOT EXISTS (SELECT 1 FROM pg_indexes WHERE indexname = 'idx_prog_versao_oficial_unique') THEN
                CREATE TABLE IF NOT EXISTS programacao_versao_oficial_cleanup_backup AS
                    SELECT *, NOW() AS backup_em FROM programacao_versao WHERE oficial = TRUE;
                UPDATE programacao_versao SET oficial = FALSE
                WHERE oficial = TRUE AND id NOT IN (
                    SELECT DISTINCT ON (plano_versao_id) id
                    FROM programacao_versao
                    WHERE oficial = TRUE
                    ORDER BY plano_versao_id, COALESCE(oficial_em, created_at) DESC
                );
                CREATE UNIQUE INDEX idx_prog_versao_oficial_unique
                    ON programacao_versao(plano_versao_id) WHERE oficial = TRUE;
            END IF;
        END $$;
    """)
    conn.commit()
    cur.close()
    conn.close()


def _build_prog_snapshot(cur, plano_versao_id: str) -> dict:
    """Lê o estado atual do board (itens/lotes/componentes/peças-hora) de uma versão do plano."""
    cur.execute(
        "SELECT cod_item, COALESCE(sequencia,0), maquina_id, ordem, COALESCE(lote,0), qtd, observacao, ativo, previsao_termino "
        "FROM programacao_board WHERE versao_id = %s ORDER BY maquina_id, lote, ordem", (plano_versao_id,)
    )
    # 17/06: filtra itens do backlog "A PROGRAMAR" (maquina_id=0/NULL) — não estão alocados
    # em máquina, não geram liberação. Operador edita no board live, mas não vai pro snapshot
    # oficial nem pra compute_liberacoes. Evita lib NULL silenciosa contaminando o Otimizador.
    itens = [{
        "cod_item": r[0], "sequencia": int(r[1] or 0), "maquina_id": r[2], "ordem": r[3],
        "lote": int(r[4] or 0), "qtd": (float(r[5]) if r[5] is not None else None),
        "observacao": r[6], "ativo": (True if r[7] is None else bool(r[7])),
        "previsao_termino": (r[8].isoformat() if r[8] is not None else None),
    } for r in cur.fetchall() if r[2] is not None and int(r[2] or 0) != 0]

    cur.execute(
        "SELECT maquina_id, lote, data_inicio, data_fim, ordem FROM programacao_lote WHERE versao_id = %s ORDER BY maquina_id, ordem, lote",
        (plano_versao_id,)
    )
    lotes = [{"maquina_id": r[0], "lote": r[1], "data_inicio": r[2], "data_fim": r[3], "ordem": (r[4] or 0)} for r in cur.fetchall()]

    cur.execute(
        "SELECT cod_item, cod_componente, descricao, tipo_comp, qtd_usar FROM programacao_uso_componente WHERE versao_id = %s",
        (plano_versao_id,)
    )
    componentes = [{"cod_item": r[0], "cod_componente": r[1], "descricao": r[2], "tipo_comp": r[3],
                    "qtd_usar": (float(r[4]) if r[4] is not None else None)} for r in cur.fetchall()]

    # peças/hora das máquinas envolvidas — congela o cálculo automático da previsão
    maq_ids = sorted({it["maquina_id"] for it in itens if it.get("maquina_id") is not None})
    tempos = []
    if maq_ids:
        cur.execute("SELECT maquina_id, cod_item, pecas_hora FROM maquina_produto_tempo WHERE maquina_id = ANY(%s)", (maq_ids,))
        tempos = [{"maquina_id": r[0], "cod_item": r[1], "pecas_hora": (float(r[2]) if r[2] is not None else None)} for r in cur.fetchall()]

    return {"plano_versao_id": str(plano_versao_id), "itens": itens, "lotes": lotes,
            "componentes": componentes, "tempos": tempos}


def _snapshot_hash(snapshot: dict) -> str:
    blob = json.dumps(snapshot, sort_keys=True, ensure_ascii=False, default=str)
    return hashlib.sha256(blob.encode("utf-8")).hexdigest()


def _auto_versao_silenciosa(cur, plano_versao_id: str, uid: Optional[int]) -> dict:
    """17/06: chamada após salvar_board pra congelar histórico automaticamente.
    Nunca marca oficial=TRUE — só cria versão se conteúdo mudou (dedup por hash).
    NÃO commita: caller controla a transação. Falha silenciosa (logger.warning)
    pra não bloquear o save do board."""
    try:
        snapshot = _build_prog_snapshot(cur, plano_versao_id)
        h = _snapshot_hash(snapshot)
        cur.execute(
            "SELECT id, hash FROM programacao_versao WHERE plano_versao_id = %s "
            "ORDER BY created_at DESC LIMIT 1",
            (plano_versao_id,)
        )
        ult = cur.fetchone()
        if ult and ult[1] == h:
            return {"criada": False, "vid": str(ult[0]), "hash": h, "motivo": "sem mudanças"}
        nome = _user_name(cur, uid) if uid else None
        cur.execute(
            "INSERT INTO programacao_versao (plano_versao_id, created_by, created_by_name, "
            "snapshot, hash, oficial) VALUES (%s, %s, %s, %s::jsonb, %s, FALSE) RETURNING id",
            (plano_versao_id, (str(uid) if uid else None), nome,
             json.dumps(snapshot, ensure_ascii=False, default=str), h),
        )
        vid = cur.fetchone()[0]
        return {"criada": True, "vid": str(vid), "hash": h, "n_itens": len(snapshot.get("itens") or [])}
    except Exception as _e:
        logger.warning(f"_auto_versao_silenciosa falhou ({_e})")
        return {"criada": False, "erro": str(_e)}


class SalvarVersaoBody(BaseModel):
    plano_versao_id: str
    oficial: Optional[bool] = False


class OficialVersaoBody(BaseModel):
    oficial: bool = True


@router.post("/salvar-versao")
def salvar_versao(body: SalvarVersaoBody, user_id: Optional[str] = Depends(get_user_id_from_session)):
    """Save explícito da operação: congela o board atual numa versão (snapshot) da Programação.
    Se nada mudou desde a última versão (mesmo hash), não duplica. Pode marcar como oficial."""
    uid = _uid(user_id, edit=True)
    if body.oficial and not check_sector_permission(user_id, 'Fábrica'):
        raise HTTPException(status_code=403, detail="Apenas usuários da Fábrica podem definir a versão oficial.")
    _ensure_board()
    _ensure_prog_versao()
    pid = (body.plano_versao_id or "").strip()
    if not pid:
        raise HTTPException(status_code=400, detail="plano_versao_id é obrigatório")
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        snapshot = _build_prog_snapshot(cur, pid)
        h = _snapshot_hash(snapshot)
        nome = _user_name(cur, uid)
        # dedup: se a última versão deste plano tem o mesmo conteúdo, não cria outra
        cur.execute(
            "SELECT id, hash, oficial FROM programacao_versao WHERE plano_versao_id = %s ORDER BY created_at DESC LIMIT 1",
            (pid,)
        )
        ult = cur.fetchone()
        if ult and ult[1] == h:
            vid = ult[0]
            if body.oficial and not ult[2]:
                cur.execute("UPDATE programacao_versao SET oficial = FALSE WHERE plano_versao_id = %s", (pid,))
                cur.execute(
                    "UPDATE programacao_versao SET oficial = TRUE, oficial_em = NOW(), oficial_por = %s, oficial_por_nome = %s WHERE id = %s",
                    (str(uid), nome, vid),
                )
                conn.commit()
            return {"id": str(vid), "criada": False, "oficial": bool(body.oficial or ult[2]),
                    "motivo": "sem mudanças desde a última versão", "hash": h}
        # cria nova versão
        if body.oficial:
            cur.execute("UPDATE programacao_versao SET oficial = FALSE WHERE plano_versao_id = %s", (pid,))
        cur.execute(
            "INSERT INTO programacao_versao (plano_versao_id, created_by, created_by_name, snapshot, hash, oficial) "
            "VALUES (%s, %s, %s, %s::jsonb, %s, %s) RETURNING id, created_at",
            (pid, str(uid), nome, json.dumps(snapshot, ensure_ascii=False, default=str), h, bool(body.oficial)),
        )
        row = cur.fetchone()
        vid = row[0]
        if body.oficial:
            cur.execute(
                "UPDATE programacao_versao SET oficial_em = NOW(), oficial_por = %s, oficial_por_nome = %s WHERE id = %s",
                (str(uid), nome, vid),
            )
        conn.commit()
        return {"id": str(vid), "criada": True, "oficial": bool(body.oficial),
                "created_at": (row[1].isoformat() if row[1] else None), "hash": h,
                "n_itens": len(snapshot["itens"])}
    finally:
        cur.close()
        conn.close()


@router.get("/versoes-salvas")
def versoes_salvas(plano_versao_id: Optional[str] = None, user_id: Optional[str] = Depends(get_user_id_from_session)):
    """Lista versões salvas (snapshots) da Programação, da mais nova para a mais antiga.
    Filtra por plano_versao_id quando informado."""
    _uid(user_id)
    _ensure_prog_versao()
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        base = ("SELECT id, plano_versao_id, created_at, created_by_name, oficial, oficial_em, oficial_por_nome, hash "
                "FROM programacao_versao")
        if plano_versao_id:
            cur.execute(base + " WHERE plano_versao_id = %s ORDER BY created_at DESC", (plano_versao_id,))
        else:
            cur.execute(base + " ORDER BY created_at DESC LIMIT 100")
        versoes = [{
            "id": str(r[0]), "plano_versao_id": r[1],
            "created_at": (r[2].isoformat() if r[2] else None),
            "created_by_name": r[3], "oficial": bool(r[4]),
            "oficial_em": (r[5].isoformat() if r[5] else None),
            "oficial_por_nome": r[6], "hash": r[7],
        } for r in cur.fetchall()]
    finally:
        cur.close()
        conn.close()
    return {"versoes": versoes, "total": len(versoes)}


@router.put("/versoes-salvas/{versao_id}/oficial")
def marcar_versao_oficial(versao_id: str, body: OficialVersaoBody, user_id: Optional[str] = Depends(get_user_id_from_session)):
    """Marca/desmarca uma versão salva da Programação como oficial. Só uma oficial por plano."""
    uid = _uid(user_id, edit=True)
    if not check_sector_permission(user_id, 'Fábrica'):
        raise HTTPException(status_code=403, detail="Apenas usuários da Fábrica podem definir a versão oficial.")
    _ensure_prog_versao()
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("SELECT plano_versao_id FROM programacao_versao WHERE id = %s", (versao_id,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Versão da programação não encontrada")
        pid = row[0]
        nome = _user_name(cur, uid)
        if body.oficial:
            cur.execute("UPDATE programacao_versao SET oficial = FALSE WHERE plano_versao_id = %s", (pid,))
            cur.execute(
                "UPDATE programacao_versao SET oficial = TRUE, oficial_em = NOW(), oficial_por = %s, oficial_por_nome = %s WHERE id = %s",
                (str(uid), nome, versao_id),
            )
        else:
            cur.execute(
                "UPDATE programacao_versao SET oficial = FALSE, oficial_em = NULL, oficial_por = NULL, oficial_por_nome = NULL WHERE id = %s",
                (versao_id,),
            )
        conn.commit()
    finally:
        cur.close()
        conn.close()
    return {"ok": True, "oficial": body.oficial}


@router.post("/versoes-salvas/{versao_id}/restaurar")
def restaurar_versao(versao_id: str, user_id: Optional[str] = Depends(get_user_id_from_session)):
    """Restaura uma versão salva (snapshot) de volta para o quadro atual da Programação:
    sobrescreve o board (itens), os lotes e os componentes da versão do plano com o conteúdo
    congelado naquele snapshot. Operação destrutiva sobre a montagem atual."""
    uid = _uid(user_id, edit=True)
    _ensure_board()
    _ensure_prog_versao()
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("SELECT plano_versao_id, snapshot FROM programacao_versao WHERE id = %s", (versao_id,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Versão da programação não encontrada")
        pid = row[0]
        snap = row[1]
        if isinstance(snap, str):
            snap = json.loads(snap)
        snap = snap or {}
        itens = snap.get("itens", []) or []
        lotes = snap.get("lotes", []) or []
        componentes = snap.get("componentes", []) or []
        u = str(uid) if uid else None

        # itens -> board (substitui tudo da versão)
        cur.execute("DELETE FROM programacao_board WHERE versao_id = %s", (pid,))
        for it in itens:
            cod = (str(it.get("cod_item") or "")).strip()
            if not cod:
                continue
            cur.execute(
                "INSERT INTO programacao_board (versao_id, cod_item, sequencia, maquina_id, ordem, lote, qtd, observacao, ativo, previsao_termino, updated_by) "
                "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) ON CONFLICT (versao_id, cod_item, sequencia) DO UPDATE SET "
                "maquina_id = EXCLUDED.maquina_id, ordem = EXCLUDED.ordem, lote = EXCLUDED.lote, qtd = EXCLUDED.qtd, "
                "observacao = EXCLUDED.observacao, ativo = EXCLUDED.ativo, previsao_termino = EXCLUDED.previsao_termino, "
                "updated_at = NOW(), updated_by = EXCLUDED.updated_by",
                (pid, cod, int(it.get("sequencia") or 0), it.get("maquina_id"), it.get("ordem"),
                 int(it.get("lote") or 0), it.get("qtd"), it.get("observacao"),
                 (True if it.get("ativo") is None else bool(it.get("ativo"))), (it.get("previsao_termino") or None), u),
            )

        # lotes -> programacao_lote (substitui tudo da versão)
        cur.execute("DELETE FROM programacao_lote WHERE versao_id = %s", (pid,))
        for lt in lotes:
            cur.execute(
                "INSERT INTO programacao_lote (versao_id, maquina_id, lote, data_inicio, data_fim, ordem, updated_by) "
                "VALUES (%s, %s, %s, %s, %s, %s, %s) ON CONFLICT (versao_id, maquina_id, lote) DO UPDATE SET "
                "data_inicio = EXCLUDED.data_inicio, data_fim = EXCLUDED.data_fim, ordem = EXCLUDED.ordem, updated_at = NOW(), updated_by = EXCLUDED.updated_by",
                (pid, lt.get("maquina_id"), lt.get("lote"), (lt.get("data_inicio") or None), (lt.get("data_fim") or None), (lt.get("ordem") or 0), u),
            )

        # componentes -> programacao_uso_componente (substitui tudo da versão)
        cur.execute("DELETE FROM programacao_uso_componente WHERE versao_id = %s", (pid,))
        for c in componentes:
            cod = (str(c.get("cod_item") or "")).strip()
            comp = (str(c.get("cod_componente") or "")).strip()
            if not cod or not comp:
                continue
            cur.execute(
                "INSERT INTO programacao_uso_componente (versao_id, cod_item, cod_componente, descricao, tipo_comp, qtd_usar) "
                "VALUES (%s, %s, %s, %s, %s, %s) ON CONFLICT (versao_id, cod_item, cod_componente) DO UPDATE SET "
                "descricao = EXCLUDED.descricao, tipo_comp = EXCLUDED.tipo_comp, qtd_usar = EXCLUDED.qtd_usar",
                (pid, cod, comp, c.get("descricao"), c.get("tipo_comp"), c.get("qtd_usar")),
            )

        conn.commit()
    finally:
        cur.close()
        conn.close()
    return {"ok": True, "plano_versao_id": str(pid), "n_itens": len(itens)}


# ============================================================================
# Liberações por produto (split por máquina) — base do Otimizador de Faturamento.
# Mesmo produto produzido em N máquinas => N liberações (cada uma com sua data de
# término efetiva e quantidade). É a representação autoritativa que o faturamento
# consome (1 linha lógica por produto, várias liberações).
# ============================================================================

_BR_TZ = timezone(timedelta(hours=-3))


def _parse_dt_flex(s):
    """Parse flexível: 'YYYY-MM-DD', 'YYYY-MM-DDTHH:MM[:SS]', ISO com offset/Z.
    Naive assume horário do Brasil (UTC-3). Retorna datetime aware ou None."""
    if not s:
        return None
    s = str(s).strip().replace(' ', 'T').replace('Z', '+00:00')
    for cand in (s, s + ':00'):
        try:
            dt = datetime.fromisoformat(cand)
            return dt if dt.tzinfo else dt.replace(tzinfo=_BR_TZ)
        except Exception:
            continue
    try:
        dt = datetime.fromisoformat(s.split('T')[0])
        return dt.replace(tzinfo=_BR_TZ)
    except Exception:
        return None


def compute_liberacoes(snapshot: dict) -> dict:
    """Deriva, por produto, as liberações (data de término efetiva + qtd + máquina) de um
    snapshot da Programação. Previsão MANUAL sobrescreve; senão calcula via peças/hora
    (24h contínuo, item independente): término = início do lote + qtd / peças_hora."""
    itens = snapshot.get("itens", []) or []
    lotes = snapshot.get("lotes", []) or []
    tempos = snapshot.get("tempos", []) or []
    lotes_by = {(l.get("maquina_id"), int(l.get("lote") or 0)): l for l in lotes}
    ph_by = {(t.get("maquina_id"), str(t.get("cod_item"))): t.get("pecas_hora") for t in tempos}
    produtos = {}
    for it in itens:
        if it.get("ativo") is False:
            continue
        cod = str(it.get("cod_item"))
        maq = it.get("maquina_id")
        lote = int(it.get("lote") or 0)
        qtd = float(it.get("qtd") or 0)
        lt = lotes_by.get((maq, lote))
        previsao = it.get("previsao_termino")  # manual (override) — término de produção (início+peças/hora)
        auto = False
        if not previsao:
            ph = ph_by.get((maq, cod))
            inicio = _parse_dt_flex(lt.get("data_inicio")) if lt else None
            if ph and float(ph) > 0 and inicio and qtd > 0:
                previsao = (inicio + timedelta(hours=qtd / float(ph))).isoformat()
                auto = True
        # data de ENTREGA (usada pelo faturamento): override manual OU "Entrega" (data_fim) do lote
        data_fim = lt.get("data_fim") if lt else None
        entrega = it.get("previsao_termino") or data_fim or None
        p = produtos.setdefault(cod, {"cod_item": cod, "qtd_total": 0.0, "liberacoes": [], "entrega": None})
        p["qtd_total"] += qtd
        p["liberacoes"].append({
            "maquina_id": maq, "lote": lote, "qtd": qtd,
            "previsao_termino": previsao, "entrega": entrega, "auto": auto, "sem_previsao": previsao is None,
        })
    for p in produtos.values():
        ents = [l["entrega"] for l in p["liberacoes"] if l.get("entrega")]
        p["entrega"] = max(ents, key=lambda s: (_parse_dt_flex(s) or datetime.min.replace(tzinfo=_BR_TZ))) if ents else None
        p["liberacoes"].sort(key=lambda x: (x["previsao_termino"] is None, x["previsao_termino"] or ""))
    return {"produtos": produtos, "total_produtos": len(produtos)}


@router.get("/liberacoes")
def liberacoes(plano_versao_id: Optional[str] = None, versao_salva_id: Optional[str] = None,
               user_id: Optional[str] = Depends(get_user_id_from_session)):
    """Liberações por produto (split por máquina): data de término efetiva + qtd por máquina.
    Use `versao_salva_id` (snapshot congelado) OU `plano_versao_id` (board ao vivo)."""
    _uid(user_id)
    _ensure_board()
    _ensure_prog_versao()
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        if versao_salva_id:
            cur.execute("SELECT snapshot FROM programacao_versao WHERE id = %s", (versao_salva_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Versão salva da programação não encontrada")
            snap = row[0]
            if isinstance(snap, str):
                snap = json.loads(snap)
        elif plano_versao_id:
            snap = _build_prog_snapshot(cur, plano_versao_id)
        else:
            raise HTTPException(status_code=400, detail="Informe versao_salva_id ou plano_versao_id")
    finally:
        cur.close()
        conn.close()
    return compute_liberacoes(snap)


@router.post("/sugestoes")
def sugestoes(body: SugestoesBody, user_id: Optional[str] = Depends(get_user_id_from_session)):
    """Para os códigos do plano: lista de máquinas (ativas) e a(s) máquina(s) sugerida(s) por código."""
    _uid(user_id)
    # máquinas ativas
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT id, nome, cor FROM maquinas WHERE ativo = TRUE ORDER BY nome")
    maquinas_lst = [{"id": r[0], "nome": r[1], "cor": r[2]} for r in cur.fetchall()]
    cur.close()
    conn.close()
    # resolução produto -> máquinas (regras + exceções)
    try:
        _, maquinas_por_produto, _ = _maquinas.resolver_associacoes()
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Erro ao resolver máquinas: {e}")
    pedidos = set(str(c).strip() for c in (body.codigos or []))
    por_codigo = {c: maquinas_por_produto.get(c, []) for c in pedidos}
    return {"maquinas": maquinas_lst, "por_codigo": por_codigo}


@router.get("/board")
def obter_board(versao_id: str, user_id: Optional[str] = Depends(get_user_id_from_session)):
    """Montagem salva (compartilhada) de uma versão: item -> máquina + ordem."""
    _uid(user_id)
    _ensure_board()
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT cod_item, maquina_id, ordem, qtd, observacao, ativo, COALESCE(lote, 0), COALESCE(sequencia, 0), previsao_termino, concluido, qtd_produzida FROM programacao_board WHERE versao_id = %s ORDER BY maquina_id, lote, ordem", (versao_id,))
    itens = [{"cod_item": r[0], "maquina_id": r[1], "ordem": r[2],
              "qtd": (float(r[3]) if r[3] is not None else None),
              "observacao": r[4], "ativo": (True if r[5] is None else bool(r[5])),
              "lote": (int(r[6]) if r[6] is not None else 0),
              "sequencia": (int(r[7]) if r[7] is not None else 0),
              "previsao_termino": (r[8].isoformat() if r[8] is not None else None),
              "concluido": bool(r[9]) if r[9] is not None else False,
              "qtd_produzida": (float(r[10]) if r[10] is not None else None)} for r in cur.fetchall()]
    cur.execute("SELECT maquina_id, data_inicio FROM programacao_maquina_inicio WHERE versao_id = %s", (versao_id,))
    maquinas = [{"maquina_id": r[0], "data_inicio": r[1]} for r in cur.fetchall()]
    cur.execute("ALTER TABLE programacao_lote ADD COLUMN IF NOT EXISTS data_fim TEXT")
    cur.execute("SELECT maquina_id, lote, data_inicio, data_fim, ordem FROM programacao_lote WHERE versao_id = %s ORDER BY maquina_id, ordem, lote", (versao_id,))
    lotes = [{"maquina_id": r[0], "lote": r[1], "data_inicio": r[2], "data_fim": r[3], "ordem": (r[4] or 0)} for r in cur.fetchall()]
    # Compat: se não há lotes salvos, deriva 1 lote (0) por máquina a partir das datas legadas.
    if not lotes:
        datas = {m["maquina_id"]: m["data_inicio"] for m in maquinas}
        maqs_com_item = sorted({it["maquina_id"] for it in itens})
        lotes = [{"maquina_id": mid, "lote": 0, "data_inicio": datas.get(mid), "ordem": 0} for mid in maqs_com_item]
    cur.close()
    conn.close()
    return {"versao_id": versao_id, "itens": itens, "maquinas": maquinas, "lotes": lotes, "total": len(itens)}


@router.put("/board")
def salvar_board(body: BoardBody, user_id: Optional[str] = Depends(get_user_id_from_session)):
    """Substitui a montagem da versão (lista completa item->máquina+ordem)."""
    uid = _uid(user_id, edit=True)
    _ensure_board()
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM programacao_board WHERE versao_id = %s", (body.versao_id,))
    for it in (body.itens or []):
        cod = (it.cod_item or "").strip()
        if not cod:
            continue
        seq = int(it.sequencia or 0)
        cur.execute(
            "INSERT INTO programacao_board (versao_id, cod_item, sequencia, maquina_id, ordem, lote, qtd, observacao, ativo, previsao_termino, concluido, qtd_produzida, updated_by) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) ON CONFLICT (versao_id, cod_item, sequencia) DO UPDATE SET "
            "maquina_id = EXCLUDED.maquina_id, ordem = EXCLUDED.ordem, lote = EXCLUDED.lote, qtd = EXCLUDED.qtd, "
            "observacao = EXCLUDED.observacao, ativo = EXCLUDED.ativo, previsao_termino = EXCLUDED.previsao_termino, "
            "concluido = EXCLUDED.concluido, qtd_produzida = EXCLUDED.qtd_produzida, "
            "updated_at = NOW(), updated_by = EXCLUDED.updated_by",
            (body.versao_id, cod, seq, it.maquina_id, it.ordem, (it.lote or 0), it.qtd, it.observacao,
             (True if it.ativo is None else bool(it.ativo)), (it.previsao_termino or None),
             bool(it.concluido), it.qtd_produzida,
             str(uid) if uid else None),
        )
    # lotes (programações por máquina, com data) — substitui tudo da versão
    cur.execute("DELETE FROM programacao_lote WHERE versao_id = %s", (body.versao_id,))
    for lt in (body.lotes or []):
        cur.execute(
            "INSERT INTO programacao_lote (versao_id, maquina_id, lote, data_inicio, data_fim, ordem, updated_by) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s) ON CONFLICT (versao_id, maquina_id, lote) DO UPDATE SET "
            "data_inicio = EXCLUDED.data_inicio, data_fim = EXCLUDED.data_fim, ordem = EXCLUDED.ordem, updated_at = NOW(), updated_by = EXCLUDED.updated_by",
            (body.versao_id, lt.maquina_id, lt.lote, (lt.data_inicio or None), (lt.data_fim or None), (lt.ordem or 0), str(uid) if uid else None),
        )
    # datas de início por máquina (legado / compat)
    for m in (body.maquinas or []):
        cur.execute(
            "INSERT INTO programacao_maquina_inicio (versao_id, maquina_id, data_inicio, updated_by) "
            "VALUES (%s, %s, %s, %s) ON CONFLICT (versao_id, maquina_id) DO UPDATE SET "
            "data_inicio = EXCLUDED.data_inicio, updated_at = NOW(), updated_by = EXCLUDED.updated_by",
            (body.versao_id, m.maquina_id, (m.data_inicio or None), str(uid) if uid else None),
        )
    # 17/06: auto-versão silenciosa (não-oficial) — congela histórico a cada save.
    # Dedup por hash garante zero duplicata quando nada mudou. Falha não bloqueia save.
    auto_v = _auto_versao_silenciosa(cur, body.versao_id, uid)
    conn.commit()
    cur.close()
    conn.close()
    return {"ok": True, "total": len(body.itens or []), "auto_versao": auto_v}


@router.get("/uso-componentes")
def obter_uso_componentes(versao_id: str, cod_item: str, user_id: Optional[str] = Depends(get_user_id_from_session)):
    """Quantidades 'a utilizar' (encarte/taba) salvas para um item de uma versão."""
    _uid(user_id)
    _ensure_board()
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "SELECT cod_componente, descricao, tipo_comp, qtd_usar, maquina_id FROM programacao_uso_componente "
        "WHERE versao_id = %s AND cod_item = %s",
        (versao_id, str(cod_item).strip()),
    )
    itens = [{"cod_componente": r[0], "descricao": r[1], "tipo_comp": r[2],
              "qtd_usar": (float(r[3]) if r[3] is not None else None),
              "maquina_id": r[4]} for r in cur.fetchall()]
    cur.close()
    conn.close()
    return {"versao_id": versao_id, "cod_item": cod_item, "itens": itens}


@router.put("/uso-componente")
def salvar_uso_componente(body: UsoComponenteIn, user_id: Optional[str] = Depends(get_user_id_from_session)):
    """Salva (upsert) a quantidade 'a utilizar' de um componente (encarte/taba) de um item."""
    uid = _uid(user_id, edit=True)
    _ensure_board()
    cod_item = (body.cod_item or "").strip()
    cod_comp = (body.cod_componente or "").strip()
    if not cod_item or not cod_comp:
        raise HTTPException(status_code=400, detail="Item e componente são obrigatórios")
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO programacao_uso_componente (versao_id, cod_item, cod_componente, descricao, tipo_comp, qtd_usar, maquina_id, updated_by) "
        "VALUES (%s, %s, %s, %s, %s, %s, %s, %s) ON CONFLICT (versao_id, cod_item, cod_componente) DO UPDATE SET "
        "descricao = EXCLUDED.descricao, tipo_comp = EXCLUDED.tipo_comp, qtd_usar = EXCLUDED.qtd_usar, "
        "maquina_id = EXCLUDED.maquina_id, updated_at = NOW(), updated_by = EXCLUDED.updated_by",
        (body.versao_id, cod_item, cod_comp, body.descricao, body.tipo_comp, body.qtd_usar, body.maquina_id, str(uid) if uid else None),
    )
    conn.commit()
    cur.close()
    conn.close()
    return {"ok": True}


# ============================================================================
# Calendário de produção (global) — turno, pausas, fim de semana, feriados.
# Usado para calcular a previsão de término respeitando o calendário de trabalho.
# ============================================================================
DEFAULT_CALENDARIO = {
    "turnos": [                           # 1..N turnos por dia (união); turno noturno cruza a meia-noite
        {"inicio": "06:00", "fim": "14:00"},
        {"inicio": "14:00", "fim": "22:00"},
        {"inicio": "22:00", "fim": "06:00"},
    ],
    "setup_min": 69,                      # minutos fixos somados ao tempo de produção
    "pausas": [
        {"inicio": "10:00", "fim": "10:40"},
        {"inicio": "13:40", "fim": "14:00"},
        {"inicio": "17:00", "fim": "17:40"},
        {"inicio": "21:40", "fim": "22:00"},
    ],
    "dias_semana_folga": [0],             # 0=domingo .. 6=sábado (mesma base do getDay do JS)
    "feriados": [],                       # [{"data":"2026-05-01","folga":true} | {"data":"...","inicio":"06:00","fim":"14:00"}]
    "variaveis": [],                      # campos livres do usuário: [{"nome":"...","valor":"..."}]
}


class CalendarioIn(BaseModel):
    config: dict
    maquina_id: Optional[int] = None   # None = calendário GERAL; preenchido = override por máquina
    limpar: Optional[bool] = False     # se True (com maquina_id), remove o override e volta ao geral


def _ensure_calendario():
    """Garante a tabela e a linha única (id=1) do calendário; retorna o config atual."""
    import json
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS programacao_calendario (
            id INTEGER PRIMARY KEY DEFAULT 1,
            config JSONB NOT NULL,
            updated_at TIMESTAMP DEFAULT NOW(),
            updated_by TEXT
        )
        """
    )
    # Override de calendário por máquina (ex.: Injetora, Sopro, Injetora 3...). Ausência = usa o geral.
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS programacao_calendario_maquina (
            maquina_id INTEGER PRIMARY KEY,
            config JSONB NOT NULL,
            updated_at TIMESTAMP DEFAULT NOW(),
            updated_by TEXT
        )
        """
    )
    conn.commit()   # commita os CREATE TABLE sempre (senão, no caminho "linha já existe" sem commit,
                    # o CREATE da tabela por-máquina sofre rollback ao fechar a conexão -> GET dá 500)
    cur.execute("SELECT config FROM programacao_calendario WHERE id = 1")
    row = cur.fetchone()
    if not row:
        cur.execute(
            "INSERT INTO programacao_calendario (id, config) VALUES (1, %s::jsonb) ON CONFLICT (id) DO NOTHING",
            (json.dumps(DEFAULT_CALENDARIO),),
        )
        conn.commit()
        cfg = DEFAULT_CALENDARIO
    else:
        cfg = row[0]
    cur.close()
    conn.close()
    return cfg


@router.get("/calendario")
def obter_calendario(user_id: Optional[str] = Depends(get_user_id_from_session)):
    """Calendário GERAL + overrides por máquina (turno, pausas, fim de semana, feriados, setup)."""
    _uid(user_id)
    geral = _ensure_calendario()
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT maquina_id, config FROM programacao_calendario_maquina")
    por_maquina = {str(r[0]): r[1] for r in cur.fetchall()}
    cur.close()
    conn.close()
    return {"config": geral, "por_maquina": por_maquina}


@router.put("/calendario")
def salvar_calendario(body: CalendarioIn, user_id: Optional[str] = Depends(get_user_id_from_session)):
    """Salva o calendário GERAL (maquina_id ausente) OU o override de uma máquina (maquina_id preenchido).
    Com maquina_id + limpar=True, remove o override (a máquina volta a usar o geral)."""
    import json
    uid = _uid(user_id, edit=True)
    _ensure_calendario()
    conn = get_db_connection()
    cur = conn.cursor()
    if body.maquina_id is not None:
        if body.limpar:
            cur.execute("DELETE FROM programacao_calendario_maquina WHERE maquina_id = %s", (body.maquina_id,))
            conn.commit(); cur.close(); conn.close()
            return {"ok": True, "maquina_id": body.maquina_id, "removido": True}
        cfg = {**DEFAULT_CALENDARIO, **(body.config or {})}
        cur.execute(
            "INSERT INTO programacao_calendario_maquina (maquina_id, config, updated_by) VALUES (%s, %s::jsonb, %s) "
            "ON CONFLICT (maquina_id) DO UPDATE SET config = EXCLUDED.config, updated_at = NOW(), updated_by = EXCLUDED.updated_by",
            (body.maquina_id, json.dumps(cfg), str(uid) if uid else None),
        )
        conn.commit(); cur.close(); conn.close()
        return {"ok": True, "maquina_id": body.maquina_id, "config": cfg}
    cfg = {**DEFAULT_CALENDARIO, **(body.config or {})}
    cur.execute(
        "UPDATE programacao_calendario SET config = %s::jsonb, updated_at = NOW(), updated_by = %s WHERE id = 1",
        (json.dumps(cfg), str(uid) if uid else None),
    )
    conn.commit(); cur.close(); conn.close()
    return {"ok": True, "config": cfg}


@router.get("/tabas-sopro")
def tabas_sopro(versao_id: str, user_id: Optional[str] = Depends(get_user_id_from_session)):
    """Cards de taba no Sopro: 1 por (produto × sequência), lidos da programacao_taba_sopro."""
    _uid(user_id)
    _ensure_board()
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "SELECT cod_item, sequencia, cod_componente, descricao, qtd, maquina_id, ordem, lote, inicio "
        "FROM programacao_taba_sopro WHERE versao_id = %s "
        "ORDER BY maquina_id, lote, ordem, sequencia",
        (versao_id,),
    )
    tabas = [{"cod_item": r[0], "sequencia": int(r[1] or 0), "cod_componente": r[2],
              "descricao": r[3], "qtd": float(r[4]) if r[4] is not None else 0.0,
              "maquina_id": r[5], "ordem": (r[6] or 0), "lote": int(r[7] or 0),
              "inicio": (r[8].isoformat() if r[8] is not None else None)} for r in cur.fetchall()]
    cur.close()
    conn.close()
    return {"tabas": tabas}


@router.get("/uso-tabas")
def uso_tabas(versao_id: str, user_id: Optional[str] = Depends(get_user_id_from_session)):
    """Overrides manuais (qtd e máquina) de tabas, por (cod_item, cod_componente), de toda a versão.
    Usado pela programação automática para respeitar a quantidade/máquina que o usuário definiu no detalhe."""
    _uid(user_id)
    _ensure_board()
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "SELECT cod_item, cod_componente, qtd_usar, maquina_id FROM programacao_uso_componente "
        "WHERE versao_id = %s AND tipo_comp = 'taba'",
        (versao_id,),
    )
    itens = [{"cod_item": r[0], "cod_componente": r[1],
              "qtd_usar": (float(r[2]) if r[2] is not None else None),
              "maquina_id": r[3]} for r in cur.fetchall()]
    cur.close()
    conn.close()
    return {"itens": itens}


class TabaSoproCard(BaseModel):
    cod_item: str
    sequencia: int = 0
    cod_componente: str
    descricao: Optional[str] = None
    qtd: float = 0
    maquina_id: Optional[int] = None
    ordem: int = 0
    lote: int = 0
    inicio: Optional[str] = None


class TabasSoproIn(BaseModel):
    versao_id: str
    cards: List[TabaSoproCard] = []


@router.put("/tabas-sopro")
def salvar_tabas_sopro(body: TabasSoproIn, user_id: Optional[str] = Depends(get_user_id_from_session)):
    """Substitui (replace) os cards de taba no Sopro de uma versão. Usado pela programação automática
    e pelo mover/reordenar de cards. Só persiste cards com máquina (Sopro) definida."""
    _uid(user_id, edit=True)
    _ensure_board()
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM programacao_taba_sopro WHERE versao_id = %s", (body.versao_id,))
    rows = [(body.versao_id, c.cod_item, c.sequencia, c.cod_componente, c.descricao,
             c.qtd, c.maquina_id, c.ordem, c.lote, c.inicio)
            for c in body.cards if c.maquina_id is not None]
    if rows:
        from psycopg2.extras import execute_values
        execute_values(
            cur,
            "INSERT INTO programacao_taba_sopro "
            "(versao_id, cod_item, sequencia, cod_componente, descricao, qtd, maquina_id, ordem, lote, inicio) VALUES %s "
            "ON CONFLICT (versao_id, cod_item, sequencia, cod_componente) DO UPDATE SET "
            "descricao = EXCLUDED.descricao, qtd = EXCLUDED.qtd, maquina_id = EXCLUDED.maquina_id, "
            "ordem = EXCLUDED.ordem, lote = EXCLUDED.lote, inicio = EXCLUDED.inicio, updated_at = NOW()",
            rows,
        )
    conn.commit()
    cur.close()
    conn.close()
    return {"ok": True, "total": len(rows)}


class ExcelItemIn(BaseModel):
    ordem: int = 0
    seq: Optional[int] = None
    codigo: str = ""
    produto: str = ""
    qtd: float = 0
    pecas_hora: Optional[float] = None
    inicio: Optional[str] = None
    termino: Optional[str] = None
    ops: Optional[str] = None
    pedidos: Optional[str] = None
    encarte_cod: Optional[str] = None
    encarte_nome: Optional[str] = None
    encarte_qtd: Optional[str] = None
    encarte_un: Optional[str] = None
    taba_cod: Optional[str] = None
    taba_nome: Optional[str] = None
    taba_qtd: Optional[str] = None
    taba_un: Optional[str] = None


class ExcelLoteIn(BaseModel):
    titulo: str = ""
    itens: List[ExcelItemIn] = []


class ExcelMaquinaIn(BaseModel):
    nome: str
    lotes: List[ExcelLoteIn] = []


class ExcelProgIn(BaseModel):
    nome_versao: Optional[str] = None
    maquinas: List[ExcelMaquinaIn] = []


@router.post("/xlsx")
def gerar_xlsx_programacao(body: ExcelProgIn, user_id: Optional[str] = Depends(get_user_id_from_session)):
    """Gera a planilha estilizada da Programação: uma ABA por máquina (só as que têm programação),
    com cores e os lotes (turnos) separados. Recebe os dados já montados do board (com término)."""
    _uid(user_id)
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="1E40AF")
    header_font = Font(bold=True, color="FFFFFF")
    lote_fill = PatternFill("solid", fgColor="DBEAFE")
    lote_font = Font(bold=True, color="1E3A8A")
    titulo_font = Font(bold=True, size=14, color="111827")
    zebra = PatternFill("solid", fgColor="F3F6FB")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cols = ["Ordem", "Seq. otim.", "Código", "Produto", "Qtd. a produzir", "Peças/h", "Início Previsto", "Término", "Nº OPs", "Pedidos",
            "Encarte cód.", "Encarte nome", "Encarte qtd", "Encarte un",
            "Taba cód.", "Taba nome", "Taba qtd", "Taba un"]
    widths = [7, 10, 13, 36, 13, 9, 17, 17, 20, 34, 13, 30, 12, 8, 13, 30, 12, 8]
    wrap_cols = {9, 10, 11, 12, 13, 14, 15, 16, 17, 18}   # OPs/Pedidos + colunas de encarte/taba
    usados = set()

    for maq in body.maquinas:
        if not any(l.itens for l in maq.lotes):
            continue  # só máquinas com programação geram aba
        base = (maq.nome or "Maquina")[:28]
        nome_aba = base
        k = 2
        while nome_aba in usados:
            nome_aba = f"{base[:26]} {k}"
            k += 1
        usados.add(nome_aba)
        ws = wb.create_sheet(title=nome_aba)
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        r = 1
        ws.cell(r, 1, maq.nome).font = titulo_font
        r += 2
        for lote in maq.lotes:
            if not lote.itens:
                continue
            c0 = ws.cell(r, 1, lote.titulo or "Lote")
            c0.font = lote_font
            for c in range(1, len(cols) + 1):
                ws.cell(r, c).fill = lote_fill
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(cols))
            r += 1
            for c, h in enumerate(cols, 1):
                cc = ws.cell(r, c, h)
                cc.fill = header_fill
                cc.font = header_font
                cc.alignment = Alignment(horizontal="center", vertical="center")
                cc.border = border
            r += 1
            # Início previsto por produto: usa o enviado pelo front; senão, encadeia pelo
            # término do produto anterior (e, no 1º item, o início do lote, vindo do título).
            lote_inicio = ""
            if lote.titulo and "Início:" in lote.titulo:
                _li = lote.titulo.split("Início:")[-1].strip()
                lote_inicio = "" if _li.lower() in ("sem data", "") else _li
            prev_term = None
            for z, it in enumerate(lote.itens):
                inicio_val = it.inicio or prev_term or lote_inicio
                vals = [it.ordem, it.seq, it.codigo, it.produto, it.qtd, it.pecas_hora,
                        inicio_val or "", it.termino or "", it.ops or "", it.pedidos or "",
                        it.encarte_cod or "", it.encarte_nome or "", it.encarte_qtd or "", it.encarte_un or "",
                        it.taba_cod or "", it.taba_nome or "", it.taba_qtd or "", it.taba_un or ""]
                for c, v in enumerate(vals, 1):
                    cc = ws.cell(r, c, v)
                    cc.border = border
                    if z % 2 == 1:
                        cc.fill = zebra
                    if c in wrap_cols:           # OPs/Pedidos + colunas de encarte/taba — quebra de linha
                        cc.alignment = Alignment(wrap_text=True, vertical="top")
                    elif c in (1, 2, 5, 6):
                        cc.alignment = Alignment(horizontal="center", vertical="top")
                    else:
                        cc.alignment = Alignment(vertical="top")
                if it.termino:
                    prev_term = it.termino
                r += 1
            r += 1  # espaço entre lotes

    if not wb.sheetnames:
        ws = wb.create_sheet("Sem programação")
        ws.cell(1, 1, "Nenhuma máquina com programação para exportar.")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    nome = f"programacao_{(body.nome_versao or 'producao')}.xlsx".replace(" ", "_")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome}"'},
    )


@router.delete("/uso-componente")
def remover_uso_componente(versao_id: str, cod_item: str, cod_componente: str, user_id: Optional[str] = Depends(get_user_id_from_session)):
    """Remove uma entrada de uso de componente (encarte/taba) de um item."""
    _uid(user_id, edit=True)
    _ensure_board()
    cod_item_s = (cod_item or "").strip()
    cod_comp_s = (cod_componente or "").strip()
    if not cod_item_s or not cod_comp_s:
        raise HTTPException(status_code=400, detail="Item e componente são obrigatórios")
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "DELETE FROM programacao_uso_componente WHERE versao_id = %s AND cod_item = %s AND cod_componente = %s",
        (versao_id, cod_item_s, cod_comp_s),
    )
    conn.commit()
    cur.close()
    conn.close()
    return {"ok": True}


@router.get("/ops")
def obter_ops(refresh: bool = False, user_id: Optional[str] = Depends(get_user_id_from_session)):
    """OPs em produção (situação 8), uma por NUMERO_DA_OP."""
    _uid(user_id)
    try:
        ops = carregar_ops(refresh=refresh)
    except Exception as e:
        logger.error(f"Erro ao carregar OPs: {e}")
        raise HTTPException(status_code=502, detail=f"Erro ao consultar OPs no BigQuery: {e}")
    return {"ops": ops, "total": len(ops)}
