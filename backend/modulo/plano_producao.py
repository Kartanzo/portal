"""
Plano de Producao — endpoints + porte do otimizador PuLP.

Fonte: gerar_plano_producao.py (EMPRESA).
Adapta carregar_dados para receber DataFrames (BigQuery direto).
Persiste cada execucao em portal_chamado(_homolog).plano_producao_versoes.
Retencao: 30 dias (limpeza automatica a cada gerar()).

MODO DUMMY (sem fontes externas): a unica fonte externa deste modulo era o
BigQuery, acessado por _bq_client()/carregar_dados_bq(). carregar_dados_bq()
agora monta (ped, est) deterministicamente via core.dummy (_dados_dummy),
preservando as mesmas colunas/tipos; _bq_client() ficou desativado. Postgres
(db_utils / plano_producao_versoes) permanece intacto.
"""
import io
import os
import re
import json
import base64
import logging
from datetime import datetime, date, timedelta, timezone
from collections import defaultdict
from typing import Optional

import pandas as pd
import pulp
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from db_utils import get_db_connection
from permission_utils import check_module_permission, check_sector_permission
from auth_utils import get_user_id_from_session

router = APIRouter()
logger = logging.getLogger(__name__)

MODULE_ID = 'plano_producao'
MODULE_FAT = 'otimizador_faturamento'
RETENCAO_DIAS = 30


def ensure_plano_producao_table():
    """Cria a tabela de versoes se ainda nao existir (idempotente).
    Usa o search_path ja configurado pela conexao (portal_chamado ou portal_chamado_homolog)."""
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS plano_producao_versoes (
                id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
                created_at TIMESTAMP DEFAULT NOW(),
                created_by UUID,
                created_by_name TEXT,
                hoje DATE NOT NULL,
                elapsed_seconds REAL,
                totais JSONB NOT NULL,
                plano JSONB NOT NULL,
                pedidos_completos JSONB NOT NULL,
                detalhe_alocacao JSONB NOT NULL,
                notes TEXT
            )
        """)
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_ppv_created_at
                ON plano_producao_versoes(created_at DESC)
        """)
        # Flag de versão oficial (migração idempotente)
        for ddl in [
            "ALTER TABLE plano_producao_versoes ADD COLUMN IF NOT EXISTS oficial BOOLEAN DEFAULT FALSE",
            "ALTER TABLE plano_producao_versoes ADD COLUMN IF NOT EXISTS oficial_em TIMESTAMP",
            "ALTER TABLE plano_producao_versoes ADD COLUMN IF NOT EXISTS oficial_por TEXT",
            "ALTER TABLE plano_producao_versoes ADD COLUMN IF NOT EXISTS oficial_por_nome TEXT",
        ]:
            cur.execute(ddl)
        # Fase C (16/06): lock real — só UMA `plano_producao_versoes` pode ser oficial globalmente.
        # Antes era lock aplicativo. Migração idempotente: limpa legado uma vez + cria UNIQUE INDEX parcial.
        cur.execute("""
            DO $$
            BEGIN
                IF NOT EXISTS (SELECT 1 FROM pg_indexes WHERE indexname = 'idx_plano_oficial_unique') THEN
                    CREATE TABLE IF NOT EXISTS plano_producao_versoes_oficial_cleanup_backup AS
                        SELECT *, NOW() AS backup_em FROM plano_producao_versoes WHERE oficial = TRUE;
                    UPDATE plano_producao_versoes SET oficial = FALSE
                    WHERE oficial = TRUE AND id NOT IN (
                        SELECT id FROM plano_producao_versoes
                        WHERE oficial = TRUE
                        ORDER BY COALESCE(oficial_em, created_at) DESC LIMIT 1
                    );
                    CREATE UNIQUE INDEX idx_plano_oficial_unique
                        ON plano_producao_versoes(oficial) WHERE oficial = TRUE;
                END IF;
            END $$;
        """)
        conn.commit()
        logger.info("plano_producao_versoes: tabela, indices e lock oficial OK")
    except Exception as e:
        conn.rollback()
        logger.error(f"Falha ao criar plano_producao_versoes: {e}")
    finally:
        cur.close()
        conn.close()

# =============================================================================
# BIGQUERY CLIENT (mesmo padrao de sac.py / importation.py)
# =============================================================================
_BQ_KEY_FILE = os.path.join(os.path.dirname(__file__), "..", "projeto-rpa-empresa-2023-16b15891f73c.json")
_BQ_PROJECT = "projeto-rpa-empresa-2023"

def _bq_client():
    # MODO DUMMY: sem fontes externas. O cliente BigQuery deixa de ser necessario
    # (carregar_dados_bq monta os DataFrames a partir de core.dummy). Mantido como
    # no-op lazy para nao quebrar import/credenciais; nao deve ser chamado.
    raise RuntimeError("BigQuery desativado (modo dummy): use carregar_dados_bq().")


# =============================================================================
# PORT DO gerar_plano_producao.py
# =============================================================================
MESES = {'jan': 1, 'fev': 2, 'mar': 3, 'abr': 4, 'mai': 5, 'jun': 6,
         'jul': 7, 'ago': 8, 'set': 9, 'out': 10, 'nov': 11, 'dez': 12}

def parse_data_pt(s):
    if pd.isna(s):
        return pd.NaT
    if isinstance(s, (pd.Timestamp, datetime)):
        return pd.Timestamp(s)
    m = re.match(r'(\d+) de (\w+)\.? de (\d+)', str(s).strip())
    if not m:
        try:
            txt = str(s).strip().split(' ')[0]
            if re.match(r'^\d{4}-\d{2}-\d{2}$', txt):
                return pd.to_datetime(txt, format='%Y-%m-%d')
            return pd.to_datetime(s, dayfirst=True)
        except Exception:
            return pd.NaT
    dia, mes, ano = int(m.group(1)), MESES.get(m.group(2).lower()), int(m.group(3))
    if not mes:
        return pd.NaT
    return pd.Timestamp(year=ano, month=mes, day=dia)


def _dados_dummy() -> tuple:
    """Monta (ped, est) deterministicamente a partir de core.dummy.

    Preserva EXATAMENTE as colunas que vinham do BigQuery:
      ped: RAZAO, EMISSAO, ENTREGA, EMISSAO_ORIGINAL, PEDIDO, DESCRICAO_PRODUTO,
           CODIGO_PRODUTO, TOTAL_ITEM, QUANTIDADE, DESC_TIPODOCUMENTO,
           STATUS_PEDIDO, SITUACAO, GERENCIA_REGIONAL
      est: CODIGO_PRODUTO, QUANTIDADE_DISPONIVEL, QUANTIDADE_RESERVA
    Pedidos cobrem os 12 meses de 2026 (lancamentos em TODOS os meses).
    """
    from core import dummy

    r = dummy.rng('plano_producao', 'pedidos', dummy.ANO_BASE)
    ano = dummy.ANO_BASE
    hoje = date(ano, 6, 24)  # referencia p/ produzir atrasados/imediatos/programados

    # SKUs canonicos (codigo sem 'BR'); descricao do pool de produtos
    prods = dummy.PRODUTOS
    tipos_doc = ['VENDA', 'DEPÓSITO ANTECIPADO', 'SAC', 'BONIFICAÇÃO', 'TROCA']
    regionais = ['REGIONAL SUL', 'REGIONAL SUDESTE', 'REGIONAL NORDESTE', 'REGIONAL CENTRO-OESTE']

    ped_rows = []
    n_ped = 0
    # 4 pedidos por mes -> 48 pedidos, garantindo lancamentos em todos os meses
    for mes in range(1, 13):
        for k in range(4):
            n_ped += 1
            pedido = f"PV{ano%100:02d}{n_ped:05d}"
            cliente = dummy.escolher(r, dummy.CLIENTES)
            tipo = dummy.escolher(r, tipos_doc)
            # 1 em cada 8 e' pedido interno (RAZAO especial -> grupo INTERNO)
            if n_ped % 8 == 0:
                cliente = 'EMPRESA PEDIDOS INTERNOS'
            emissao = dummy.dia_aleatorio(ano, mes, r)
            # entrega: distribui atrasado (passado), imediato (=emissao), programado (futuro)
            faixa = n_ped % 3
            if faixa == 0:        # atrasado: entrega < hoje
                entrega = emissao if emissao < hoje else dummy.dia_aleatorio(ano, max(1, mes-1), r)
                emissao_orig = entrega
            elif faixa == 1:      # imediato: entrega == emissao_orig == emissao
                entrega = emissao
                emissao_orig = emissao
            else:                 # programado: entrega no futuro
                fut_mes = min(12, mes + 2)
                entrega = dummy.dia_aleatorio(ano, fut_mes, r)
                emissao_orig = emissao
            status = dummy.escolher(r, ['1', '4'])          # apos filtro WHERE
            regional = dummy.escolher(r, regionais)
            # 2 a 4 itens por pedido
            n_itens = r.randint(2, 4)
            skus = r.sample(prods, n_itens)
            for (cod, desc, _un, _cat) in skus:
                qtd = r.randint(5, 60)
                preco = dummy.valor(r, base=120.0, var=0.5)
                total = round(qtd * preco, 2)
                # exercita o strip de 'BR' em ~1/4 dos SKUs (BQ trazia codigos com prefixo BR)
                cod_ped = ('BR' + cod) if (r.random() < 0.25) else cod
                ped_rows.append({
                    'RAZAO': cliente,
                    'EMISSAO': emissao.strftime('%Y-%m-%d'),
                    'ENTREGA': entrega.strftime('%Y-%m-%d'),
                    'EMISSAO_ORIGINAL': emissao_orig.strftime('%Y-%m-%d'),
                    'PEDIDO': pedido,
                    'DESCRICAO_PRODUTO': desc,
                    'CODIGO_PRODUTO': cod_ped,
                    'TOTAL_ITEM': total,
                    'QUANTIDADE': qtd,
                    'DESC_TIPODOCUMENTO': tipo,
                    'STATUS_PEDIDO': status,
                    'SITUACAO': 'ATIVO',
                    'GERENCIA_REGIONAL': regional,
                })
    ped = pd.DataFrame(ped_rows, columns=[
        'RAZAO', 'EMISSAO', 'ENTREGA', 'EMISSAO_ORIGINAL', 'PEDIDO', 'DESCRICAO_PRODUTO',
        'CODIGO_PRODUTO', 'TOTAL_ITEM', 'QUANTIDADE', 'DESC_TIPODOCUMENTO',
        'STATUS_PEDIDO', 'SITUACAO', 'GERENCIA_REGIONAL'])

    # Estoque: uma linha por SKU canonico. Estoque deliberadamente parcial para
    # forcar producao (qty_produzir > 0) em parte dos itens.
    re = dummy.rng('plano_producao', 'estoque', ano)
    est_rows = []
    for (cod, _desc, _un, _cat) in prods:
        disp = re.randint(0, 200)
        reserva = re.randint(0, 40)
        est_rows.append({
            'CODIGO_PRODUTO': cod,
            'QUANTIDADE_DISPONIVEL': disp,
            'QUANTIDADE_RESERVA': reserva,
        })
    est = pd.DataFrame(est_rows, columns=['CODIGO_PRODUTO', 'QUANTIDADE_DISPONIVEL', 'QUANTIDADE_RESERVA'])
    return ped, est


def carregar_dados_bq() -> tuple:
    """Carrega estoque + pedidos (MODO DUMMY: gerados em core.dummy) e normaliza.

    Antes vinha do BigQuery. Agora os DataFrames sao montados deterministicamente
    a partir de core.dummy, preservando EXATAMENTE as colunas/tipos que o restante
    do pipeline e o frontend consomem. SQL original mantido como referencia abaixo.
    """
    # SQL original (referencia — nao executado em modo dummy):
    # STATUS_PEDIDO eh STRING no BigQuery, comparar como string
    sql_pedidos = """
        SELECT RAZAO, EMISSAO, ENTREGA, EMISSAO_ORIGINAL, PEDIDO, DESCRICAO_PRODUTO,
               CODIGO_PRODUTO, TOTAL_ITEM, QUANTIDADE, DESC_TIPODOCUMENTO,
               STATUS_PEDIDO, SITUACAO, GERENCIA_REGIONAL
        FROM `projeto-rpa-empresa-2023.VENDAS.Metas_por_faturamento`
        WHERE STATUS_PEDIDO IN ('1','4')
          AND DESC_TIPODOCUMENTO IS NOT NULL
          AND DESC_TIPODOCUMENTO NOT IN ('DISPLAY','CAMPANHAS','RAPEL','MOSTRUARIO','None','CONTRATOS')
          AND PEDIDO NOT LIKE 'PVM%'
    """
    # estoque_logistica tem multiplas linhas por item (varios locais);
    # somar DISPONIVEL e renomear colunas para o esquema canonico esperado
    sql_estoque = """
        SELECT CODIGO_ITEM AS CODIGO_PRODUTO,
               SUM(DISPONIVEL) AS QUANTIDADE_DISPONIVEL,
               SUM(RESERVA) AS QUANTIDADE_RESERVA
        FROM `projeto-rpa-empresa-2023.VENDAS.estoque_logistica`
        WHERE CODIGO_ITEM IS NOT NULL
        GROUP BY CODIGO_ITEM
    """
    ped, est = _dados_dummy()

    # Normalizacao (mesma logica do script original)
    ped['COD'] = ped['CODIGO_PRODUTO'].astype(str).str.replace('BR', '', regex=False).str.strip()
    est['COD'] = est['CODIGO_PRODUTO'].astype(str).str.strip()
    ped['DATA'] = ped['EMISSAO'].apply(parse_data_pt)
    ped['DATA_ENTREGA'] = ped['ENTREGA'].apply(parse_data_pt)
    ped['DATA_EMISSAO_ORIG'] = ped['EMISSAO_ORIGINAL'].apply(parse_data_pt)
    ped['QUANTIDADE'] = pd.to_numeric(ped['QUANTIDADE'], errors='coerce').fillna(0)
    ped['TOTAL_ITEM'] = pd.to_numeric(ped['TOTAL_ITEM'], errors='coerce').fillna(0)
    est['QUANTIDADE_DISPONIVEL'] = pd.to_numeric(est['QUANTIDADE_DISPONIVEL'], errors='coerce').fillna(0)
    est['QUANTIDADE_RESERVA'] = pd.to_numeric(est['QUANTIDADE_RESERVA'], errors='coerce').fillna(0)
    return est, ped


def otimizar(ped, est, hoje, time_limit=300, msg=False):
    lin = ped.groupby(['PEDIDO', 'COD', 'RAZAO', 'DATA', 'DATA_ENTREGA', 'DATA_EMISSAO_ORIG', 'DESCRICAO_PRODUTO'],
                      as_index=False).agg(QTD=('QUANTIDADE', 'sum'),
                                          VAL=('TOTAL_ITEM', 'sum'))

    stock = est.set_index('COD')['QUANTIDADE_DISPONIVEL'].to_dict()
    reserva = est.set_index('COD')['QUANTIDADE_RESERVA'].to_dict() if 'QUANTIDADE_RESERVA' in est.columns else {}
    def S(c): return int(stock.get(c, 0))
    def R(c): return int(reserva.get(c, 0))
    def _normaliza_tipo(s):
        s = str(s or '').upper().strip()
        if not s:
            return ''
        if 'SAC' in s:
            return 'SAC'
        if 'BONIFIC' in s:
            return 'BONIFICACAO'
        if 'TROCA' in s:
            return 'TROCA'
        return ''  # cliente padrão (sem cor especial)
    def _normaliza_razao(razao: str) -> str:
        return 'INTERNO' if str(razao or '').strip() == 'EMPRESA PEDIDOS INTERNOS' else ''
    # Tipo por pedido: inclui 'INTERNO' para pedidos de uso interno EMPRESA
    tipo_por_ped_raw = {}
    if 'DESC_TIPODOCUMENTO' in ped.columns:
        for _, row_t in ped[['PEDIDO','DESC_TIPODOCUMENTO','RAZAO']].drop_duplicates().iterrows():
            t = _normaliza_tipo(str(row_t['DESC_TIPODOCUMENTO'] or ''))
            r = _normaliza_razao(str(row_t['RAZAO'] or ''))
            if r: t = r  # INTERNO sobrescreve tipo
            if t: tipo_por_ped_raw[row_t['PEDIDO']] = t
    tipo_por_ped = tipo_por_ped_raw
    # Pedidos com tipo de documento "DEPÓSITO ANTECIPADO" — máxima prioridade quando atrasados
    dep_antecipado = set()
    if 'DESC_TIPODOCUMENTO' in ped.columns:
        for _, row_d in ped[['PEDIDO', 'DESC_TIPODOCUMENTO']].drop_duplicates().iterrows():
            if 'ANTECIPAD' in str(row_d['DESC_TIPODOCUMENTO'] or '').upper():
                dep_antecipado.add(row_d['PEDIDO'])
    situacao_por_ped = (ped.dropna(subset=['STATUS_PEDIDO'])
                          .groupby('PEDIDO')['STATUS_PEDIDO'].first()
                          .astype(str).to_dict()
                        if 'STATUS_PEDIDO' in ped.columns else {})
    regional_por_ped = (ped.dropna(subset=['GERENCIA_REGIONAL'])
                          .groupby('PEDIDO')['GERENCIA_REGIONAL'].first()
                          .astype(str).to_dict()
                        if 'GERENCIA_REGIONAL' in ped.columns else {})

    meta = lin.groupby('PEDIDO').agg(
        cliente=('RAZAO', 'first'), emissao=('DATA', 'min'),
        entrega=('DATA_ENTREGA', 'min'),
        emissao_orig=('DATA_EMISSAO_ORIG', 'min'),
        valor=('VAL', 'sum'), n_itens=('COD', 'count')).reset_index()
    corte_5d = hoje - pd.Timedelta(days=5)
    mesma_data = meta['entrega'] == meta['emissao_orig']
    meta['atraso'] = mesma_data & (meta['entrega'] < corte_5d) | ~mesma_data & (meta['entrega'] < hoje)

    pedidos = meta['PEDIDO'].tolist()
    val = dict(zip(meta['PEDIDO'], meta['valor']))
    atr = dict(zip(meta['PEDIDO'], meta['atraso']))

    dem = defaultdict(float)
    val_item = defaultdict(float)  # valor (R$) da linha do item por (pedido, SKU)
    itens_por_ped = defaultdict(list)
    peds_por_sku = defaultdict(list)
    for r in lin.itertuples():
        dem[(r.PEDIDO, r.COD)] += r.QTD
        val_item[(r.PEDIDO, r.COD)] += r.VAL
        itens_por_ped[r.PEDIDO].append(r.COD)
        peds_por_sku[r.COD].append(r.PEDIDO)
    skus = sorted(set(lin['COD']))

    def build(fix_over=None, fix_tot=None):
        m = pulp.LpProblem('plano', pulp.LpMaximize)
        x = {p: pulp.LpVariable(f'x_{p}', cat='Binary') for p in pedidos}
        q = {s: pulp.LpVariable(f'q_{s}', lowBound=0, cat='Integer') for s in skus}
        for s in skus:
            # Estoque disponivel (S) + reserva (R) + producao (q) cobre demanda
            # qty_produzir = max(0, demanda - estoque_disponivel - reserva) = max(0, demanda - estoque_fisico)
            m += pulp.lpSum(dem[(p, s)] * x[p] for p in peds_por_sku[s]) <= S(s) + R(s) + q[s]
        over = [p for p in pedidos if atr[p]]
        onti = [p for p in pedidos if not atr[p]]
        oo = pulp.lpSum(x[p] for p in over)
        ot = pulp.lpSum(x[p] for p in onti)
        rv = pulp.lpSum(val[p] * x[p] for p in pedidos)
        if fix_over is not None: m += oo >= fix_over
        if fix_tot is not None: m += (oo + ot) >= fix_tot
        return m, x, q, oo, ot, rv

    solver = pulp.PULP_CBC_CMD(msg=msg, timeLimit=time_limit)

    m, x, q, oo, ot, rv = build()
    m += oo; m.solve(solver)
    best_over = int(round(pulp.value(oo)))

    m, x, q, oo, ot, rv = build(fix_over=best_over)
    m += oo + ot; m.solve(solver)
    best_tot = int(round(pulp.value(oo + ot)))

    m, x, q, oo, ot, rv = build(fix_over=best_over, fix_tot=best_tot)
    # Tiebreaker: dentro do mesmo faturamento, minimiza produção (q) — força
    # q[s] = max(0, demanda - S - R) e evita o solver parar em valor maior.
    # Valor de pedido (rv) e' em reais; produção em unidades. eps pequeno
    # garante que rv domina e q so' desempata.
    eps = 1e-4
    m += rv - eps * pulp.lpSum(q[s] for s in skus)
    m.solve(solver)

    fulfilled = {p for p in pedidos if pulp.value(x[p]) > 0.5}
    produce = {s: int(round(pulp.value(q[s]))) for s in skus if pulp.value(q[s]) > 0.5}

    return {'fulfilled': fulfilled, 'produce': produce, 'val': val, 'val_item': val_item, 'atr': atr,
            'dem': dem, 'itens_por_ped': itens_por_ped, 'peds_por_sku': peds_por_sku,
            'pedidos': pedidos, 'skus': skus, 'stock_fn': S, 'reserva_fn': R,
            'tipo_por_ped': tipo_por_ped, 'situacao_por_ped': situacao_por_ped,
            'regional_por_ped': regional_por_ped, 'dep_antecipado': dep_antecipado,
            'lin': lin, 'meta': meta}


def alocar_estoque_real(sol):
    meta = sol['meta']
    meta_sorted = meta.sort_values(['atraso', 'valor'], ascending=[False, False])
    lin = sol['lin']
    stock = {s: sol['stock_fn'](s) for s in sol['skus']}
    running = dict(stock)
    alloc_order, alloc_detail = [], []
    for _, pm in meta_sorted.iterrows():
        p = pm['PEDIDO']
        items = lin[lin['PEDIDO'] == p]
        can = all(row['QTD'] <= running.get(row['COD'], 0) for _, row in items.iterrows())
        if can:
            for _, row in items.iterrows():
                antes = running[row['COD']]
                running[row['COD']] = antes - row['QTD']
                alloc_detail.append({
                    'pedido': p, 'sku': row['COD'], 'qtd': int(row['QTD']),
                    'est_antes': antes, 'est_depois': antes - row['QTD']
                })
            alloc_order.append(p)
    return alloc_order, alloc_detail


def sequenciar(sol, hoje=None):
    """Sequencia producao em 3 grupos: atrasados -> imediatos -> programados.

    Um mesmo SKU pode aparecer em ate 3 linhas (uma por grupo com demanda).
    Dentro de cada grupo:
      - Atrasados: ordenados por dias de atraso desc, depois valor desc.
      - Imediatos (emissao_orig == entrega): ordenados por valor desc.
      - Programados (demais): ordenados por valor desc.
    """
    fulfilled     = sol['fulfilled']
    produce       = sol['produce']
    dem           = sol['dem']
    val           = sol['val']
    atr           = sol['atr']
    meta          = sol['meta']
    itens_por_ped = sol['itens_por_ped']
    dep_antecipado = sol.get('dep_antecipado', set())

    if hoje is None:
        hoje = pd.Timestamp.today().normalize()

    entrega_map      = dict(zip(meta['PEDIDO'], meta['entrega']))
    emissao_orig_map = dict(zip(meta['PEDIDO'], meta['emissao_orig']))

    def tipo_pedido(p):
        en = entrega_map.get(p)
        # Atrasado = qualquer pedido cuja data de entrega já passou (independente de ser imediato)
        if pd.notna(en) and en < hoje:
            # Atrasado + Depósito Antecipado = máxima prioridade
            return 'atraso_deposito' if p in dep_antecipado else 'atrasado'
        eo = emissao_orig_map.get(p)
        if pd.notna(en) and pd.notna(eo) and en == eo:
            return 'imediato'
        return 'programado'

    def dias_atraso(p):
        en = entrega_map.get(p)
        return max(0, (hoje - en).days) if pd.notna(en) else 0

    # Grupo (prioridade) de cada pedido — usado para que cada pedido apareça só na linha do seu grupo.
    sol['grupo_por_ped'] = {p: tipo_pedido(p) for p in fulfilled}

    # Acumular qty, valor e max_dias por (sku, grupo)
    sku_grupo_qty  = defaultdict(lambda: defaultdict(float))
    sku_grupo_val  = defaultdict(lambda: defaultdict(float))
    sku_grupo_dias = defaultdict(lambda: defaultdict(float))

    for p in fulfilled:
        tipo = tipo_pedido(p)
        d    = dias_atraso(p)
        for s in itens_por_ped[p]:
            q = dem.get((p, s), 0)
            if q <= 0:
                continue
            sku_grupo_qty[s][tipo] += q
            sku_grupo_val[s][tipo] += val[p]
            if tipo in ('atraso_deposito', 'atrasado'):
                sku_grupo_dias[s][tipo] = max(sku_grupo_dias[s][tipo], d)

    # Simula estoque para rastrear quais pedidos ficam completos a cada passo.
    # completed_set começa VAZIO — todos os pedidos (inclusive os cobertos por estoque)
    # são tratados como ativos e aparecem na demanda de cada passo.
    S = sol['stock_fn']
    R_fn = sol.get('reserva_fn', lambda c: 0)
    # av inclui disponível + reserva para que ordens cobertas por reserva completem corretamente
    av            = {s: S(s) + R_fn(s) for s in sol['skus']}
    orders_needs  = {p: {s: dem[(p, s)] for s in itens_por_ped[p]} for p in fulfilled}
    completed_set = set()

    # Cada SKU aparece UMA vez (no grupo de maior prioridade que tem demanda).
    # qty = produce[s] completo → garante DEMANDA - FÍSICO = QTD sempre.
    # Estoque restante para cada SKU (S+R), decrementado conforme os grupos consomem
    stock_restante: dict = {s: float(S(s) + R_fn(s)) for s in produce}
    # Componentes do restante (disponível e reserva) p/ exibir disp/reserva/físico em cascata
    disp_restante: dict = {s: float(S(s)) for s in produce}
    res_restante: dict = {s: float(R_fn(s)) for s in produce}

    seq = []

    for grupo in ('atraso_deposito', 'atrasado', 'imediato', 'programado'):
        skus_g = [(s, sku_grupo_qty[s][grupo])
                  for s in produce if sku_grupo_qty[s][grupo] > 0]
        if grupo in ('atraso_deposito', 'atrasado'):
            skus_g.sort(key=lambda x: (-sku_grupo_dias[x[0]][grupo],
                                       -sku_grupo_val[x[0]][grupo]))
        else:
            skus_g.sort(key=lambda x: -sku_grupo_val[x[0]][grupo])

        for s, D_grupo in skus_g:
            done_before = set(completed_set)
            # Ignora se não há pedido ativo que utilize este SKU neste grupo
            if not any(s in orders_needs[p] for p in fulfilled - done_before):
                stock_restante[s] = max(0.0, stock_restante[s] - D_grupo)
                _usa_d = min(disp_restante[s], D_grupo)
                disp_restante[s] -= _usa_d
                res_restante[s] = max(0.0, res_restante[s] - (D_grupo - _usa_d))
                continue

            sr = stock_restante[s]
            disp_grupo = disp_restante[s]            # disponível remanescente entrando neste grupo
            res_grupo  = res_restante[s]             # reserva remanescente entrando neste grupo
            fisico_grupo = min(D_grupo, sr)          # estoque que cobre este grupo
            actual_qty   = max(0.0, D_grupo - sr)    # produção necessária para este grupo
            stock_restante[s] = max(0.0, sr - D_grupo)
            _usa_d = min(disp_grupo, D_grupo)        # consome disponível primeiro, depois reserva
            disp_restante[s] = disp_grupo - _usa_d
            res_restante[s] = max(0.0, res_grupo - (D_grupo - _usa_d))

            av_antes = av[s]
            if actual_qty > 0:
                av[s] += actual_qty

            # Só pedidos do grupo atual completam neste step (evita que programados completem no atrasado)
            newly = [p for p in fulfilled - completed_set
                     if tipo_pedido(p) == grupo
                     and s in orders_needs[p]
                     and all(av[s2] >= need for s2, need in orders_needs[p].items())]
            completed_set.update(newly)

            # Só adiciona à seq se há produção OU pedidos ativos para mostrar.
            # Máx. Prioridade (atraso+depósito) só aparece se houver PRODUÇÃO (Qtd > 0):
            # se o físico já atende a demanda, não é prioridade de produção.
            # Só entra no plano quem tem PRODUÇÃO (Qtd > 0). Item coberto por estoque (Qtd 0)
            # não é produção, então não aparece — e nenhum pedido completa num item de Qtd 0.
            mostra = actual_qty > 0.5
            if mostra:
                seq.append({
                    'sku': s, 'qty': round(actual_qty), 'grupo': grupo,
                    'newly': newly, 'done_before': done_before, 'av_antes': av_antes,
                    'demanda_grupo': D_grupo, 'fisico_grupo': fisico_grupo,
                    'disp_grupo': disp_grupo, 'res_grupo': res_grupo,
                })

    # Reconciliação: garante que TODO item produzido apareça na sequência.
    # Itens cuja produção é puxada por demanda fora dos grupos exibidos (ex.: componentes
    # como CONJ PARAFUSO) antes ficavam invisíveis — e a conclusão de pedidos que ocorria
    # nesses passos sumia junto. Vale para QUALQUER item nessa situação (não só um pedido).
    _ordem_grupos = ('atraso_deposito', 'atrasado', 'imediato', 'programado')
    emitidos = {st['sku'] for st in seq}
    for s in [c for c in produce if produce[c] > 0 and c not in emitidos]:
        g = next((gg for gg in _ordem_grupos if sku_grupo_qty[s][gg] > 0), 'programado')
        av_antes = av[s]
        av[s] += produce[s]
        newly = [p for p in fulfilled - completed_set
                 if tipo_pedido(p) == g and s in orders_needs[p]
                 and all(av[s2] >= need for s2, need in orders_needs[p].items())]
        completed_set.update(newly)
        step = {
            'sku': s, 'qty': int(produce[s]), 'grupo': g, 'newly': newly,
            'done_before': set(completed_set) - set(newly), 'av_antes': av_antes,
            'demanda_grupo': sku_grupo_qty[s][g], 'fisico_grupo': S(s) + R_fn(s),
            'disp_grupo': S(s), 'res_grupo': R_fn(s),
        }
        # posição: logo após o último passo do grupo de prioridade do item
        pos = max((i for i, st in enumerate(seq) if st['grupo'] == g), default=len(seq) - 1) + 1
        seq.insert(pos, step)

    # Conclusão no ÚLTIMO produto da sequência do pedido (não no primeiro passo onde o
    # estoque nominalmente cobre). Cada pedido que completa é marcado na ÚLTIMA linha
    # exibida em que ele aparece (= último item da sua sequência de produção).
    _peds_por_sku = sol['peds_por_sku']
    ultima_linha = {}
    for i, st in enumerate(seq):
        for p in _peds_por_sku[st['sku']]:
            if p in completed_set and tipo_pedido(p) == st['grupo']:
                ultima_linha[p] = i
    for st in seq:
        st['newly'] = []
    for p, i in ultima_linha.items():
        seq[i]['newly'].append(p)

    return seq, completed_set


def montar_linhas(seq, sol, n_estoque_real, hoje=None):
    dem, val, atr = sol['dem'], sol['val'], sol['atr']
    val_item = sol['val_item']
    fulfilled, peds_por_sku = sol['fulfilled'], sol['peds_por_sku']
    grupo_por_ped = sol.get('grupo_por_ped', {})
    S = sol['stock_fn']
    R = sol.get('reserva_fn', lambda c: 0)
    desc_map = sol['lin'].drop_duplicates('COD').set_index('COD')['DESCRICAO_PRODUTO'].to_dict()
    initial = n_estoque_real
    cum = initial
    rows = []
    # Dias de atraso por pedido (para mostrar no plano)
    if hoje is None:
        hoje = pd.Timestamp.today().normalize()
    meta = sol['meta']
    entrega_map = dict(zip(meta['PEDIDO'], meta['entrega']))
    dias_por_ped = {}
    for p in fulfilled:
        en = entrega_map.get(p)
        try:
            dias_por_ped[p] = max(0, (hoje - en).days) if en is not None and pd.notna(en) else 0
        except Exception:
            dias_por_ped[p] = 0
    for i, step in enumerate(seq, 1):
        s, qty, newly = step['sku'], step['qty'], step['newly']
        done_before = step.get('done_before', set())
        grupo = step.get('grupo')
        # Deduplica peds_por_sku[s] (pode ter duplicatas por múltiplas linhas no BigQuery).
        # Mostra TODOS os pedidos do grupo desta linha que usam o SKU (cada pedido aparece
        # apenas na linha do seu grupo de prioridade — evita o mesmo pedido em 2 sequências,
        # mas NÃO esconde os que já completaram em passos anteriores).
        imp = list(dict.fromkeys(
            p for p in peds_por_sku[s]
            if p in fulfilled and (grupo_por_ped.get(p) == grupo if grupo else True)
        ))
        # Todos os fulfilled (para DEMANDA_TOTAL = av_antes + QTD → fórmula correta)
        all_imp = list(dict.fromkeys(p for p in peds_por_sku[s] if p in fulfilled))
        cum += len(newly)
        # Entrega mais cedo (deadline) entre os pedidos atendidos por este passo — usada pela
        # Programação para sinalizar se o término planejado vai passar da entrega (atraso).
        _entr_dts = [entrega_map.get(p) for p in (imp or all_imp)]
        _entr_dts = [e for e in _entr_dts if e is not None and pd.notna(e)]
        _entrega_row = min(_entr_dts).strftime('%d/%m/%Y') if _entr_dts else ''
        rows.append({
            'SEQUENCIA': i,
            'ENTREGA': _entrega_row,
            'CODIGO_PRODUTO': s,
            'DESCRICAO': desc_map.get(s, ''),
            'QTD_PRODUZIR': qty,
            'ESTOQUE_ATUAL': int(step.get('disp_grupo', S(s)) + step.get('res_grupo', R(s))),
            'ESTOQUE_DISPONIVEL': int(step.get('disp_grupo', S(s))),
            'RESERVA_ATUAL': int(step.get('res_grupo', R(s))),
            'ESTOQUE_FISICO': int(step.get('disp_grupo', S(s)) + step.get('res_grupo', R(s))),  # físico = disponível + reserva (cascata)
            # Demanda total = soma da demanda de TODOS os pedidos que usam o SKU neste passo
            # (a mesma lista de PEDIDOS_QUE_USAM_SKU), não só a fatia do grupo/que completa.
            'DEMANDA_TOTAL_PLANO': int(sum(dem[(p, s)] for p in imp)),
            'DEMANDA_POR_PEDIDO': {str(p): int(dem[(p, s)]) for p in imp},
            # Valor do item (linha do SKU) em TODOS os pedidos que usam o SKU (não só os que completam)
            'VALOR_ITEM_POR_PEDIDO': {str(p): round(float(val_item[(p, s)]), 2) for p in imp},
            'N_PEDIDOS_QUE_USAM_SKU': len(imp),
            'PEDIDOS_QUE_USAM_SKU': ', '.join(sorted(imp)),
            'N_PEDIDOS_COMPLETOS_APOS_ESTE_ITEM': len(newly),
            'PEDIDOS_COMPLETOS_APOS_ESTE_ITEM': ', '.join(sorted(newly)),
            # Valor do ITEM (linha do SKU) nos pedidos que completam, não o total do pedido
            'VALOR_PEDIDOS_COMPLETOS_APOS': round(sum(val_item[(p, s)] for p in newly), 2),
            'VALOR_POR_PEDIDO_COMPLETO': {str(p): round(float(val_item[(p, s)]), 2) for p in newly},
            'N_ATRASADOS_COMPLETOS_APOS': sum(1 for p in newly if atr[p]),
            'CUM_PEDIDOS_COMPLETOS': cum,
            'CUM_PCT': round(cum / len(sol['pedidos']) * 100, 2),
            'GRUPO': step.get('grupo', ''),
            'DIAS_MAX_ATRASO': max((dias_por_ped.get(p, 0) for p in imp if atr.get(p)), default=0),
        })
    return rows, initial, cum


def _estilos():
    thin = Side(style='thin', color='BFBFBF')
    return {
        'bd': Border(left=thin, right=thin, top=thin, bottom=thin),
        'title_font': Font(name='Arial', size=14, bold=True, color='FFFFFF'),
        'title_fill': PatternFill('solid', start_color='1F4E78'),
        'lbl': Font(name='Arial', size=10, bold=True),
        'vfn': Font(name='Arial', size=10),
        'hdr_font': Font(name='Arial', size=10, bold=True, color='FFFFFF'),
        'hdr_fill': PatternFill('solid', start_color='2E75B6'),
        'atr_fill': PatternFill('solid', start_color='FCE4D6'),
        'ok_fill': PatternFill('solid', start_color='E2EFDA'),
        'total_fill': PatternFill('solid', start_color='DDEBF7'),
    }


def gerar_xlsx_bytes(plano_rows, totais, hoje_str, pedidos_completos, detalhe_alocacao) -> bytes:
    """Recria o XLSX (3 abas) a partir dos JSONs armazenados no banco."""
    st = _estilos()
    wb = Workbook()

    # ABA 1: Plano_Producao
    ws = wb.active; ws.title = 'Plano_Producao'
    ws.merge_cells('A1:O1')
    ws['A1'] = 'PLANO DE PRODUÇÃO — Otimização PuLP (lex: atrasados → total → faturamento)'
    ws['A1'].font = st['title_font']; ws['A1'].fill = st['title_fill']
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    resumo = [
        ('Data de referência (hoje)', hoje_str),
        ('Pedidos na carteira', totais['n_pedidos']),
        ('Pedidos atendidos no plano', totais['n_atendidos']),
        ('  dos quais atrasados', f"{totais['n_atr_atend']} / {totais['n_atr_tot']}"),
        ('Pedidos completos só com estoque atual', totais['n_ini']),
        ('Pedidos completos após executar o plano', totais['n_fim']),
        ('Valor da carteira (R$)', f"R$ {totais['val_tot']:,.2f}"),
        ('Valor atendido no plano (R$)', f"R$ {totais['val_atend']:,.2f}"),
        ('SKUs a produzir', totais['n_skus']),
        ('Unidades a produzir', f"{totais['n_unid']:,}"),
    ]
    r = 3
    for k, v in resumo:
        ws.cell(row=r, column=1, value=k).font = st['lbl']
        ws.cell(row=r, column=2, value=v).font = st['vfn']
        r += 1

    start = r + 2
    headers = ['Seq', 'Código', 'Descrição', 'Qtd a Produzir', 'Estoque Disponível', 'Reserva', 'Estoque Físico',
               'Demanda Total (plano)', '# Pedidos que usam SKU', 'Pedidos que usam SKU',
               '# Pedidos Completos APÓS este item', 'Pedidos Completos APÓS este item',
               'Valor do Item', '# Atrasados Completos APÓS',
               'Cum. Pedidos Completos', 'Cum. %', 'Grupo', 'Max Dias Atraso']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=start, column=i, value=h)
        c.font = st['hdr_font']; c.fill = st['hdr_fill']
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = st['bd']
    ws.row_dimensions[start].height = 36

    for ri, row in enumerate(plano_rows, 1):
        rr = start + ri
        vals = [row['SEQUENCIA'], row['CODIGO_PRODUTO'], row['DESCRICAO'], row['QTD_PRODUZIR'],
                row.get('ESTOQUE_DISPONIVEL', row.get('ESTOQUE_ATUAL', 0)), row.get('RESERVA_ATUAL', 0),
                row.get('ESTOQUE_FISICO', row.get('ESTOQUE_ATUAL', 0) + row.get('RESERVA_ATUAL', 0)),
                row['DEMANDA_TOTAL_PLANO'],
                row['N_PEDIDOS_QUE_USAM_SKU'], row['PEDIDOS_QUE_USAM_SKU'],
                row['N_PEDIDOS_COMPLETOS_APOS_ESTE_ITEM'],
                row['PEDIDOS_COMPLETOS_APOS_ESTE_ITEM'],
                row['VALOR_PEDIDOS_COMPLETOS_APOS'],
                row['N_ATRASADOS_COMPLETOS_APOS'],
                row['CUM_PEDIDOS_COMPLETOS'], row['CUM_PCT'] / 100,
                row.get('GRUPO', ''), row.get('DIAS_MAX_ATRASO', 0)]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=rr, column=ci, value=v)
            c.font = st['vfn']; c.border = st['bd']
            c.alignment = Alignment(vertical='top', wrap_text=(ci in (3, 9, 11)))
        ws.cell(row=rr, column=4).number_format = '#,##0'
        ws.cell(row=rr, column=5).number_format = '#,##0'
        ws.cell(row=rr, column=6).number_format = '#,##0'
        ws.cell(row=rr, column=7).number_format = '#,##0'
        ws.cell(row=rr, column=12).number_format = 'R$ #,##0.00;(R$ #,##0.00);-'
        ws.cell(row=rr, column=15).number_format = '0.0%'
        if row['N_PEDIDOS_COMPLETOS_APOS_ESTE_ITEM'] > 0:
            for ci in range(1, 16):
                ws.cell(row=rr, column=ci).fill = st['ok_fill']

    widths = [6, 12, 45, 13, 12, 12, 15, 14, 55, 14, 55, 18, 14, 12, 9]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=start + 1, column=1)


    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =============================================================================
# PIPELINE COMPLETO
# =============================================================================
def _executar_otimizacao(hoje: pd.Timestamp, time_limit: int = 300) -> dict:
    """Roda o pipeline completo e devolve os 4 datasets serializaveis em JSON."""
    t0 = datetime.now()
    est, ped = carregar_dados_bq()
    if ped.empty:
        raise ValueError("Carteira de pedidos vazia (sem pedidos com status 1/4 apos filtros).")

    sol = otimizar(ped, est, hoje, time_limit=time_limit)
    alloc_order, alloc_detail = alocar_estoque_real(sol)
    seq, completed = sequenciar(sol, hoje=hoje)
    plano_rows, n_ini, n_fim = montar_linhas(seq, sol, len(alloc_order), hoje=hoje)

    val_d = sol['val']; atr_d = sol['atr']
    meta = sol['meta']
    cli = dict(zip(meta['PEDIDO'], meta['cliente']))
    emi = dict(zip(meta['PEDIDO'], meta['emissao']))
    lin = sol['lin']
    desc_map = lin.drop_duplicates('COD').set_index('COD')['DESCRICAO_PRODUTO'].to_dict()

    # Pedidos_Completos_Estoque (dataset serializavel)
    pedidos_completos = []
    for p in alloc_order:
        items = lin[lin['PEDIDO'] == p]
        dt = emi[p]
        pedidos_completos.append({
            'pedido': p, 'cliente': cli[p],
            'emissao': dt.strftime('%d/%m/%Y') if pd.notna(dt) else '',
            'atrasado': bool(atr_d[p]),
            'valor': round(float(val_d[p]), 2),
            'n_itens': int(len(items)),
            'skus': ', '.join(sorted(items['COD'].unique())),
            'qtd_total': int(items['QTD'].sum()),
        })

    # Detalhe_Alocacao_Estoque
    detalhe_alocacao = []
    for d in alloc_detail:
        p = d['pedido']
        item_val = float(lin[(lin['PEDIDO'] == p) & (lin['COD'] == d['sku'])]['VAL'].sum())
        detalhe_alocacao.append({
            'pedido': p, 'cliente': cli.get(p, ''),
            'sku': d['sku'], 'descricao': desc_map.get(d['sku'], ''),
            'qtd': int(d['qtd']), 'est_antes': int(d['est_antes']),
            'atrasado': bool(atr_d[p]), 'valor_linha': round(item_val, 2),
        })

    tipo_por_ped = sol.get('tipo_por_ped', {})
    situacao_por_ped = sol.get('situacao_por_ped', {})
    totais = {
        'n_pedidos': len(sol['pedidos']),
        'n_atendidos': len(sol['fulfilled']),
        'n_ini': n_ini, 'n_fim': n_fim,
        'n_atr_tot': sum(1 for p in sol['pedidos'] if sol['atr'][p]),
        'n_atr_atend': sum(1 for p in sol['fulfilled'] if sol['atr'][p]),
        'val_tot': round(float(sum(sol['val'].values())), 2),
        'val_atend': round(float(sum(sol['val'][p] for p in sol['fulfilled'])), 2),
        'n_skus': len(sol['produce']),
        'n_unid': int(sum(sol['produce'].values())),
        'pedidos_tipos': {str(p): str(t) for p, t in tipo_por_ped.items()},
        'pedidos_situacao': {str(p): str(s) for p, s in situacao_por_ped.items()},
        'pedidos_valores': {str(p): round(float(v), 2) for p, v in sol['val'].items()},
        'pedidos_atrasados': {str(p): bool(sol['atr'][p]) for p in sol['pedidos']},
        'pedidos_regionais': {str(p): str(r) for p, r in sol.get('regional_por_ped', {}).items()},
        'pedidos_clientes': {str(p): str(c) for p, c in zip(meta['PEDIDO'], meta['cliente'])},
        'pedidos_entregas': {str(p): (e.strftime('%d/%m/%Y') if not pd.isna(e) else '') for p, e in zip(meta['PEDIDO'], meta['entrega'])},
        'pedidos_emissao':  {str(p): (e.strftime('%d/%m/%Y') if not pd.isna(e) else '') for p, e in zip(meta['PEDIDO'], meta['emissao'])},
        'pedidos_emissao_orig': {str(p): (e.strftime('%d/%m/%Y') if not pd.isna(e) else '') for p, e in zip(meta['PEDIDO'], meta['emissao_orig'])},
        'pedidos_dias_atraso': {str(p): int((hoje - e).days) if not pd.isna(e) and bool(a) else 0 for p, e, a in zip(meta['PEDIDO'], meta['entrega'], meta['atraso'])},
        'pedidos_completos_estoque': [str(p) for p in alloc_order],
        'pedidos_atendidos': [str(p) for p in sol['fulfilled']],
    }
    # Mapeamento pedido→itens: fallback para pedidos 100% cobertos por estoque (sem producao)
    _desc_map = sol['lin'].drop_duplicates('COD').set_index('COD')['DESCRICAO_PRODUTO'].to_dict()
    totais['pedidos_itens'] = {
        str(p): list({s: {'sku': s,
                          'demanda': int(sol['dem'].get((p, s), 0) or 0),
                          'descricao': _desc_map.get(s, ''),
                          'valor': round(float(sol['val_item'].get((p, s), 0) or 0), 2)}
                      for s in sol['itens_por_ped'].get(p, [])}.values())
        for p in sol['fulfilled']
    }

    elapsed = (datetime.now() - t0).total_seconds()
    return {
        'totais': totais,
        'plano': plano_rows,
        'pedidos_completos': pedidos_completos,
        'detalhe_alocacao': detalhe_alocacao,
        'elapsed_seconds': elapsed,
    }


def _purge_versoes_antigas(cur):
    """Remove versoes com mais de RETENCAO_DIAS dias (NUNCA apaga a versao oficial)."""
    cutoff = datetime.now() - timedelta(days=RETENCAO_DIAS)
    cur.execute(
        "DELETE FROM plano_producao_versoes WHERE created_at < %s AND oficial IS NOT TRUE",
        (cutoff,)
    )


def _user_name(cur, user_id: Optional[str]) -> Optional[str]:
    if not user_id:
        return None
    cur.execute("SELECT name FROM users WHERE id = %s", (str(user_id),))
    row = cur.fetchone()
    return row[0] if row else None


# =============================================================================
# ENDPOINTS
# =============================================================================
class GerarPlanoBody(BaseModel):
    hoje: Optional[str] = None  # YYYY-MM-DD
    time_limit: Optional[int] = 300
    oficial: Optional[bool] = False  # marca a versão gerada como oficial


class OficialBody(BaseModel):
    oficial: bool


@router.post("/fabrica/plano-producao/gerar")
def gerar_plano(body: GerarPlanoBody, user_id: Optional[str] = Depends(get_user_id_from_session)):
    if not check_module_permission(user_id or '', MODULE_ID):
        raise HTTPException(status_code=403, detail="Acesso negado.")

    try:
        hoje_ts = pd.Timestamp(body.hoje) if body.hoje else pd.Timestamp.today().normalize()
    except Exception:
        raise HTTPException(status_code=400, detail="Data 'hoje' invalida. Use YYYY-MM-DD.")

    try:
        out = _executar_otimizacao(hoje_ts, time_limit=body.time_limit or 300)
    except ValueError as ve:
        raise HTTPException(status_code=400, detail=str(ve))
    except Exception as e:
        logger.error(f"Erro no otimizador: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Erro ao gerar plano: {e}")

    # Persiste versao
    conn = get_db_connection(); cur = conn.cursor()
    try:
        _purge_versoes_antigas(cur)
        author_name = _user_name(cur, user_id)
        eh_oficial = bool(getattr(body, 'oficial', False))
        cur.execute(
            """
            INSERT INTO plano_producao_versoes
                (created_by, created_by_name, hoje, elapsed_seconds,
                 totais, plano, pedidos_completos, detalhe_alocacao,
                 oficial, oficial_em, oficial_por, oficial_por_nome)
            VALUES (%s, %s, %s, %s, %s::jsonb, %s::jsonb, %s::jsonb, %s::jsonb,
                    %s, CASE WHEN %s THEN NOW() ELSE NULL END, %s, %s)
            RETURNING id, created_at
            """,
            (str(user_id) if user_id else None, author_name,
             hoje_ts.strftime('%Y-%m-%d'), out['elapsed_seconds'],
             json.dumps(out['totais']),
             json.dumps(out['plano']),
             json.dumps(out['pedidos_completos']),
             json.dumps(out['detalhe_alocacao']),
             eh_oficial, eh_oficial,
             (str(user_id) if (eh_oficial and user_id) else None),
             (author_name if eh_oficial else None))
        )
        new_id, created_at = cur.fetchone()
        if eh_oficial:
            # Só pode haver UMA oficial: desmarca as demais
            cur.execute("UPDATE plano_producao_versoes SET oficial = FALSE, oficial_em = NULL WHERE oficial = TRUE AND id <> %s", (new_id,))
        conn.commit()
    except Exception as e:
        conn.rollback()
        logger.error(f"Erro ao persistir versao: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Erro ao salvar versao: {e}")
    finally:
        cur.close(); conn.close()

    return {
        'id': str(new_id),
        'oficial': eh_oficial,
        'created_at': created_at.replace(tzinfo=timezone.utc).isoformat(),
        'created_by_name': author_name,
        'hoje': hoje_ts.strftime('%Y-%m-%d'),
        'elapsed_seconds': out['elapsed_seconds'],
        'totais': out['totais'],
        'plano': out['plano'],
        'pedidos_completos': out['pedidos_completos'],
        'detalhe_alocacao': out['detalhe_alocacao'],
    }


@router.get("/fabrica/plano-producao/versoes")
def listar_versoes(user_id: Optional[str] = Depends(get_user_id_from_session)):
    if not check_module_permission(user_id or '', MODULE_ID):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    conn = get_db_connection(); cur = conn.cursor()
    try:
        cur.execute("""
            SELECT id, created_at, created_by_name, hoje, elapsed_seconds, totais,
                   oficial, oficial_em, oficial_por_nome
            FROM plano_producao_versoes
            ORDER BY created_at DESC
            LIMIT 200
        """)
        rows = cur.fetchall()
        result = []
        for row in rows:
            totais = row[5] if isinstance(row[5], dict) else (json.loads(row[5]) if row[5] else {})
            # A LISTA só precisa dos totais escalares (resumo do histórico). Os mapas grandes
            # por-pedido (pedidos_tipos/situacao/entregas/valores/...) ficam apenas na versão
            # completa — senão a lista de N versões vira dezenas de MB e falha/lenta ao carregar.
            totais = {k: v for k, v in totais.items() if not isinstance(v, (dict, list))}
            result.append({
                'id': str(row[0]),
                'created_at': (row[1].replace(tzinfo=timezone.utc).isoformat() if row[1] else None),
                'created_by_name': row[2],
                'hoje': row[3].strftime('%Y-%m-%d') if row[3] else None,
                'elapsed_seconds': row[4],
                'totais': totais,
                'oficial': bool(row[6]),
                'oficial_em': (row[7].replace(tzinfo=timezone.utc).isoformat() if row[7] else None),
                'oficial_por_nome': row[8],
            })
        return result
    finally:
        cur.close(); conn.close()


@router.put("/fabrica/plano-producao/versoes/{versao_id}/oficial")
def marcar_oficial(versao_id: str, body: OficialBody, user_id: Optional[str] = Depends(get_user_id_from_session)):
    if not check_module_permission(user_id or '', MODULE_ID, 'can_edit'):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    if not check_sector_permission(user_id, 'Fábrica'):
        raise HTTPException(status_code=403, detail="Apenas usuários da Fábrica podem definir a versão oficial.")
    conn = get_db_connection(); cur = conn.cursor()
    try:
        nome = _user_name(cur, user_id)
        if body.oficial:
            # Só pode haver UMA versão oficial: desmarca as demais
            cur.execute("UPDATE plano_producao_versoes SET oficial = FALSE, oficial_em = NULL WHERE oficial = TRUE AND id <> %s", (versao_id,))
        cur.execute(
            """UPDATE plano_producao_versoes
                  SET oficial = %s,
                      oficial_em = CASE WHEN %s THEN NOW() ELSE NULL END,
                      oficial_por = %s,
                      oficial_por_nome = %s
                WHERE id = %s""",
            (bool(body.oficial), bool(body.oficial),
             (str(user_id) if (body.oficial and user_id) else None),
             (nome if body.oficial else None), versao_id),
        )
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Versão não encontrada.")
        conn.commit()
        return {"ok": True, "oficial": bool(body.oficial)}
    finally:
        cur.close(); conn.close()


@router.get("/fabrica/plano-producao/versoes/{versao_id}")
def obter_versao(versao_id: str, user_id: Optional[str] = Depends(get_user_id_from_session)):
    if not check_module_permission(user_id or '', MODULE_ID):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    conn = get_db_connection(); cur = conn.cursor()
    try:
        cur.execute("""
            SELECT id, created_at, created_by_name, hoje, elapsed_seconds,
                   totais, plano, pedidos_completos, detalhe_alocacao
            FROM plano_producao_versoes WHERE id = %s
        """, (versao_id,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Versao nao encontrada.")
        def _j(v): return v if isinstance(v, (dict, list)) else (json.loads(v) if v else [])
        return {
            'id': str(row[0]),
            'created_at': (row[1].replace(tzinfo=timezone.utc).isoformat() if row[1] else None),
            'created_by_name': row[2],
            'hoje': row[3].strftime('%Y-%m-%d') if row[3] else None,
            'elapsed_seconds': row[4],
            'totais': _j(row[5]),
            'plano': _j(row[6]),
            'pedidos_completos': _j(row[7]),
            'detalhe_alocacao': _j(row[8]),
        }
    finally:
        cur.close(); conn.close()


# =============================================================================
# OTIMIZADOR DE FATURAMENTO — visualização baseada num modelo salvo de produção
# =============================================================================
def _pedidos_completos(totais: dict, plano: list) -> set:
    """Pedidos que o plano marca como COMPLETOS (coluna 'pedidos completos'):
    completos só com estoque ∪ completos após cada item produzido."""
    s = set(totais.get('pedidos_completos_estoque') or [])
    for row in (plano or []):
        s.update((row.get('VALOR_POR_PEDIDO_COMPLETO') or {}).keys())
    return s


@router.get("/fabrica/otimizador-faturamento")
def otimizador_faturamento(versao_id: Optional[str] = None, refresh: bool = False,
                           user_id: Optional[str] = Depends(get_user_id_from_session)):
    """Faturamento dos pedidos COMPLETOS no modelo (default: última versão).
    Ordem = sequência em que o plano completa os pedidos. Saldos AO VIVO do BigQuery;
    % por ALOCAÇÃO sequencial do disponível (mesma ordem), descontando entre pedidos.
    Resultado fica em CACHE: só recalcula (BigQuery) quando refresh=True."""
    if not check_module_permission(user_id or '', MODULE_FAT):
        raise HTTPException(status_code=403, detail="Acesso negado.")

    conn = get_db_connection(); cur = conn.cursor()
    try:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS otimizador_faturamento_cache (
                versao_id UUID PRIMARY KEY,
                computed_at TIMESTAMP DEFAULT NOW(),
                payload JSONB NOT NULL
            )""")
        conn.commit()
        if versao_id:
            cur.execute("""SELECT id, created_at, created_by_name, hoje, totais, plano
                             FROM plano_producao_versoes WHERE id = %s""", (versao_id,))
        else:
            cur.execute("""SELECT id, created_at, created_by_name, hoje, totais, plano
                             FROM plano_producao_versoes ORDER BY created_at DESC LIMIT 1""")
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Nenhum plano de produção no histórico.")
        vid = str(row[0])
        if not refresh:
            cur.execute("SELECT payload, computed_at FROM otimizador_faturamento_cache WHERE versao_id = %s", (vid,))
            cached = cur.fetchone()
            if cached:
                payload = cached[0] if isinstance(cached[0], dict) else json.loads(cached[0])
                payload['cached'] = True
                payload['computed_at'] = cached[1].replace(tzinfo=timezone.utc).isoformat() if cached[1] else None
                return payload
    finally:
        cur.close(); conn.close()

    def _j(v): return v if isinstance(v, (dict, list)) else (json.loads(v) if v else {})
    totais, plano = _j(row[4]), _j(row[5])

    # ORDEM EXATA = coluna "Pedidos Completos APÓS" do plano, lida por sequência de produção
    # (seq 1, depois 2, …) e dentro de cada passo na mesma ordem da grade (alfabética).
    # NÃO inclui os 'completos só com estoque' (não estão nessa coluna).
    ordem_seq = []
    for r in sorted(plano, key=lambda x: x.get('SEQUENCIA') or 0):
        for p in sorted((r.get('VALOR_POR_PEDIDO_COMPLETO') or {}).keys()):
            ordem_seq.append(str(p))
    completos = set(ordem_seq)

    # Mapa de produção por SKU (COD normalizado): seq na fila + qtd a produzir
    produzindo = {}
    for r in plano:
        cod = str(r.get('CODIGO_PRODUTO'))
        if (r.get('QTD_PRODUZIR') or 0) > 0:
            produzindo[cod] = {'seq': r.get('SEQUENCIA'), 'qtd': int(r.get('QTD_PRODUZIR') or 0)}

    # Dados AO VIVO do BigQuery
    est, ped = carregar_dados_bq()
    ped = ped[ped['PEDIDO'].isin(completos)]
    disp = est.set_index('COD')['QUANTIDADE_DISPONIVEL'].to_dict()
    res = est.set_index('COD')['QUANTIDADE_RESERVA'].to_dict() if 'QUANTIDADE_RESERVA' in est.columns else {}

    def _fmt(dt):
        return dt.strftime('%d/%m/%Y') if pd.notna(dt) else ''

    # Pré-agrupa por pedido
    atr_map = totais.get('pedidos_atrasados', {}) or {}
    val_map = totais.get('pedidos_valores', {}) or {}
    grupos = {}
    for pid, g in ped.groupby('PEDIDO'):
        prod = g.groupby(['COD', 'CODIGO_PRODUTO'], as_index=False).agg(
            QUANTIDADE=('QUANTIDADE', 'sum'), TOTAL=('TOTAL_ITEM', 'sum'),
            DESCRICAO=('DESCRICAO_PRODUTO', 'first'))
        grupos[str(pid)] = {
            'razao': str(g['RAZAO'].iloc[0]) if 'RAZAO' in g.columns else '',
            'emissao': _fmt(g['DATA'].min()), 'entrega': _fmt(g['DATA_ENTREGA'].min()),
            'prod': prod,
        }

    # EXIBIÇÃO segue a fila da coluna "completos" do plano (ordem_seq).
    # Cobertura/"em produção" usam o disponível CHEIO do SKU por pedido (SEM desconto entre
    # pedidos) — igual ao Otimizador de Produção, para os dois baterem.
    pedidos_out = []
    skus_vistos = set()
    tot_qtd = tot_val = tot_cob = 0.0
    tot_fis = tot_res = tot_disp = 0.0
    posicao = 0
    for pid in ordem_seq:
        if pid not in grupos:
            continue
        posicao += 1
        info = grupos[pid]
        produtos, p_val, p_cob = [], 0.0, 0.0
        for r in info['prod'].itertuples():
            cod = r.COD
            d = float(disp.get(cod, 0)); rs = float(res.get(cod, 0))
            qtd = float(r.QUANTIDADE); total = float(r.TOTAL)
            # cobertura pelo disponível CHEIO do SKU (por pedido, sem desconto entre pedidos)
            usado = min(qtd, max(0.0, d))
            cobertura = (usado / qtd) if qtd > 0 else 0.0
            p_val += total; p_cob += total * cobertura
            prodinfo = produzindo.get(cod)
            # "Em produção" só se ESTE pedido depende da produção (demanda dele > disponível do SKU)
            precisa_prod = bool(prodinfo) and qtd > d
            produtos.append({
                'codigo': str(r.CODIGO_PRODUTO), 'descricao': str(r.DESCRICAO or ''),
                'quantidade': int(qtd), 'total': round(total, 2),
                'saldo_fisico': int(d + rs), 'reserva': int(rs), 'saldo_disponivel': int(d),
                'coberto_agora': int(usado),
                'em_producao': precisa_prod,
                'seq_producao': (prodinfo['seq'] if precisa_prod else None),
                'qtd_producao': (prodinfo['qtd'] if precisa_prod else None),
            })
            # KPIs de estoque: soma por SKU distinto (evita dupla contagem)
            if cod not in skus_vistos:
                skus_vistos.add(cod)
                tot_fis += d + rs; tot_res += rs; tot_disp += d
        pct = round(p_cob / p_val * 100, 2) if p_val > 0 else 0.0
        tot_qtd += sum(p['quantidade'] for p in produtos); tot_val += p_val; tot_cob += p_cob
        pedidos_out.append({
            'seq': posicao,  # contador 1,2,3… na ordem da coluna completos do plano
            'pedido': pid, 'razao': info['razao'], 'emissao': info['emissao'], 'entrega': info['entrega'],
            'total_pedido': round(p_val, 2), 'pct_completo': pct,
            'atrasado': bool(atr_map.get(pid)), 'produtos': produtos,
        })

    # pedidos_out já está na ordem exata da coluna "completos" do plano (não re-ordenar).

    result = {
        'versao': {
            'id': vid,
            'created_at': (row[1].replace(tzinfo=timezone.utc).isoformat() if row[1] else None),
            'created_by_name': row[2],
            'hoje': row[3].strftime('%Y-%m-%d') if row[3] else None,
        },
        'pedidos': pedidos_out,
        'kpis': {
            'total_faturar': round(tot_val, 2),
            'total_fisico': int(tot_fis),
            'total_reserva': int(tot_res),
            'total_disponivel': int(tot_disp),
            'quantidade': int(tot_qtd),
            'total_pedidos': len(pedidos_out),
            'qtd_skus': len(skus_vistos),
        },
        'total_geral': {
            'quantidade': int(tot_qtd), 'total': round(tot_val, 2),
            'pct_completo': round(tot_cob / tot_val * 100, 2) if tot_val > 0 else 0.0,
        },
    }

    # Persiste no cache (recalculado só quando refresh=True)
    conn = get_db_connection(); cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO otimizador_faturamento_cache (versao_id, computed_at, payload)
            VALUES (%s, NOW(), %s::jsonb)
            ON CONFLICT (versao_id) DO UPDATE SET computed_at = NOW(), payload = EXCLUDED.payload
            RETURNING computed_at
        """, (vid, json.dumps(result)))
        ca = cur.fetchone()[0]
        conn.commit()
    except Exception as e:
        conn.rollback(); logger.warning(f"cache faturamento falhou: {e}"); ca = None
    finally:
        cur.close(); conn.close()

    result['cached'] = False
    result['computed_at'] = ca.replace(tzinfo=timezone.utc).isoformat() if ca else None
    return result


@router.get("/fabrica/plano-producao/versoes/{versao_id}/xlsx")
def baixar_xlsx(versao_id: str, user_id: Optional[str] = Depends(get_user_id_from_session)):
    if not check_module_permission(user_id or '', MODULE_ID):
        raise HTTPException(status_code=403, detail="Acesso negado.")
    conn = get_db_connection(); cur = conn.cursor()
    try:
        cur.execute("""
            SELECT hoje, totais, plano, pedidos_completos, detalhe_alocacao
            FROM plano_producao_versoes WHERE id = %s
        """, (versao_id,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Versao nao encontrada.")
        def _j(v): return v if isinstance(v, (dict, list)) else (json.loads(v) if v else [])
        hoje_str = row[0].strftime('%Y-%m-%d') if row[0] else ''
        xlsx_bytes = gerar_xlsx_bytes(_j(row[2]), _j(row[1]), hoje_str, _j(row[3]), _j(row[4]))
    finally:
        cur.close(); conn.close()

    filename = f"Otimizador_Producao_{hoje_str}_{versao_id[:8]}.xlsx"
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# =============================================================================
# ENVIO WHATSAPP (WAHA)
# =============================================================================
def _fmt_moeda(v) -> str:
    try:
        s = f"{float(v):,.2f}"
        return "R$ " + s.replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "R$ 0,00"


def _html_escape(s) -> str:
    if s is None:
        return ""
    return (str(s)
            .replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            .replace('"', "&quot;").replace("'", "&#39;"))


def _gerar_html_plano(hoje_str: str, totais: dict, plano: list,
                      pedidos_completos: list, detalhe_alocacao: list) -> str:
    """Gera HTML auto-contido (Bootstrap CDN) com resumo + 3 tabelas."""
    def _table(rows: list, max_rows: int = 200) -> str:
        if not rows:
            return "<p class='text-muted'>Sem dados.</p>"
        cols = list(rows[0].keys())
        ths = "".join(f"<th>{_html_escape(c)}</th>" for c in cols)
        body_rows = []
        for r in rows[:max_rows]:
            tds = "".join(f"<td>{_html_escape(r.get(c, ''))}</td>" for c in cols)
            body_rows.append(f"<tr>{tds}</tr>")
        extra = ""
        if len(rows) > max_rows:
            extra = (f"<tr><td colspan='{len(cols)}' class='text-center text-muted'>"
                     f"... +{len(rows)-max_rows} linhas omitidas</td></tr>")
        return (f"<div class='table-responsive'><table class='table table-sm table-striped'>"
                f"<thead><tr>{ths}</tr></thead><tbody>{''.join(body_rows)}{extra}</tbody></table></div>")

    kpis = [
        ("Pedidos", totais.get('n_pedidos', 0)),
        ("Atendidos", totais.get('n_atendidos', 0)),
        ("Pedidos Atrasados", totais.get('n_atr_tot', 0)),
        ("Atrasados Atendidos", totais.get('n_atr_atend', 0)),
        ("Valor Total", _fmt_moeda(totais.get('val_tot', 0))),
        ("Valor Atendido", _fmt_moeda(totais.get('val_atend', 0))),
        ("SKUs a Produzir", totais.get('n_skus', 0)),
        ("Unidades a Produzir", totais.get('n_unid', 0)),
    ]
    kpis_html = "".join(
        f"<div class='col-6 col-md-3 mb-3'><div class='card'><div class='card-body p-2'>"
        f"<div class='text-muted small'>{_html_escape(lbl)}</div>"
        f"<div class='fs-5 fw-bold'>{_html_escape(val)}</div></div></div></div>"
        for lbl, val in kpis
    )

    return f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Otimizador de Producao {hoje_str}</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
</head>
<body class="bg-light">
<div class="container py-4">
  <h1 class="h3 mb-1">Otimizador de Producao</h1>
  <p class="text-muted">Data base: {_html_escape(hoje_str)}</p>
  <div class="row">{kpis_html}</div>
  <h2 class="h5 mt-4">Otimizador de Producao</h2>
  {_table(plano, max_rows=500)}
  <h2 class="h5 mt-4">Pedidos Completos (Estoque)</h2>
  {_table(pedidos_completos, max_rows=300)}
  <h2 class="h5 mt-4">Detalhe Alocacao Estoque</h2>
  {_table(detalhe_alocacao, max_rows=500)}
  <p class="text-muted small mt-4">Gerado pelo Portal EMPRESA.</p>
</div>
</body>
</html>"""


def _gerar_caption_plano(hoje_str: str, totais: dict) -> str:
    return (
        f"*Otimizador de Producao* — {hoje_str}\n"
        f"Pedidos: {totais.get('n_pedidos', 0)} | Atendidos: {totais.get('n_atendidos', 0)}\n"
        f"Atrasados: {totais.get('n_atr_tot', 0)} (atendidos: {totais.get('n_atr_atend', 0)})\n"
        f"Valor total: {_fmt_moeda(totais.get('val_tot', 0))}\n"
        f"Valor atendido: {_fmt_moeda(totais.get('val_atend', 0))}\n"
        f"SKUs a produzir: {totais.get('n_skus', 0)} | Unidades: {totais.get('n_unid', 0)}"
    )


class EnviarWhatsAppBody(BaseModel):
    numero: str  # so digitos (10 ou 11)


@router.post("/fabrica/plano-producao/versoes/{versao_id}/enviar-whatsapp")
def enviar_whatsapp(versao_id: str, body: EnviarWhatsAppBody,
                    user_id: Optional[str] = Depends(get_user_id_from_session)):
    # Permissao do modulo plano_producao (can_edit gate p/ disparo)
    if not check_module_permission(user_id or '', MODULE_ID, 'can_edit'):
        raise HTTPException(status_code=403, detail="Acesso negado.")

    conn = get_db_connection(); cur = conn.cursor()
    try:
        cur.execute("""
            SELECT hoje, totais, plano, pedidos_completos, detalhe_alocacao
            FROM plano_producao_versoes WHERE id = %s
        """, (versao_id,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Versao nao encontrada.")
        def _j(v): return v if isinstance(v, (dict, list)) else (json.loads(v) if v else [])
        hoje_str = row[0].strftime('%Y-%m-%d') if row[0] else ''
        totais = _j(row[1]) or {}
        plano = _j(row[2]) or []
        pedidos_completos = _j(row[3]) or []
        detalhe_alocacao = _j(row[4]) or []
    finally:
        cur.close(); conn.close()

    xlsx_bytes = gerar_xlsx_bytes(plano, totais, hoje_str, pedidos_completos, detalhe_alocacao)
    data_b64 = base64.b64encode(xlsx_bytes).decode('ascii')
    caption = _gerar_caption_plano(hoje_str, totais)
    filename = f"Otimizador_Producao_{hoje_str}_{versao_id[:8]}.xlsx"

    from modulo.whatsapp_config import enviar_arquivo_whatsapp
    result = enviar_arquivo_whatsapp(
        user_id=user_id,
        numero=body.numero,
        origem='plano_producao',
        referencia_id=versao_id,
        caption=caption,
        filename=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        data_base64=data_b64,
    )
    return result
